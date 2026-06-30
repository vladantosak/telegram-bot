import json
import os
import sqlite3
import asyncio
import logging
import csv
import io
from logging.handlers import RotatingFileHandler
from datetime import datetime
import datetime as dt_module
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
import hashlib
import re
import html

try:
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError:
    gspread = None
    Credentials = None

from groq import Groq
from telegram import (
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ForceReply,
)
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
)

# ── Логирование ──────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(),
        RotatingFileHandler("bot.log", maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

def split_message(text: str, max_len: int = 4000) -> list[str]:
    """Разбивает длинное сообщение на части, не ломая предложения."""
    chunks = []
    while len(text) > max_len:
        split_at = text.rfind('\n', 0, max_len)
        if split_at == -1:
            split_at = max_len
        chunks.append(text[:split_at])
        text = text[split_at:].lstrip()
    chunks.append(text)
    return chunks


# ── Настройки и переменные окружения ─────────────────────────────────────────
TOKEN = os.environ.get("TELEGRAM_TOKEN")

# Чтение ADMIN_IDS из переменных окружения
ADMIN_IDS_RAW = os.environ.get("ADMIN_IDS", "")
ADMIN_IDS = []
if ADMIN_IDS_RAW:
    for x in ADMIN_IDS_RAW.split(","):
        x = x.strip()
        if x.replace("-", "").isdigit():
            ADMIN_IDS.append(int(x))

DB_PATH = os.environ.get("DB_PATH", "workers.db")
DEFAULT_GROUP_ID = int(os.environ.get("GROUP_ID", "-1003804380536"))
SUMMARY_CHAT_ID = int(os.environ.get("SUMMARY_CHAT_ID", "0")) or (ADMIN_IDS[0] if ADMIN_IDS else 0)
GROQ_API_KEY = os.environ.get("GROQ_API_KEY")

# ── Часовой пояс ──────────────────────────────────────────────────────────────
LOCAL_TZ = ZoneInfo("Europe/Chisinau")  # UTC+2 / UTC+3 (летнее время)

def now_local() -> datetime:
    """Текущее время в локальном часовом поясе."""
    return datetime.now(LOCAL_TZ)

groq_client = Groq(api_key=GROQ_API_KEY) if GROQ_API_KEY else None

LATE_THRESHOLD_MIN = 15
SCHEDULE_A = ["10:00", "12:00", "15:00", "17:00"]
SCHEDULE_B = ["11:00", "13:00", "16:00", "18:00"]
SCHEDULES = {"A": SCHEDULE_A, "B": SCHEDULE_B}

# Состояния диалогов
(
    ASK_WORKER_ID,
    ASK_LASTNAME,
    ASK_FIRSTNAME,
    ASK_POSITION,
    ASK_GROUP,
    ASK_SCHEDULE,
    ASK_NEEDS_DAILY_FACT,
    ASK_REMOVE_DEPARTMENT,
    ASK_REMOVE_WORKER,
    ASK_DEPARTMENT,
    ASK_REPORT_TIME,
    ASK_LIST_DEPARTMENT,
    ASK_LIST_WORKER,
    ASK_EDIT_FIELD,
    ASK_EDIT_VALUE,
    ASK_EDIT_SCHEDULE,
    ASK_EDIT_DAILY_FACT,
    ASK_EDIT_GROUP_VALUE,
    ASK_ORDER_DEPARTMENT,
    ASK_EDIT_SORT_ORDER,
    ASK_CONFIRM_DELETE,
    ASK_EDIT_STATUS_WORK,
    ASK_EXPORT_TYPE,
    ASK_EXPORT_DEPARTMENT,
    ASK_EXPORT_FORMAT,
    ASK_GSHEETS_URL,
    ASK_GSHEETS_CREDS,
    ASK_IMPORT_FILE,
    ASK_MOVE_POSITION_ORDER,
    ASK_SUMMARY_DATE,
    ASK_EDIT_SCHEDULE_DEPT,
    ASK_DEPT_SCHEDULE_VAL,
    ASK_REG_LAST_NAME,
    ASK_REG_FIRST_NAME,
) = range(34)

# Клавиатуры
MAIN_MENU = ReplyKeyboardMarkup(
    [
        ["📋 Сотрудники", "📊 Сводка сейчас", "📅 Сводка за дату"],
        ["➕ Добавить сотрудника", "➖ Удалить сотрудника"],
        ["🏢 Сотрудники отдела", "🔄 График отдела"],
        ["⏰ Время сводки", "🆔 ID чата", "📣 Напомнить всем"],
        ["📥 Выгрузить отчеты", "📥 Импорт сотрудников"],
    ],
    resize_keyboard=True,
)

CANCEL_KEYBOARD = ReplyKeyboardMarkup([["❌ Отмена"]], resize_keyboard=True)
SCHEDULE_KEYBOARD = ReplyKeyboardMarkup([["A", "B"], ["❌ Отмена"]], resize_keyboard=True)
YES_NO_KEYBOARD = ReplyKeyboardMarkup([["Да", "Нет"], ["❌ Отмена"]], resize_keyboard=True)
CANCEL_TEXT = "❌ Отмена"
DIALOG_TEXT = filters.TEXT & ~filters.COMMAND & ~filters.Regex(f"^{CANCEL_TEXT}$")


# ══════════════════════════════════════════════════════════════════════════════
# База данных
# ══════════════════════════════════════════════════════════════════════════════

import sqlite3
import asyncio

# Глобальные асинхронные блокировки пользователей для предотвращения race condition в отчетах
user_locks = {}

def get_user_lock(user_id: int) -> asyncio.Lock:
    if user_id not in user_locks:
        user_locks[user_id] = asyncio.Lock()
    return user_locks[user_id]

def is_quiet_mode_enabled() -> bool:
    try:
        conn = get_db()
        row = conn.execute("SELECT value FROM settings WHERE key = 'quiet_mode_enabled'").fetchone()
        conn.close()
        if row:
            return row["value"] == "1"
    except Exception:
        pass
    return False

def set_quiet_mode(enabled: bool):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES ('quiet_mode_enabled', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        ("1" if enabled else "0",)
    )
    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
    except Exception:
        pass
    return conn


async def run_db(func, *args, **kwargs):
    """Выполняет блокирующую функцию работы с БД в отдельном потоке.

    Любая существующая синхронная функция (get_worker, save_report, calculate_worker_stats и т.д.)
    использует sqlite3 синхронно. Вызванная напрямую из async-хендлера, она блокирует event loop
    на время диска I/O — при нескольких пользователях бот начинает "подвисать" на сообщениях.

    Использование: вместо
        worker = get_worker(telegram_id)
    в async-функции пишем:
        worker = await run_db(get_worker, telegram_id)

    Это не требует переписывать сами DB-функции — только точки их вызова из async-хендлеров.
    """
    return await asyncio.to_thread(func, *args, **kwargs)

def init_db():
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS workers (
            telegram_id INTEGER PRIMARY KEY,
            last_name TEXT NOT NULL,
            first_name TEXT NOT NULL,
            position TEXT NOT NULL DEFAULT 'Не указано',
            group_id INTEGER NOT NULL,
            schedule TEXT NOT NULL DEFAULT 'A',
            needs_daily_fact INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1
        )
        """
    )

    cols = {row["name"] for row in conn.execute("PRAGMA table_info(workers)").fetchall()}
    for col, definition in [
        ("schedule", "TEXT NOT NULL DEFAULT 'A'"),
        ("needs_daily_fact", "INTEGER NOT NULL DEFAULT 1"),
        ("sort_order", "INTEGER NOT NULL DEFAULT 0"),
        ("is_active", "INTEGER NOT NULL DEFAULT 1"),
        ("object_id", "TEXT NOT NULL DEFAULT 'Основной'"),
    ]:
        if col not in cols:
            conn.execute(f"ALTER TABLE workers ADD COLUMN {col} {definition}")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS objects (
            object_id TEXT PRIMARY KEY,
            group_id INTEGER NOT NULL DEFAULT 0
        )
        """
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS groups (
            group_id INTEGER PRIMARY KEY,
            group_name TEXT NOT NULL
        )
        """
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER NOT NULL,
            report_date TEXT NOT NULL,
            report_type TEXT NOT NULL,
            slot_time TEXT,
            received_at TEXT NOT NULL,
            is_ok INTEGER NOT NULL,
            is_late INTEGER NOT NULL DEFAULT 0,
            format_comment TEXT,
            required_action TEXT,
            raw_text TEXT NOT NULL DEFAULT ''
        )
        """
    )

    cols_reports = {row["name"] for row in conn.execute("PRAGMA table_info(reports)").fetchall()}
    if "raw_text" not in cols_reports:
        conn.execute("ALTER TABLE reports ADD COLUMN raw_text TEXT NOT NULL DEFAULT ''")
    # Колонки для отслеживания текстового сообщения-оценки в группе: при дополнении статуса
    # старое текстовое сообщение удаляется и пересоздаётся, а пересланные медиа (видео/голос)
    # остаются как есть и не трогаются.
    if "group_chat_id" not in cols_reports:
        conn.execute("ALTER TABLE reports ADD COLUMN group_chat_id INTEGER")
    if "group_message_id" not in cols_reports:
        conn.execute("ALTER TABLE reports ADD COLUMN group_message_id INTEGER")

    # Таблица для отслеживания каждого пересланного в группу видео/голосового по отчету.
    # Нужна, чтобы при дополнении статуса В ПРЕДЕЛАХ временного окна (см. MEDIA_MERGE_WINDOW_MINUTES)
    # можно было удалить СТАРЫЕ видео из группы и переслать заново уже ВСЕ видео этого статуса вместе,
    # с одним общим комментарием под ними.
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS report_media (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER NOT NULL,
            source_chat_id INTEGER NOT NULL,
            source_message_id INTEGER NOT NULL,
            group_message_id INTEGER,
            position INTEGER NOT NULL,
            added_at TEXT NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_report_media_report ON report_media(report_id)")

    # Таблица для настроек (для сохранения расписания сводок)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )

    # Таблица для зафиксированных отправленных напоминаний
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sent_reminders (
            telegram_id INTEGER,
            report_date TEXT,
            slot_time TEXT,
            PRIMARY KEY (telegram_id, report_date, slot_time)
        )
        """
    )

    # Таблица для зафиксированных отправленных предварительных напоминаний (за 10 минут)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sent_pre_reminders (
            telegram_id INTEGER,
            report_date TEXT,
            slot_time TEXT,
            PRIMARY KEY (telegram_id, report_date, slot_time)
        )
        """
    )

    # Таблица для сохранения данных незарегистрированных сотрудников
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS pending_unregistered_users (
            telegram_id INTEGER PRIMARY KEY,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            username TEXT NOT NULL,
            timestamp TEXT NOT NULL,
            text_content TEXT NOT NULL
        )
        """
    )

    # Индексы
    conn.execute("CREATE INDEX IF NOT EXISTS idx_reports_date ON reports(report_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_reports_worker_date ON reports(telegram_id, report_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_workers_pos ON workers(position)")

    # Очистка тестовых или отмененных данных сотрудника ("Отмена")
    conn.execute(
        "DELETE FROM workers WHERE last_name LIKE '%Отмена%' OR first_name LIKE '%Отмена%' OR position LIKE '%Отмена%'"
    )
    
    # Очистка старых напоминаний старше 7 дней
    conn.execute("DELETE FROM sent_reminders WHERE report_date < date('now', '-7 days')")
    conn.execute("DELETE FROM sent_pre_reminders WHERE report_date < date('now', '-7 days')")

    conn.commit()
    conn.close()

def get_worker(telegram_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (telegram_id,)).fetchone()
    conn.close()
    return row

def find_unregistered_workers_by_lastname(last_name: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM workers WHERE telegram_id < 0 AND LOWER(last_name) = LOWER(?)",
        (last_name.strip(),)
    ).fetchall()
    conn.close()
    return rows

def bind_worker_id(old_id: int, new_id: int):
    conn = get_db()
    try:
        conn.execute("UPDATE workers SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE reports SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE sent_reminders SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE sent_pre_reminders SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.commit()
    except Exception as e:
        logger.error(f"Error binding worker ID: {e}")
        conn.rollback()
        raise e
    finally:
        conn.close()

def save_pending_unregistered_user(telegram_id: int, first_name: str, last_name: str, username: str, timestamp: str, text_content: str):
    conn = get_db()
    conn.execute(
        """
        INSERT OR REPLACE INTO pending_unregistered_users (telegram_id, first_name, last_name, username, timestamp, text_content)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (telegram_id, first_name, last_name, username, timestamp, text_content)
    )
    conn.commit()
    conn.close()

def get_pending_unregistered_user(telegram_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM pending_unregistered_users WHERE telegram_id = ?", (telegram_id,)).fetchone()
    conn.close()
    return row

def delete_pending_unregistered_user(telegram_id: int):
    conn = get_db()
    conn.execute("DELETE FROM pending_unregistered_users WHERE telegram_id = ?", (telegram_id,))
    conn.commit()
    conn.close()

def get_all_workers():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM workers ORDER BY position, sort_order, last_name, first_name"
    ).fetchall()
    conn.close()
    return rows

def get_workers_by_position(position: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM workers WHERE lower(position) = lower(?) ORDER BY sort_order, last_name, first_name",
        (position,),
    ).fetchall()
    conn.close()
    return rows

def upsert_worker(telegram_id: int, last_name: str, first_name: str, position: str, group_id: int, schedule: str, needs_daily_fact: bool, sort_order: int = 0, is_active: int = 1, object_id: str = 'Основной'):
    conn = get_db()
    conn.execute(
        """
        INSERT INTO workers (telegram_id, last_name, first_name, position, group_id, schedule, needs_daily_fact, sort_order, is_active, object_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(telegram_id) DO UPDATE SET
            last_name=excluded.last_name,
            first_name=excluded.first_name,
            position=excluded.position,
            group_id=excluded.group_id,
            schedule=excluded.schedule,
            needs_daily_fact=excluded.needs_daily_fact,
            sort_order=excluded.sort_order,
            is_active=excluded.is_active,
            object_id=excluded.object_id
        """,
        (telegram_id, last_name, first_name, position, group_id, schedule, int(needs_daily_fact), sort_order, is_active, object_id),
    )
    conn.commit()
    conn.close()
    try:
        loop = asyncio.get_running_loop()
        if loop.is_running():
            loop.create_task(async_sync_gsheets_background())
    except RuntimeError:
        pass

def get_object_group(object_id: str) -> int:
    conn = get_db()
    row = conn.execute("SELECT group_id FROM objects WHERE object_id = ?", (object_id,)).fetchone()
    conn.close()
    if row:
        return row["group_id"]
    return 0

def save_object_group(object_id: str, group_id: int):
    conn = get_db()
    conn.execute(
        """
        INSERT INTO objects (object_id, group_id) VALUES (?, ?)
        ON CONFLICT(object_id) DO UPDATE SET group_id=excluded.group_id
        """,
        (object_id, group_id)
    )
    conn.commit()
    conn.close()

def get_worker_target_group(worker) -> int:
    try:
        w_dict = dict(worker)
    except (TypeError, ValueError):
        return DEFAULT_GROUP_ID
    if "object_id" in w_dict and w_dict["object_id"]:
        obj_group = get_object_group(w_dict["object_id"])
        if obj_group and obj_group != 0:
            return obj_group
    return w_dict.get("group_id") or DEFAULT_GROUP_ID

def read_excel(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    headers = []
    for cell in sheet[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip().lower())
        else:
            headers.append("")
            
    workers = []
    
    idx_id = -1
    idx_lastname = -1
    idx_firstname = -1
    idx_name = -1
    idx_position = -1
    idx_group = -1
    idx_schedule = -1
    idx_daily_fact = -1
    idx_object = -1
    
    for idx, h in enumerate(headers):
        if h in ("telegram_id", "id", "tg id", "telegram", "телеграм id", "тг id"):
            idx_id = idx
        elif h in ("last_name", "lastname", "фамилия"):
            idx_lastname = idx
        elif h in ("first_name", "firstname", "имя"):
            idx_firstname = idx
        elif h in ("name", "фио", "сотрудник", "имя фамилия"):
            idx_name = idx
        elif h in ("position", "должность", "отдел", "специальность"):
            idx_position = idx
        elif h in ("group_id", "группа", "id группы", "чат", "чат id", "brigade"):
            idx_group = idx
        elif h in ("schedule", "график", "расписание"):
            idx_schedule = idx
        elif h in ("needs_daily_fact", "факт дня", "ежедневный факт"):
            idx_daily_fact = idx
        elif h in ("object", "объект"):
            idx_object = idx
            
    for row_idx in range(2, sheet.max_row + 1):
        row_values = []
        for col_idx in range(1, len(headers) + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            row_values.append(val)
        
        if not any(v is not None for v in row_values):
            continue
            
        tg_id_val = row_values[idx_id] if idx_id != -1 and idx_id < len(row_values) else None
        lastname_val = row_values[idx_lastname] if idx_lastname != -1 and idx_lastname < len(row_values) else None
        firstname_val = row_values[idx_firstname] if idx_firstname != -1 and idx_firstname < len(row_values) else None
        name_val = row_values[idx_name] if idx_name != -1 and idx_name < len(row_values) else None
        pos_val = row_values[idx_position] if idx_position != -1 and idx_position < len(row_values) else None
        group_val = row_values[idx_group] if idx_group != -1 and idx_group < len(row_values) else None
        schedule_val = row_values[idx_schedule] if idx_schedule != -1 and idx_schedule < len(row_values) else None
        daily_fact_val = row_values[idx_daily_fact] if idx_daily_fact != -1 and idx_daily_fact < len(row_values) else None
        
        # We also check if user has name/position/brigade/object columns which we map gracefully:
        # object column could be mapped to position/department if not specified
        object_id = "Основной"
        if idx_object != -1 and idx_object < len(row_values) and row_values[idx_object] is not None:
            obj_val = str(row_values[idx_object]).strip()
            object_id = obj_val
            if pos_val is None:
                pos_val = obj_val
            else:
                pos_val = f"{str(pos_val).strip()} ({obj_val})"

        tg_id = None
        if tg_id_val is not None:
            try:
                tg_id = int(float(str(tg_id_val).strip()))
            except ValueError:
                pass
                
        last_name = ""
        first_name = ""
        if lastname_val is not None:
            last_name = str(lastname_val).strip()
        if firstname_val is not None:
            first_name = str(firstname_val).strip()
            
        if not last_name and not first_name and name_val is not None:
            parts = str(name_val).strip().split()
            if len(parts) >= 2:
                last_name = parts[0]
                first_name = " ".join(parts[1:])
            elif len(parts) == 1:
                first_name = parts[0]
                last_name = "Сотрудник"
                
        if not last_name:
            last_name = "Не указана"
        if not first_name:
            first_name = "Не указано"
            
        position = str(pos_val).strip() if pos_val is not None else "Не указано"
        
        group_id = DEFAULT_GROUP_ID
        if group_val is not None:
            try:
                # If they wrote e.g. "Бригада 1", it isn't an integer, so we ignore or map to DEFAULT_GROUP_ID
                group_id = int(float(str(group_val).strip()))
            except ValueError:
                pass
                
        schedule = str(schedule_val).strip().upper() if schedule_val is not None else "A"
        if schedule not in ("A", "B"):
            schedule = "A"
            
        needs_daily_fact = True
        if daily_fact_val is not None:
            df_str = str(daily_fact_val).strip().lower()
            if df_str in ("0", "false", "нет", "no"):
                needs_daily_fact = False
                
        # Generate a fallback tg_id if not present
        if tg_id is None:
            name_str = f"{last_name} {first_name} {position}"
            tg_id_hash = int(hashlib.md5(name_str.encode('utf-8')).hexdigest()[:7], 16)
            tg_id = -tg_id_hash

        workers.append({
            "telegram_id": tg_id,
            "last_name": last_name,
            "first_name": first_name,
            "position": position,
            "group_id": group_id,
            "schedule": schedule,
            "needs_daily_fact": needs_daily_fact,
            "object_id": object_id
        })
        
    return workers

def update_worker_field(telegram_id: int, field: str, value):
    allowed = {"last_name", "first_name", "position", "group_id", "schedule", "needs_daily_fact", "sort_order", "is_active", "object_id"}
    if field not in allowed:
        raise ValueError(f"Недопустимое поле: {field}")
    conn = get_db()
    
    # Решение проблемы 9: Сброс sort_order в 0 при смене отдела (position)
    if field == "position":
        next_order = get_next_sort_order(value)
        conn.execute("UPDATE workers SET position = ?, sort_order = ? WHERE telegram_id = ?", (value, next_order, telegram_id))
    else:
        conn.execute(f"UPDATE workers SET {field} = ? WHERE telegram_id = ?", (value, telegram_id))
        
    conn.commit()
    conn.close()
    try:
        loop = asyncio.get_running_loop()
        if loop.is_running():
            loop.create_task(async_sync_gsheets_background())
    except RuntimeError:
        pass

def swap_sort_order(id1: int, id2: int):
    conn = get_db()
    r1 = conn.execute("SELECT sort_order FROM workers WHERE telegram_id = ?", (id1,)).fetchone()
    r2 = conn.execute("SELECT sort_order FROM workers WHERE telegram_id = ?", (id2,)).fetchone()
    if r1 and r2:
        conn.execute("UPDATE workers SET sort_order = ? WHERE telegram_id = ?", (r2["sort_order"], id1))
        conn.execute("UPDATE workers SET sort_order = ? WHERE telegram_id = ?", (r1["sort_order"], id2))
        conn.commit()
    conn.close()

def delete_worker(telegram_id: int) -> bool:
    conn = get_db()
    try:
        conn.execute("DELETE FROM reports WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM sent_reminders WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM sent_pre_reminders WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM pending_unregistered_users WHERE telegram_id = ?", (telegram_id,))
        cur = conn.execute("DELETE FROM workers WHERE telegram_id = ?", (telegram_id,))
        conn.commit()
        deleted = cur.rowcount > 0
    except Exception as e:
        logger.error(f"Error in delete_worker {telegram_id}: {e}")
        conn.rollback()
        deleted = False
    finally:
        conn.close()
    try:
        loop = asyncio.get_running_loop()
        if loop.is_running():
            loop.create_task(async_sync_gsheets_background())
    except RuntimeError:
        pass
    return deleted

def save_group_name(group_id: int, group_name: str):
    conn = get_db()
    conn.execute(
        """
        INSERT INTO groups (group_id, group_name) VALUES (?, ?)
        ON CONFLICT(group_id) DO UPDATE SET group_name=excluded.group_name
        """,
        (group_id, group_name),
    )
    conn.commit()
    conn.close()

def get_group_name(group_id: int) -> str:
    conn = get_db()
    row = conn.execute("SELECT group_name FROM groups WHERE group_id = ?", (group_id,)).fetchone()
    conn.close()
    return row["group_name"] if row else str(group_id)

async def get_group_name_async(bot, group_id: int) -> str:
    """Решение проблемы 8: автоматическое подтягивание названия некэшированной группы из Telegram."""
    conn = get_db()
    row = conn.execute("SELECT group_name FROM groups WHERE group_id = ?", (group_id,)).fetchone()
    conn.close()
    if row:
        return row["group_name"]
    try:
        chat = await bot.get_chat(group_id)
        name = chat.title or str(group_id)
        save_group_name(group_id, name)
        return name
    except Exception:
        return str(group_id)

def get_all_group_names() -> dict:
    conn = get_db()
    rows = conn.execute("SELECT group_id, group_name FROM groups").fetchall()
    conn.close()
    res = {row["group_id"]: row["group_name"] for row in rows}
    if DEFAULT_GROUP_ID not in res:
        res[DEFAULT_GROUP_ID] = str(DEFAULT_GROUP_ID)
    return res

async def fetch_and_save_group_name(bot, group_id: int) -> str:
    try:
        chat = await bot.get_chat(group_id)
        name = chat.title or str(group_id)
    except Exception:
        name = str(group_id)
    save_group_name(group_id, name)
    return name

def save_report(telegram_id: int, report_date: str, report_type: str, slot_time: str | None, received_at: str, is_ok: bool, is_late: bool, format_comment: str, required_action: str, raw_text: str = "") -> int:
    """Слегка изменена для возврата ID новой записи (для исправления оценок ИИ)."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO reports (telegram_id, report_date, report_type, slot_time, received_at, is_ok, is_late, format_comment, required_action, raw_text)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (telegram_id, report_date, report_type, slot_time, received_at, int(is_ok), int(is_late), format_comment, required_action, raw_text),
    )
    conn.commit()
    inserted_id = cur.lastrowid
    conn.close()
    try:
        loop = asyncio.get_running_loop()
        if loop.is_running():
            loop.create_task(async_sync_gsheets_background())
    except RuntimeError:
        pass
    return inserted_id

def update_report_text_and_ai(report_id: int, is_ok: bool, format_comment: str, required_action: str, raw_text: str, received_at: str):
    conn = get_db()
    conn.execute(
        """
        UPDATE reports
        SET is_ok = ?, format_comment = ?, required_action = ?, raw_text = ?, received_at = ?
        WHERE id = ?
        """,
        (int(is_ok), format_comment, required_action, raw_text, received_at, report_id)
    )
    conn.commit()
    conn.close()
    try:
        loop = asyncio.get_running_loop()
        if loop.is_running():
            loop.create_task(async_sync_gsheets_background())
    except RuntimeError:
        pass

def set_report_group_message(report_id: int, chat_id: int, message_id: int):
    """Сохраняет id текстового сообщения-оценки в группе, чтобы при следующем дополнении
    к этому же отчёту можно было удалить старое сообщение и отправить новое.
    Сообщения с пересланными медиа (видео/голос) сюда не относятся — они не трогаются."""
    conn = get_db()
    conn.execute(
        "UPDATE reports SET group_chat_id = ?, group_message_id = ? WHERE id = ?",
        (chat_id, message_id, report_id)
    )
    conn.commit()
    conn.close()

def get_report_group_message(report_id: int):
    """Возвращает (group_chat_id, group_message_id) ранее сохраненного сообщения-оценки в группе."""
    conn = get_db()
    row = conn.execute("SELECT group_chat_id, group_message_id FROM reports WHERE id = ?", (report_id,)).fetchone()
    conn.close()
    return row

def get_submitted_status_slots(telegram_id: int, report_date: str) -> set:
    """Слоты, по которым на сегодня у сотрудника уже есть отчет-статус."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT slot_time FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'status'",
        (telegram_id, report_date)
    ).fetchall()
    conn.close()
    return {r["slot_time"] for r in rows}

def pick_target_status_slot(schedule: list[str], now: datetime, submitted_slots: set):
    """Выбирает слот, к которому нужно отнести статус-отчет.

    Если есть уже прошедшие, но ещё не сданные слоты — берём САМЫЙ РАННИЙ из них.
    Это нужно, чтобы несколько видео/сообщений подряд (например, вечером, когда сотрудник
    "догоняется" по пропущенным статусам) последовательно закрывали пропуски по порядку,
    а не все скидывались в один и тот же (последний) слот.

    Если пропущенных слотов нет — обычная логика "ближайший по времени слот"
    (стандартная сдача отчета в реальном времени)."""
    current_mins = now.hour * 60 + now.minute
    missing_passed = []
    for slot in schedule:
        if slot in submitted_slots:
            continue
        h, m = map(int, slot.split(":"))
        if h * 60 + m <= current_mins:
            missing_passed.append(slot)
    if missing_passed:
        missing_passed.sort(key=lambda s: tuple(map(int, s.split(":"))))
        return missing_passed[0], True
    return find_nearest_slot(schedule, now)

def get_existing_report_row(telegram_id: int, report_date: str, report_type: str, slot_time: str | None = None):
    """Находит уже существующий отчет за слот (status) или за день (daily_fact), если он есть."""
    conn = get_db()
    if report_type == "status":
        row = conn.execute(
            "SELECT * FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'status' AND slot_time = ?",
            (telegram_id, report_date, slot_time)
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT * FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'daily_fact'",
            (telegram_id, report_date)
        ).fetchone()
    conn.close()
    return row

def build_addon_text(existing_raw: str, new_text: str, use_video_label: bool) -> str:
    """Склеивает текст дополнения со старым текстом отчета, нумеруя видео ("[Видео N]")
    или помечая текстовое дополнение ("[Дополнение]"), чтобы и в чате, и для самого ИИ
    было понятно, что это N-я по счету запись, а не разрозненный текст."""
    existing_raw = existing_raw or ""
    if use_video_label:
        idx = existing_raw.count("[Видео ") + 1
        if "[Видео " not in existing_raw and existing_raw:
            existing_raw = f"[Видео 1]: {existing_raw}"
            idx = 2
        label = f"[Видео {idx}]"
    else:
        label = "[Дополнение]"
    return f"{existing_raw}\n{label}: {new_text}" if existing_raw else new_text

def add_report_media(report_id: int, source_chat_id: int, source_message_id: int, group_message_id: int | None, position: int, added_at: str):
    """Запоминает, что для этого отчета в группу было переслано конкретное видео/голосовое —
    откуда оно взято (source_chat_id/source_message_id, чтобы переслать заново) и куда было
    переслано (group_message_id, чтобы можно было удалить, если потребуется)."""
    conn = get_db()
    conn.execute(
        "INSERT INTO report_media (report_id, source_chat_id, source_message_id, group_message_id, position, added_at) VALUES (?, ?, ?, ?, ?, ?)",
        (report_id, source_chat_id, source_message_id, group_message_id, position, added_at)
    )
    conn.commit()
    conn.close()

def get_report_media(report_id: int):
    """Все видео/голосовые, привязанные к отчету, в порядке добавления."""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM report_media WHERE report_id = ? ORDER BY position",
        (report_id,)
    ).fetchall()
    conn.close()
    return rows

def delete_report_media_rows(report_id: int):
    conn = get_db()
    conn.execute("DELETE FROM report_media WHERE report_id = ?", (report_id,))
    conn.commit()
    conn.close()

def get_worker_history_last_week(telegram_id: int) -> str:
    """Решение проблемы 6: Логирование истории отчетов сотрудника за неделю."""
    conn = get_db()
    rows = conn.execute(
        """
        SELECT * FROM reports 
        WHERE telegram_id = ? 
        AND report_date >= date('now', '-7 days')
        ORDER BY report_date DESC, received_at DESC
        """,
        (telegram_id,)
    ).fetchall()
    
    worker = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (telegram_id,)).fetchone()
    conn.close()
    
    if not worker:
        return "Сотрудник не найден."
        
    if not rows:
        return f"📅 У сотрудника {worker['last_name']} {worker['first_name']} нет отчетов за последние 7 дней."
        
    lines = [f"📅 История отчетов за прошедшую неделю для {worker['last_name']} {worker['first_name']}:\n"]
    for r in rows:
        r_type = "Статус" if r["report_type"] == "status" else "Итог дня"
        slot_str = f" за {r['slot_time']}" if r["slot_time"] and r["report_type"] == "status" else ""
        ok_str = "✅ ОК" if r["is_ok"] else f"⚠️ Замечание: {r['format_comment']}"
        late_str = " ⏰ Опоздание" if r["is_late"] else ""
        lines.append(
            f"📍 {r['report_date']} в {r['received_at']}\n"
            f"   Тип: {r_type}{slot_str}\n"
            f"   Результат: {ok_str}{late_str}\n"
        )
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# Решение проблемы 12 (Хранение расписания сводок в БД)
# ══════════════════════════════════════════════════════════════════════════════

def get_scheduled_times() -> list[str]:
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key = 'summary_times'").fetchone()
    conn.close()
    if row:
        try:
            return json.loads(row["value"])
        except Exception:
            return ["19:00"]
    return ["19:00"]

def save_scheduled_times(times: list[str]):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES ('summary_times', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (json.dumps(sorted(times)),)
    )
    conn.commit()
    conn.close()

def reschedule_summary_jobs(application: Application):
    job_queue = application.job_queue
    if not job_queue:
        logger.warning("JobQueue недоступен.")
        return

    # Удаление существующих задач сводки
    for job in job_queue.get_jobs_by_name("daily_summary"):
        job.schedule_removal()
    for job in job_queue.get_jobs_by_name("weekly_monthly_summary"):
        job.schedule_removal()

    times = get_scheduled_times()
    for t_str in times:
        try:
            hour, minute = map(int, t_str.split(":"))
            time_obj = dt_module.time(hour=hour, minute=minute, tzinfo=LOCAL_TZ)
            job_queue.run_daily(
                scheduled_summary_callback,
                time=time_obj,
                days=(0, 1, 2, 3, 4, 5, 6),
                name="daily_summary"
            )
            logger.info(f"Запланирована сводка на {t_str}")
        except Exception as e:
            logger.error(f"Критическая ошибка при планировании сводки на {t_str}: {e}")

    # Планируем еженедельный и ежемесячный отчет на 09:00 ежедневно
    try:
        time_obj_wm = dt_module.time(hour=9, minute=0, tzinfo=LOCAL_TZ)
        job_queue.run_daily(
            weekly_monthly_summary_callback,
            time=time_obj_wm,
            days=(0, 1, 2, 3, 4, 5, 6),
            name="weekly_monthly_summary"
        )
        logger.info("Запланирована ежедневная проверка на 09:00 для еженедельных/ежемесячных автоотчетов")
    except Exception as e:
        logger.error(f"Ошибка при планировании еженедельного/ежемесячного отчета: {e}")

    # Удаление существующих задач резервного копирования
    for job in job_queue.get_jobs_by_name("daily_backup"):
        job.schedule_removal()
        
    # Планируем бэкап на 03:00 ночи ежедневно
    try:
        time_obj_backup = dt_module.time(hour=3, minute=0, tzinfo=LOCAL_TZ)
        job_queue.run_daily(
            daily_backup_callback,
            time=time_obj_backup,
            days=(0, 1, 2, 3, 4, 5, 6),
            name="daily_backup"
        )
        logger.info("Запланировано ежедневное резервное копирование на 03:00")
    except Exception as e:
        logger.error(f"Ошибка при планировании бэкапа: {e}")

async def scheduled_summary_callback(context: ContextTypes.DEFAULT_TYPE):
    now = now_local()
    date_str = now.strftime("%Y-%m-%d")
    summary_text = f"⏰ Автоматическая запланированная сводка:\n\n" + generate_daily_summary_text(date_str)
    
    if SUMMARY_CHAT_ID:
        try:
            for part in split_message(summary_text):
                await context.bot.send_message(chat_id=SUMMARY_CHAT_ID, text=part)
        except Exception as e:
            logger.error(f"Ошибка при отправке автоматической сводки в {SUMMARY_CHAT_ID}: {e}")
            
    for admin_id in ADMIN_IDS:
        try:
            for part in split_message(summary_text):
                await context.bot.send_message(chat_id=admin_id, text=part)
        except Exception:
            pass

async def weekly_monthly_summary_callback(context: ContextTypes.DEFAULT_TYPE):
    now = now_local()
    # Если сегодня понедельник (weekday == 0), отправляем еженедельный отчет
    if now.weekday() == 0:
        logger.info("Отправка еженедельного отчета (понедельник)...")
        await send_weekly_summary(context.bot)
        
    # Если сегодня первое число месяца, отправляем ежемесячный отчет
    if now.day == 1:
        logger.info("Отправка ежемесячного отчета (первое число месяца)...")
        await send_monthly_summary(context.bot)


# ══════════════════════════════════════════════════════════════════════════════
# Вспомогательные функции авторизации и меню
# ══════════════════════════════════════════════════════════════════════════════

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

async def require_admin(update: Update) -> bool:
    if not is_admin(update.effective_user.id):
        await update.message.reply_text(
            "Эта кнопка доступна только администратору.",
            reply_markup=ReplyKeyboardRemove(),
        )
        return False
    if update.effective_chat.type != "private":
        try:
            await update.message.reply_text(
                "⚠️ Эта функция доступна только в личных сообщениях с ботом.",
                reply_markup=ReplyKeyboardRemove()
            )
        except Exception:
            pass
        return False
    return True

async def notify_admins_new_registration(bot, w_fio: str, position: str, username: str, user_id: int):
    w_fio_esc = html.escape(w_fio)
    position_esc = html.escape(position)
    username_esc = html.escape(username)
    admin_msg = (
        f"🎉 <b>Зарегистрировался новый сотрудник!</b>\n\n"
        f"👤 <b>ФИО:</b> {w_fio_esc}\n"
        f"💼 <b>Должность/Отдел:</b> {position_esc}\n"
        f"📱 <b>Никнейм в TG:</b> @{username_esc}\n"
        f"🆔 <b>Telegram ID:</b> <code>{user_id}</code>"
    )
    # Notify individual admin IDs
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(chat_id=admin_id, text=admin_msg, parse_mode="HTML")
        except Exception as e:
            logger.error(f"Failed to send registration notification to admin {admin_id}: {e}")
            
    # Notify the summary chat if configured
    if SUMMARY_CHAT_ID and SUMMARY_CHAT_ID not in ADMIN_IDS:
        try:
            await bot.send_message(chat_id=SUMMARY_CHAT_ID, text=admin_msg, parse_mode="HTML")
        except Exception as e:
            logger.error(f"Failed to send registration notification to SUMMARY_CHAT_ID {SUMMARY_CHAT_ID}: {e}")

async def send_report_instruction(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    chat_type = update.effective_chat.type
    await update.message.reply_text(
    "*📹 Видео-статус*\n\n"
    "Отправляйте видео в чат:\n"
    "🕙 10:00 | 12:00 | 15:00 | 17:00\n\n"

    "*Видео должно содержать:*\n"
    "✅ Что сделали за последние 2 часа\n"
    "✅ Объём работы\n"
    "✅ Показать результат на видео и обозначить его голосом\n\n"

    "*В видео скажите голосом:*\n"
    "• Что было сделано\n"
    "• Сколько сделано\n"
    "• И где\n\n"

    "Пример:\n"
    "«За 2 часа выкопал 15 метров траншеи. "
    "Показываю выполненную работу в кадре»\n\n"

    "Всем кроме механизатовов\n"
    "(если вы работаете за рулём техники вы факт не присылаете)\n\n"
    "*В 17:00 отправьте 2 видео:*\n\n"
    "1️⃣ Статус за последние 2 часа\n"
    "— что сделали сейчас\n"
    "— показать результат\n\n"

    "2️⃣ Факт за весь день\n"
    "— что сделали за день\n"
    "— общий объём\n"
    "— итоговый результат\n\n"

    "Если нельзя показать работу:\n"
    "укажите голосом:\n"
    "• что сделали\n"
    "• где сделали\n"
    "• сколько сделали\n"
    "• почему нельзя показать\n\n"

    "⚠️ Исправляйте замечания по прошлым статусам. если бам делают замчания в противном случае.\n"
    "При повторном несоблюдении требований информация передаётся руководству для принятия дальнейшего решения.\n",

parse_mode="Markdown",
    
    parse_mode="Markdown",
    reply_markup=menu_for_user(user_id, chat_type)
)

def menu_for_user(user_id: int, chat_type: str = "private"):
    if is_admin(user_id) and chat_type == "private":
        return MAIN_MENU
    if chat_type != "private":
        return ReplyKeyboardRemove()
    if get_worker(user_id) is not None:
        return ReplyKeyboardMarkup(
            [
                ["📋 Инструкция по сдаче видео-статуса"],
                ["🛌 Не работаю сегодня"]
            ],
            resize_keyboard=True
        )
    return ReplyKeyboardMarkup([["🔑 Начать регистрацию"]], resize_keyboard=True)

def positions_keyboard(rows):
    positions = sorted({row["position"] for row in rows})
    keyboard = [[p] for p in positions]
    keyboard.append(["❌ Отмена"])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True), positions

def numbered_workers_keyboard(rows):
    keyboard = []
    for i, row in enumerate(rows, 1):
        keyboard.append([f"{i}. {row['last_name']} {row['first_name']}"])
    keyboard.append(["❌ Отмена"])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def departments_reply_keyboard(context: ContextTypes.DEFAULT_TYPE | None = None) -> ReplyKeyboardMarkup:
    if context is not None and "depts_cache" in context.user_data:
        depts = context.user_data["depts_cache"]
    else:
        conn = get_db()
        rows = conn.execute("SELECT DISTINCT position FROM workers WHERE position IS NOT NULL AND position != ''").fetchall()
        conn.close()
        depts = sorted(list({row["position"] for row in rows if row["position"]}))
        if context is not None:
            context.user_data["depts_cache"] = depts
            
    keyboard = [[dept] for dept in depts if dept]
    keyboard.append(["❌ Отмена"])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_next_sort_order(position: str) -> int:
    conn = get_db()
    row = conn.execute("SELECT MAX(sort_order) as max_val FROM workers WHERE lower(position) = lower(?)", (position,)).fetchone()
    conn.close()
    if row and row["max_val"] is not None:
        return row["max_val"] + 1
    return 1


# ══════════════════════════════════════════════════════════════════════════════
# Работа с ИИ (Groq / Whisper / Llama)
# ══════════════════════════════════════════════════════════════════════════════

def normalize_ai_result(data: dict, source_text: str, report_type: str | None = "status") -> dict:
    # report_type="status"/"daily_fact" — тип жёстко задан системой по времени (днём это всегда так).
    # report_type=None — "авто"-режим: вечером система не может заранее сказать, статус это или факт,
    # поэтому доверяем выбору модели из её собственного ответа (data["report_type"]).
    if report_type not in ("status", "daily_fact"):
        report_type = str(data.get("report_type", "status")).strip().lower()
        if report_type not in ("status", "daily_fact"):
            report_type = "status"

    raw_ok = data.get("is_ok", False)
    if isinstance(raw_ok, str):
        is_ok = raw_ok.strip().lower() in ("true", "1", "yes", "да")
    else:
        is_ok = bool(raw_ok)
    issue = str(data.get("issue") or data.get("format_comment") or "").strip()
    required_action = str(data.get("required_action") or "").strip()
    employee_message = str(data.get("employee_message") or "").strip()

    if is_ok:
        issue = ""
        format_comment = "всё ОК"
        required_action = "ничего не предпринимать"
        employee_message = ""
    else:
        if not issue:
            issue = "есть замечания по отчету"
        format_comment = f"не ОК, {issue}"
        required_action = f"сделал замечание сотруднику: {issue}"
        if not employee_message:
            employee_message = f"В отчете есть замечание: {issue}. В следующем отчете исправьте это."

    return {
        "report_type": report_type,
        "is_ok": is_ok,
        "format_comment": format_comment,
        "required_action": required_action,
        "employee_message": employee_message,
    }

def find_nearest_slot(schedule: list[str], now: datetime):
    current_minutes = now.hour * 60 + now.minute
    nearest_slot = None
    nearest_diff = None

    for slot in schedule:
        hour, minute = map(int, slot.split(":"))
        slot_minutes = hour * 60 + minute
        diff = abs(current_minutes - slot_minutes)
        if nearest_diff is None or diff < nearest_diff:
            nearest_diff = diff
            nearest_slot = slot

    is_late = nearest_diff is not None and nearest_diff > LATE_THRESHOLD_MIN
    return nearest_slot, is_late

def transcribe_audio(file_path: str) -> str:
    if groq_client is None:
        return "Не задан GROQ_API_KEY, аудио не распознано."
    try:
        with open(file_path, "rb") as f:
            transcription = groq_client.audio.transcriptions.create(
                file=(os.path.basename(file_path), f),
                model="whisper-large-v3",
                language="ru",
                response_format="text",
            )
        return transcription.strip()
    except Exception as e:
        return f"Ошибка распознавания аудио: {e}"

CHECK_PROMPT_TEMPLATE = """
Ты — опытный прораб строительного объекта.

Твоя задача — проверить отчёт рабочего.
Рабочие могут писать коротко, с ошибками, простыми словами.
Ты должен понимать смысл, а не искать идеальную формулировку.

Главный вопрос:
СДЕЛАЛ ЛИ ЧЕЛОВЕК РАБОТУ ИЛИ НЕТ?

{mode_instruction}


=====================
КАК ОЦЕНИВАТЬ
=====================

Отчёт считается ХОРОШИМ, если понятно:

- какую работу выполнял человек
- с чем он работал
- какой процесс выполнялся
- есть ли проблема


Не требуй обязательно цифры.

Примеры ХОРОШИХ отчётов:

"работал на дробилке"
→ хорошо, понятно что человек выполнял работу


"стоял на кране, подавал материал"
→ хорошо


"делал опалубку"
→ хорошо


"бетон заливали сегодня"
→ хорошо


"копал траншею 50 метров"
→ отлично, есть объём


=====================
КОГДА НУЖНЫ ОБЪЁМЫ
=====================

Цифры важны там, где работа измеряется:

земляные работы:
- метры
- кубы

бетон:
- кубометры

монтаж:
- количество элементов

кабель:
- метры


Но отсутствие цифр НЕ означает плохой отчёт.

Например:

"копал траншею"

может быть принят,
если понятно, что человек реально выполнял работу.


=====================
ПЛОХИЕ ОТЧЁТЫ
=====================

Отклонять:

"работаю"
"в процессе"
"нормально"
"всё сделал"
"на объекте"
"занимаюсь"

если невозможно понять, что именно делал человек.


=====================
ИСПРАВЛЕНИЕ РЕЧИ
=====================

Рабочие могут писать:

"там это, ковыряли яму"
понимай как:
"Выполнялись земляные работы"


"дробилку гоняли"
понимай как:
"Работа на дробильном оборудовании"


Ты обязан понять смысл сообщения.


=====================
ФОРМАТ
=====================

Ответь только JSON:

{{
{json_type_field}"is_ok": true или false,

"issue": "что не так",

"required_action": "что написать сотруднику",

"employee_message": "короткое сообщение сотруднику"
}}

Отчёт:

{text}
"""

CLEAN_REPORT_PROMPT = """
Ты — технический специалист, который оформляет отчёты строительной бригады.

Тебе дают сообщение рабочего.
Рабочий может писать с ошибками, сокращениями и разговорными словами.

Твоя задача:
превратить его сообщение в понятный официальный отчёт.

Правила:

1. Не придумывай работу, которой не было.
2. Сохраняй только смысл исходного сообщения.
3. Исправляй ошибки.
4. Убирай слова-паразиты.
5. Делай текст коротким и понятным.


Примеры:


Вход:
"с утра там дробилку гоняли щебень делали"

Выход:
"Выполнялась работа на дробильном оборудовании, производилась переработка материала."


Вход:
"копал там возле склада траншею"

Выход:
"Выполнялись земляные работы: разработка траншеи возле склада."


Вход:
"бетон лили колонны"

Выход:
"Выполнялись бетонные работы: заливка колонн."


Вход:
"всё норм"

Выход:
"Отчёт не содержит информации о выполненной работе."


Верни только готовый текст отчёта.


Сообщение рабочего:

{text}
"""

_ai_status_cache = {}
_ai_clean_cache = {}

def get_md5(text: str) -> str:
    return hashlib.md5(text.strip().encode("utf-8")).hexdigest()

def clean_report(text: str) -> str:
    if groq_client is None:
        return text
    h = get_md5(text)
    if h in _ai_clean_cache:
        logger.info("clean_report: cache hit!")
        return _ai_clean_cache[h]
    try:
        response = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "Ты преобразуешь сообщения рабочих в официальные отчеты строительной бригады. Верни только готовый текст отчета без каких-либо комментариев и кавычек."},
                {"role": "user", "content": CLEAN_REPORT_PROMPT.format(text=text)},
            ],
            max_tokens=400,
            temperature=0,
        )
        res = response.choices[0].message.content.strip().strip('"').strip("'")
        _ai_clean_cache[h] = res
        return res
    except Exception as e:
        logger.error(f"Ошибка при очистке отчета: {e}")
        return text

def check_status(text: str, report_type: str | None = "status") -> dict:
    is_forced = report_type in ("status", "daily_fact")

    if is_forced:
        report_type_label = "СТАТУС (отчёт о текущей работе в течение дня)" if report_type == "status" else "ИТОГ ДНЯ / ФАКТ (финальный отчёт о всей проделанной за день работе)"
        report_type_hint = (
            "Оценивай как промежуточный отчёт о том, чем человек занимается прямо сейчас."
            if report_type == "status"
            else "Оценивай как итоговый отчёт за весь рабочий день — ожидай более полного описания того, что было сделано в течение дня."
        )
        mode_instruction = (
            f"ВАЖНО: тип отчёта уже точно определён системой по времени отправки — это {report_type_label}.\n"
            f"Не пытайся определить тип отчёта сам, просто оцени содержание с учётом этого контекста:\n{report_type_hint}"
        )
        json_type_field = ""
        cache_key_prefix = report_type
    else:
        # Авто-режим: используется ВЕЧЕРОМ, когда система не может заранее сказать,
        # это запоздавший статус по конкретному времени или итог за весь день — решает сама модель по смыслу.
        mode_instruction = (
            "ВАЖНО: сейчас вечер, и система НЕ может заранее определить тип отчёта — сотрудник может присылать как\n"
            "запоздавший СТАТУС о конкретном моменте/периоде работы, так и ИТОГ ДНЯ (финальное подведение итогов за весь день).\n"
            "Определи тип САМ по смыслу сообщения:\n"
            "- если текст описывает, что человек делал в какой-то конкретный момент или период работы (похоже на обычный текущий статус) — это \"status\"\n"
            "- если текст звучит как ОБЩЕЕ подведение итогов за ВЕСЬ рабочий день целиком (что в сумме сделано за день) — это \"daily_fact\""
        )
        json_type_field = '"report_type": "status" или "daily_fact",\n\n'
        cache_key_prefix = "auto"

    if groq_client is None:
        return normalize_ai_result({"is_ok": False, "issue": "GROQ_API_KEY не задан"}, text, report_type if is_forced else None)
    h = get_md5(f"{cache_key_prefix}:{text}")
    if h in _ai_status_cache:
        logger.info("check_status: cache hit!")
        return _ai_status_cache[h]
    try:
        response = groq_client.chat.completions.create(
            # Решение проблемы 10: смена модели на llama-3.3-70b-versatile
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "Отвечай только валидным JSON без Markdown."},
                {"role": "user", "content": CHECK_PROMPT_TEMPLATE.format(text=text, mode_instruction=mode_instruction, json_type_field=json_type_field)},
            ],
            max_tokens=400,
            temperature=0,
            response_format={"type": "json_object"},
        )
        raw = response.choices[0].message.content.strip()
        res = normalize_ai_result(json.loads(raw), text, report_type if is_forced else None)
        _ai_status_cache[h] = res
        return res
    except Exception as e:
        return normalize_ai_result({"is_ok": False, "issue": f"Ошибка ИИ: {e}"}, text, report_type if is_forced else None)


async def transcribe_audio_async(file_path: str) -> str:
    return await asyncio.to_thread(transcribe_audio, file_path)


async def clean_report_async(text: str) -> str:
    return await asyncio.to_thread(clean_report, text)


async def check_status_async(text: str, report_type_override: str | None = None) -> dict:
    # report_type_override="status" или "daily_fact" — тип жёстко задан системой (днём всегда так,
    # факт дня физически не может прийти раньше последнего слота статуса).
    # report_type_override=None — "авто"-режим (только вечером): тип определяет сама модель по смыслу,
    # это позволяет различить, например, 2 видео "на статус" и 1 "на факт", присланные в одно и то же время.
    res = await asyncio.to_thread(check_status, text, report_type_override)
    return res


# ══════════════════════════════════════════════════════════════════════════════
# Статистика и автоотчеты (Еженедельные / Ежемесячные / Систематические нарушители)
# ══════════════════════════════════════════════════════════════════════════════

def calculate_worker_stats(worker, start_date_str: str, end_date_str: str, conn) -> dict:
    """
    Рассчитывает статистику сдачи отчетов сотрудником в диапазоне дат [start_date, end_date] (включительно).
    Исключает дни, когда у сотрудника был выходной (report_type = 'not_working').
    Учитывает дату первого отчета сотрудника как начало его работы, чтобы не считать пропуски до начала работы.
    """
    # Определяем дату самого первого отчета сотрудника
    join_row = conn.execute(
        "SELECT MIN(report_date) as start_date FROM reports WHERE telegram_id = ?",
        (worker["telegram_id"],)
    ).fetchone()
    
    adjusted_start_date_str = start_date_str
    if join_row and join_row["start_date"]:
        if join_row["start_date"] > start_date_str:
            adjusted_start_date_str = join_row["start_date"]
            
    # Получаем все отчеты за период
    reports = conn.execute(
        "SELECT * FROM reports WHERE telegram_id = ? AND report_date >= ? AND report_date <= ?",
        (worker["telegram_id"], adjusted_start_date_str, end_date_str)
    ).fetchall()
    
    not_working_dates = set()
    submitted_statuses = set()
    submitted_facts = set()
    total_lates = 0
    total_remarks = 0
    
    for r in reports:
        r_dict = dict(r)
        rep_date = r_dict["report_date"]
        rep_type = r_dict["report_type"]
        
        if rep_type == "not_working":
            not_working_dates.add(rep_date)
        elif rep_type == "status":
            slot = r_dict["slot_time"]
            submitted_statuses.add((rep_date, slot))
            if r_dict["is_late"]:
                total_lates += 1
            if not r_dict["is_ok"]:
                total_remarks += 1
        elif rep_type == "daily_fact":
            submitted_facts.add(rep_date)
            if r_dict["is_late"]:
                total_lates += 1
            if not r_dict["is_ok"]:
                total_remarks += 1

    start_dt = datetime.strptime(adjusted_start_date_str, "%Y-%m-%d").date()
    end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    
    expected_status_count = 0
    expected_fact_count = 0
    
    current_date_it = start_dt
    now = now_local()
    now_date_str = now.strftime("%Y-%m-%d")
    current_mins = now.hour * 60 + now.minute
    
    while current_date_it <= end_dt:
        date_it_str = current_date_it.strftime("%Y-%m-%d")
        if date_it_str not in not_working_dates:
            slots = SCHEDULES.get(worker["schedule"], SCHEDULE_A)
            if date_it_str == now_date_str:
                # Если сегодня, считаем только прошедшие временные слоты
                for slot in slots:
                    hour, minute = map(int, slot.split(":"))
                    slot_mins = hour * 60 + minute
                    if current_mins > slot_mins + LATE_THRESHOLD_MIN or (date_it_str, slot) in submitted_statuses:
                        expected_status_count += 1
                if worker["needs_daily_fact"]:
                    if date_it_str in submitted_facts:
                        expected_fact_count += 1
                    else:
                        last_slot = slots[-1]
                        hour, minute = map(int, last_slot.split(":"))
                        slot_mins = hour * 60 + minute
                        if current_mins > slot_mins + 60:
                            expected_fact_count += 1
            else:
                expected_status_count += len(slots)
                if worker["needs_daily_fact"]:
                    expected_fact_count += 1
        current_date_it += dt_module.timedelta(days=1)
        
    total_expected = expected_status_count + expected_fact_count
    total_submitted = len(submitted_statuses) + len(submitted_facts)
    missed_count = max(0, total_expected - total_submitted)
    
    percent_submitted = 100.0
    if total_expected > 0:
        percent_submitted = (total_submitted / total_expected) * 100.0
        
    return {
        "expected": total_expected,
        "submitted": total_submitted,
        "lates": total_lates,
        "remarks": total_remarks,
        "percent": percent_submitted,
        "missed": missed_count,
    }

def get_violators_threshold() -> int:
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key = 'violators_threshold'").fetchone()
    conn.close()
    if row:
        try:
            return int(row["value"])
        except Exception:
            return 3
    return 3

def save_violators_threshold(val: int):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES ('violators_threshold', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (str(val),)
    )
    conn.commit()
    conn.close()

async def send_weekly_summary(bot):
    now = now_local()
    # Период: с прошлого понедельника по вчерашнее воскресенье
    end_dt = now - dt_module.timedelta(days=1)
    start_dt = now - dt_module.timedelta(days=7)
    
    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = end_dt.strftime("%Y-%m-%d")
    
    conn = get_db()
    workers = conn.execute("SELECT * FROM workers ORDER BY position, sort_order, last_name, first_name").fetchall()
    
    summary_lines = [
        f"📈 *Еженедельный автоматический отчет* за период с {start_dt.strftime('%d.%m.%Y')} по {end_dt.strftime('%d.%m.%Y')}:\n"
    ]
    
    workers_by_dept = {}
    for w in workers:
        lastname_lower = (w["last_name"] or "").lower()
        firstname_lower = (w["first_name"] or "").lower()
        dept_lower = (w["position"] or "").lower()
        if any(x in lastname_lower or x in firstname_lower or x in dept_lower for x in ("отмена", "test", "тест")):
            continue
        if not w["is_active"]:
            continue
        dept = w["position"]
        if dept not in workers_by_dept:
            workers_by_dept[dept] = []
        workers_by_dept[dept].append(w)
        
    for dept, dept_workers in workers_by_dept.items():
        if not dept_workers:
            continue
        summary_lines.append(f"🏢 *Отдел: {dept}*")
        summary_lines.append("──────────────────────────")
        for w in dept_workers:
            stats = calculate_worker_stats(w, start_str, end_str, conn)
            name = f"{w['last_name']} {w['first_name']}"
            summary_lines.append(
                f"👨‍💻 *{name}*:\n"
                f"   • Сдано отчетов: {stats['percent']:.1f}% ({stats['submitted']} из {stats['expected']})\n"
                f"   • Опозданий: {stats['lates']}\n"
                f"   • Замечаний: {stats['remarks']}\n"
            )
        summary_lines.append("")
        
    conn.close()
    
    full_text = "\n".join(summary_lines)
    
    if SUMMARY_CHAT_ID:
        try:
            for part in split_message(full_text):
                await bot.send_message(chat_id=SUMMARY_CHAT_ID, text=part, parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Ошибка при отправке еженедельной сводки в {SUMMARY_CHAT_ID}: {e}")
            
    for admin_id in ADMIN_IDS:
        try:
            for part in split_message(full_text):
                await bot.send_message(chat_id=admin_id, text=part, parse_mode="Markdown")
        except Exception:
            pass

async def send_monthly_summary(bot):
    now = now_local()
    # Период: весь предыдущий календарный месяц
    first_day_this_month = now.replace(day=1)
    last_day_prev_month = first_day_this_month - dt_module.timedelta(days=1)
    start_day_prev_month = last_day_prev_month.replace(day=1)
    
    start_str = start_day_prev_month.strftime("%Y-%m-%d")
    end_str = last_day_prev_month.strftime("%Y-%m-%d")
    
    conn = get_db()
    workers = conn.execute("SELECT * FROM workers ORDER BY position, sort_order, last_name, first_name").fetchall()
    
    months_ru = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    month_name = months_ru.get(start_day_prev_month.month, start_day_prev_month.strftime("%B"))
    
    summary_lines = [
        f"📅 *Ежемесячный автоматический отчет за {month_name} {start_day_prev_month.year}* ({start_day_prev_month.strftime('%d.%m.%Y')} - {last_day_prev_month.strftime('%d.%m.%Y')}):\n"
    ]
    
    workers_by_dept = {}
    for w in workers:
        lastname_lower = (w["last_name"] or "").lower()
        firstname_lower = (w["first_name"] or "").lower()
        dept_lower = (w["position"] or "").lower()
        if any(x in lastname_lower or x in firstname_lower or x in dept_lower for x in ("отмена", "test", "тест")):
            continue
        if not w["is_active"]:
            continue
        dept = w["position"]
        if dept not in workers_by_dept:
            workers_by_dept[dept] = []
        workers_by_dept[dept].append(w)
        
    for dept, dept_workers in workers_by_dept.items():
        if not dept_workers:
            continue
        summary_lines.append(f"🏢 *Отдел: {dept}*")
        summary_lines.append("──────────────────────────")
        for w in dept_workers:
            stats = calculate_worker_stats(w, start_str, end_str, conn)
            name = f"{w['last_name']} {w['first_name']}"
            summary_lines.append(
                f"👨‍💻 *{name}*:\n"
                f"   • Сдано отчетов: {stats['percent']:.1f}% ({stats['submitted']} из {stats['expected']})\n"
                f"   • Опозданий: {stats['lates']}\n"
                f"   • Замечаний: {stats['remarks']}\n"
            )
        summary_lines.append("")
        
    conn.close()
    
    full_text = "\n".join(summary_lines)
    
    if SUMMARY_CHAT_ID:
        try:
            for part in split_message(full_text):
                await bot.send_message(chat_id=SUMMARY_CHAT_ID, text=part, parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Ошибка при отправке ежемесячной сводки в {SUMMARY_CHAT_ID}: {e}")
            
    for admin_id in ADMIN_IDS:
        try:
            for part in split_message(full_text):
                await bot.send_message(chat_id=admin_id, text=part, parse_mode="Markdown")
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# Решение проблемы 4 (Детализированная сводка, различающая статус и факт дня)
# ══════════════════════════════════════════════════════════════════════════════

def generate_daily_summary_text(report_date: str) -> str:
    conn = get_db()
    workers = conn.execute("SELECT * FROM workers ORDER BY position, sort_order, last_name, first_name").fetchall()
    reports = conn.execute("SELECT * FROM reports WHERE report_date = ?", (report_date,)).fetchall()
    conn.close()

    reports_by_worker = {}
    for r in reports:
        tid = r["telegram_id"]
        r_dict = dict(r)
        if tid not in reports_by_worker:
            reports_by_worker[tid] = {"status": {}, "daily_fact": [], "not_working": None}
        
        if r_dict["report_type"] == "status":
            slot = r_dict["slot_time"]
            reports_by_worker[tid]["status"][slot] = r_dict
        elif r_dict["report_type"] == "daily_fact":
            reports_by_worker[tid]["daily_fact"].append(r_dict)
        elif r_dict["report_type"] == "not_working":
            reports_by_worker[tid]["not_working"] = r_dict

    summary_lines = [
        f"📊 Сводка отчетов за {report_date}",
        ""
    ]
    
    workers_by_dept = {}
    for w in workers:
        lastname_lower = (w["last_name"] or "").lower()
        firstname_lower = (w["first_name"] or "").lower()
        dept_lower = (w["position"] or "").lower()
        
        # Исключаем тестовые записи ("отмена" или "тест")
        if any(x in lastname_lower or x in firstname_lower or x in dept_lower for x in ("отмена", "test", "тест")):
            continue
            
        dept = w["position"]
        if dept not in workers_by_dept:
            workers_by_dept[dept] = []
        workers_by_dept[dept].append(w)

    for dept, dept_workers in workers_by_dept.items():
        if not dept_workers:
            continue
            
        # Рассчитываем статистику отдела за сегодня
        dept_expected = 0
        dept_submitted = 0
        for w in dept_workers:
            if not w["is_active"]:
                continue
            w_reports = reports_by_worker.get(w["telegram_id"], {"status": {}, "daily_fact": [], "not_working": None})
            if w_reports.get("not_working"):
                continue
                
            schedule_slots = SCHEDULES.get(w["schedule"], SCHEDULE_A)
            dept_expected += len(schedule_slots)
            
            for slot in schedule_slots:
                if slot in w_reports["status"]:
                    dept_submitted += 1
                    
            if w["needs_daily_fact"]:
                dept_expected += 1
                if w_reports["daily_fact"]:
                    dept_submitted += 1
                    
        dept_percent = (dept_submitted / dept_expected * 100) if dept_expected > 0 else 100.0
        
        summary_lines.append(f"🏢 *Отдел: {dept}*")
        if dept_expected > 0:
            summary_lines.append(f"📈 *Отдел выполнил {dept_submitted}/{dept_expected} отчетов ({dept_percent:.1f}%)*")
        summary_lines.append("──────────────────────────")
        for w in dept_workers:
            tid = w["telegram_id"]
            name = f"{w['last_name']} {w['first_name']}"
            
            if not w["is_active"]:
                summary_lines.append(f"👨‍💻 {name}")
                summary_lines.append("   🏝 В отпуске / на больничном")
                summary_lines.append("")
                continue
                
            w_reports = reports_by_worker.get(tid, {"status": {}, "daily_fact": [], "not_working": None})
            
            # Проверяем, если сегодня человек не работает
            if w_reports.get("not_working"):
                reason = w_reports["not_working"]["format_comment"] or "не указана"
                summary_lines.append(f"👨‍💻 {name}")
                summary_lines.append(f"   🛌 Не работает сегодня (Причина: {reason})")
                summary_lines.append("")
                continue
                
            # 1. Почасовые статусы
            schedule_slots = SCHEDULES.get(w["schedule"], SCHEDULE_A)
            status_segments = []
            issues_list = []
            
            for slot in schedule_slots:
                rep = w_reports["status"].get(slot)
                if rep:
                    status_icon = "✅" if rep["is_ok"] else "⚠️"
                    late_icon = "⏰" if rep["is_late"] else ""
                    status_segments.append(f"{slot} {status_icon}{late_icon}")
                    if not rep["is_ok"]:
                        comment = rep["format_comment"] or "Есть замечание"
                        if comment.startswith("не ОК, "):
                            comment = comment[len("не ОК, "):]
                        elif comment.startswith("не ОК: "):
                            comment = comment[len("не ОК: "):]
                        issues_list.append(f"• {slot} — {comment}")
                else:
                    status_segments.append(f"{slot} ❌")
            
            status_str = " | ".join(status_segments)
            
            # 2. Факт дня (daily_fact)
            if w["needs_daily_fact"]:
                fact_reps = w_reports["daily_fact"]
                if fact_reps:
                    f_rep = fact_reps[-1]
                    if f_rep["is_ok"]:
                        fact_str = "✅ Сдан"
                    else:
                        comment = f_rep["format_comment"] or "Есть замечание"
                        if comment.startswith("не ОК, "):
                            comment = comment[len("не ОК, "):]
                        elif comment.startswith("не ОК: "):
                            comment = comment[len("не ОК: "):]
                        fact_str = f"⚠️ Замечание ({comment})"
                else:
                    fact_str = "❌ Не отправлен"
            else:
                fact_str = "⚪ Не требуется"

            summary_lines.append(f"👨‍💻 {name}")
            summary_lines.append(f"   ⏱ Статусы:  {status_str}")
            summary_lines.append(f"   📋 Итог дня: {fact_str}")
            if issues_list:
                summary_lines.append("   ⚠️ Замечания по статусам:")
                for issue in issues_list:
                    summary_lines.append(f"     {issue}")
            summary_lines.append("")
        summary_lines.append("")

    # ── Секция: Систематические нарушители (за последние 7 дней от даты сводки) ──
    violations_lines = []
    threshold = get_violators_threshold()
    
    try:
        end_dt = datetime.strptime(report_date, "%Y-%m-%d").date()
        start_dt = end_dt - dt_module.timedelta(days=6)
        start_date_7 = start_dt.strftime("%Y-%m-%d")
        
        violators = []
        conn = get_db()
        for w in workers:
            lastname_lower = (w["last_name"] or "").lower()
            firstname_lower = (w["first_name"] or "").lower()
            dept_lower = (w["position"] or "").lower()
            if any(x in lastname_lower or x in firstname_lower or x in dept_lower for x in ("отмена", "test", "тест")):
                continue
            if not w["is_active"]:
                continue
                
            stats = calculate_worker_stats(w, start_date_7, report_date, conn)
            total_violations = stats["missed"] + stats["remarks"]
            if total_violations > threshold:
                violators.append((w, stats, total_violations))
        conn.close()
        
        if violators:
            violations_lines.append("⚠️ *Систематические нарушители (за последние 7 дней):*")
            violations_lines.append("──────────────────────────")
            violators.sort(key=lambda x: x[2], reverse=True)
            for w, stats, total in violators:
                w_name = f"{w['last_name']} {w['first_name']}"
                violations_lines.append(
                    f"👤 *{w_name}* ({w['position']}):\n"
                    f"   • Всего нарушений: *{total}* (Пропущено: {stats['missed']}, Замечаний: {stats['remarks']})\n"
                    f"   • Дисциплина: {stats['percent']:.1f}%\n"
                )
            violations_lines.append("")
    except Exception as e:
        logger.error(f"Ошибка при расчете систематических нарушителей: {e}")

    if violations_lines:
        summary_lines.extend(violations_lines)

    return "\n".join(summary_lines)


# ══════════════════════════════════════════════════════════════════════════════
# Решение проблемы 7 (Обработчик Callback-кнопки переключения результатов)
# ══════════════════════════════════════════════════════════════════════════════

def format_show_date(date_str: str) -> str:
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%d.%m")
    except Exception:
        return date_str

def format_status_or_fact_line(report_type: str, slot_time: str | None, report_date: str) -> str:
    formatted_date = format_show_date(report_date)
    if report_type == "daily_fact":
        return f"Факт за {formatted_date}"
    else:
        slot_str = slot_time or "Неизвестно"
        return f"Статус за {slot_str} за {formatted_date}"

def update_message_metadata(original_text: str, is_ok: bool | None = None, comment: str | None = None, status_val: str | None = None, is_manual: bool = False) -> str:
    lines = original_text.split("\n")
    for i, line in enumerate(lines):
        if (line.startswith("Статус:") or line.startswith("Статус за") or line.startswith("Факт за")) and status_val is not None:
            lines[i] = status_val
        elif (line.startswith("Оценка ИИ:") or line.startswith("Оценка:")) and is_ok is not None:
            label = "Оценка" if (is_manual or line.startswith("Оценка:")) else "Оценка ИИ"
            lines[i] = f"{label}: {'ОК' if is_ok else 'НЕ ОК'}"
        elif (line.startswith("Комментарий ИИ:") or line.startswith("Комментарий:")) and comment is not None:
            label = "Комментарий" if (is_manual or line.startswith("Комментарий:")) else "Комментарий ИИ"
            lines[i] = f"{label}: {comment}"
    return "\n".join(lines)

def update_message_text_fields(original_text: str, is_ok: bool, new_comment: str) -> str:
    return update_message_metadata(original_text, is_ok=is_ok, comment=new_comment, is_manual=True)

def make_report_keyboard(report_id: int, report_type: str | None = None) -> InlineKeyboardMarkup:
    if report_type is None:
        try:
            conn = get_db()
            row = conn.execute("SELECT report_type FROM reports WHERE id = ?", (report_id,)).fetchone()
            conn.close()
            report_type = row["report_type"] if row else "status"
        except Exception:
            report_type = "status"
            
    type_btn_text = "📋 Сделать Итогом дня" if report_type == "status" else "⏱ Сделать Статусом"
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🔄 Изменить оценку (ОК / НЕ ОК)", callback_data=f"fix_toggle_{report_id}"),
            InlineKeyboardButton("✏️ Изменить комментарий", callback_data=f"edit_comment_{report_id}")
        ],
        [
            InlineKeyboardButton(type_btn_text, callback_data=f"toggle_type_{report_id}")
        ]
    ])

async def handle_callback_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    
    if not is_admin(user_id):
        await query.answer("У вас нет прав администратора для корректировки оценок.", show_alert=True)
        return
        
    data = query.data
    if data.startswith("fix_toggle_"):
        report_id = int(data.split("_")[-1])
        
        conn = get_db()
        report = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
        if not report:
            conn.close()
            await query.answer("Запись отчета в БД не найдена.", show_alert=True)
            return
            
        new_ok = 1 if report["is_ok"] == 0 else 0
        new_comment = "ОК (изменено администратором вручную)" if new_ok == 1 else "Замечание (изменено администратором вручную)"
        new_action = f"Скорректировано вручную администратором @{query.from_user.username or user_id}"
        
        conn.execute(
            "UPDATE reports SET is_ok = ?, format_comment = ?, required_action = ? WHERE id = ?",
            (new_ok, new_comment, new_action, report_id)
        )
        
        worker = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (report["telegram_id"],)).fetchone()
        conn.commit()
        conn.close()
        asyncio.create_task(async_sync_gsheets_background())
        
        worker_name = f"{worker['last_name']} {worker['first_name']}" if worker else f"ID {report['telegram_id']}"
        status_emoji = "✅" if new_ok == 1 else "⚠️"
        
        await query.answer("Оценка скорректирована!")
        
        # Обновляем текст сообщения, сохраняя историю
        original_text = query.message.text or ""
        if "Официальный отчет:" in original_text:
            new_text = update_message_text_fields(original_text, new_ok == 1, new_comment)
        else:
            new_text = (
                f"🔧 Оценка отчета изменена вручную администратором @{query.from_user.username or user_id}:\n"
                f"Сотрудник: {worker_name}\n"
                f"Дата отчета: {report['report_date']}\n"
                f"Статус: {report['slot_time'] or report['report_type']}\n"
                f"Новый статус: {status_emoji} ({new_comment})"
            )
        
        kbd = make_report_keyboard(report_id, report["report_type"] if report else None)
        
        try:
            await query.edit_message_text(text=new_text, reply_markup=kbd)
        except Exception as e:
            logger.error(f"Ошибка обновления интерактивной кнопки: {e}")

    elif data.startswith("edit_comment_"):
        report_id = int(data.split("_")[-1])
        
        conn = get_db()
        report = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
        conn.close()
        
        if not report:
            await query.answer("Запись отчета в БД не найдена.", show_alert=True)
            return
            
        context.user_data["editing_comment_report_id"] = report_id
        context.user_data["editing_comment_chat_id"] = query.message.chat_id
        context.user_data["editing_comment_message_id"] = query.message.message_id
        context.user_data["editing_comment_original_text"] = query.message.text
        
        await query.answer()
        try:
            prompt_msg = await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="✏️ Введите новый комментарий ИИ:",
                reply_markup=ForceReply(selective=True)
            )
            context.user_data["editing_comment_prompt_message_id"] = prompt_msg.message_id
        except Exception as e:
            logger.error(f"Ошибка отправки ForceReply: {e}")

    elif data.startswith("toggle_type_"):
        report_id = int(data.split("_")[-1])
        
        conn = get_db()
        try:
            report = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
            if not report:
                await query.answer("Запись отчета в БД не найдена.", show_alert=True)
                return
                
            worker = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (report["telegram_id"],)).fetchone()
            
            current_type = report["report_type"]
            if current_type == "status":
                new_type = "daily_fact"
                new_slot = None
                is_late = 0
                status_display_val = format_status_or_fact_line(new_type, new_slot, report["report_date"])
            else:
                new_type = "status"
                schedule_slots = SCHEDULES.get(worker["schedule"] if worker else "A", SCHEDULE_A)
                try:
                    parts = (report["received_at"] or "00:00:00").split(":")
                    if len(parts) >= 3:
                        h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                    elif len(parts) == 2:
                        h, m, s = int(parts[0]), int(parts[1]), 0
                    else:
                        h, m, s = 0, 0, 0
                    now_dt = now_local().replace(hour=h, minute=m, second=s)
                except Exception:
                    now_dt = now_local()
                new_slot, wait_is_late = find_nearest_slot(schedule_slots, now_dt)
                status_display_val = format_status_or_fact_line(new_type, new_slot, report["report_date"])
                is_late = int(wait_is_late)
                
            conn.execute(
                "UPDATE reports SET report_type = ?, slot_time = ?, is_late = ? WHERE id = ?",
                (new_type, new_slot, is_late, report_id)
            )
            conn.commit()
        finally:
            conn.close()
        
        asyncio.create_task(async_sync_gsheets_background())
        
        await query.answer("Тип отчета изменен!")
        
        original_text = query.message.text or ""
        new_text = update_message_metadata(original_text, status_val=status_display_val)
        
        kbd = make_report_keyboard(report_id, new_type)
        try:
            await query.edit_message_text(text=new_text, reply_markup=kbd)
        except Exception as e:
            logger.error(f"Ошибка обновления типа отчета в сообщении: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# Базовые хэндлеры команд
# ══════════════════════════════════════════════════════════════════════════════

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_type = update.effective_chat.type
    if is_admin(update.effective_user.id) and chat_type == "private":
        await update.message.reply_text("Привет! Выберите действие кнопкой ниже.", reply_markup=MAIN_MENU)
    else:
        # Для зарегистрированных работников показываем меню с кнопкой, для остальных - убираем клавиатуру
        await update.message.reply_text(
            "Привет! Отправьте видеоотчет, когда он будет готов.",
            reply_markup=menu_for_user(update.effective_user.id, chat_type)
        )

async def get_chat_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"ID чата: {update.effective_chat.id}", reply_markup=menu_for_user(update.effective_user.id, update.effective_chat.type))


# ── Нарушители: Управление порогом ──────────────────────────────────────────
async def set_threshold_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Эта команда доступна только администраторам.")
        return
    
    args = context.args
    if not args:
        current = get_violators_threshold()
        await update.message.reply_text(
            f"ℹ️ Текущий порог для систематических нарушителей: *{current}* пропусков/замечаний за 7 дней.\n\n"
            f"Чтобы изменить порог, отправьте: `/threshold <число>` (например, `/threshold 4`).",
            parse_mode="Markdown"
        )
        return
        
    try:
        val = int(args[0])
        if val < 1:
            await update.message.reply_text("❌ Порог должен быть положительным числом.")
            return
        save_violators_threshold(val)
        await update.message.reply_text(f"✅ Порог систематических нарушителей успешно изменен на *{val}*.", parse_mode="Markdown")
    except ValueError:
        await update.message.reply_text("❌ Неверный формат числа. Пример: `/threshold 4`.")


# ── Поиск сотрудника по имени ────────────────────────────────────────────────
async def find_worker_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return ConversationHandler.END
        
    query = ""
    if context.args:
        query = " ".join(context.args).strip()
    elif update.message.text and update.message.text.startswith("/find"):
        parts = update.message.text.split(None, 1)
        if len(parts) > 1:
            query = parts[1].strip()
            
    if not query:
        await update.message.reply_text(
            "📝 Пожалуйста, укажите имя или фамилию для поиска. Например: `/find Иванов`",
            parse_mode="Markdown"
        )
        return ConversationHandler.END

    conn = get_db()
    query_pattern = f"%{query}%"
    rows = conn.execute(
        "SELECT * FROM workers WHERE last_name LIKE ? OR first_name LIKE ? ORDER BY position, sort_order, last_name, first_name",
        (query_pattern, query_pattern)
    ).fetchall()
    conn.close()

    if not rows:
        await update.message.reply_text(f"❌ Сотрудники по запросу «{query}» не найдены.")
        return ConversationHandler.END

    valid_rows = []
    for r in rows:
        lastname_lower = (r["last_name"] or "").lower()
        firstname_lower = (r["first_name"] or "").lower()
        dept_lower = (r["position"] or "").lower()
        if any(x in lastname_lower or x in firstname_lower or x in dept_lower for x in ("отмена", "test", "тест")):
            continue
        valid_rows.append(r)

    if not valid_rows:
        await update.message.reply_text(f"❌ Сотрудники по запросу «{query}» не найдены.")
        return ConversationHandler.END

    if len(valid_rows) == 1:
        worker = valid_rows[0]
        dept_rows = get_workers_by_position(worker["position"])
        idx = 0
        for i, r in enumerate(dept_rows):
            if r["telegram_id"] == worker["telegram_id"]:
                idx = i
                break
                
        context.user_data["list_rows"] = dept_rows
        context.user_data["edit_worker"] = worker
        context.user_data["edit_worker_idx"] = idx

        schedule_str = ", ".join(SCHEDULES.get(worker["schedule"], SCHEDULE_A))
        fact = "да" if worker["needs_daily_fact"] else "нет"
        gname = await get_group_name_async(context.bot, worker["group_id"])
        active_str = "Активен" if worker["is_active"] else "В отпуске / на больничном"
        object_name = worker.get("object_id", "Основной")
        
        info = (
            f"🔍 Найден сотрудник:\n\n"
            f"👤 {worker['last_name']} {worker['first_name']}\n"
            f"Объект: {object_name}\n"
            f"Отдел: {worker['position']}\n"
            f"График: {worker['schedule']} ({schedule_str})\n"
            f"Группа: {gname}\n"
            f"Факт дня: {fact}\n"
            f"Статус работы: {active_str}\n\n"
            f"Что хотите сделать?"
        )

        kbd = ReplyKeyboardMarkup(
            [
                ["📅 История за неделю", "✏️ Номер в списке"],
                ["✏️ Изменить фамилию", "✏️ Изменить имя"],
                ["✏️ Изменить отдел", "✏️ Изменить график"],
                ["✏️ Изменить группу", "✏️ Изменить объект"],
                ["✏️ Факт дня", "✏️ Статус работы"],
                ["🔼 Вверх в списке", "🔽 Вниз в списке"],
                ["❌ Отмена"]
            ],
            resize_keyboard=True,
        )
        await update.message.reply_text(info, reply_markup=kbd)
        return ASK_EDIT_FIELD
    else:
        context.user_data["list_rows"] = valid_rows
        
        lines = [f"🔍 Найдено {len(valid_rows)} сотрудников по вашему запросу:\n"]
        for i, r in enumerate(valid_rows, 1):
            schedule_str = ", ".join(SCHEDULES.get(r["schedule"], SCHEDULE_A))
            fact = "да" if r["needs_daily_fact"] else "нет"
            lines.append(f"{i}. {r['last_name']} {r['first_name']} (Отдел: {r['position']})\n   График: {r['schedule']} | Факт дня: {fact}")
            
        lines.append("\n👉 Выберите сотрудника по номеру из списка:")
        await update.message.reply_text("\n".join(lines), reply_markup=numbered_workers_keyboard(valid_rows))
        return ASK_LIST_WORKER


# ── Сводка за дату ───────────────────────────────────────────────────────────
async def summary_date_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    await update.message.reply_text(
        "📅 Пожалуйста, введите интересующую вас дату в формате *ДД.ММ.ГГГГ* (например, `26.06.2026`):",
        parse_mode="Markdown",
        reply_markup=CANCEL_KEYBOARD
    )
    return ASK_SUMMARY_DATE

async def summary_date_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    raw = update.message.text.strip()
    
    parsed_date = None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            parsed_date = datetime.strptime(raw, fmt)
            break
        except ValueError:
            continue
            
    if parsed_date is None:
        await update.message.reply_text(
            "❌ Неверный формат даты. Пожалуйста, введите дату в формате *ДД.ММ.ГГГГ* (например, `26.06.2026`):",
            parse_mode="Markdown"
        )
        return ASK_SUMMARY_DATE
        
    date_str = parsed_date.strftime("%Y-%m-%d")
    await update.message.reply_text(f"⏳ Формирую сводку за {parsed_date.strftime('%d.%m.%Y')}...")
    summary_text = generate_daily_summary_text(date_str)
    
    parts = split_message(summary_text)
    for part in parts[:-1]:
        await update.message.reply_text(part)
    if parts:
        await update.message.reply_text(parts[-1], reply_markup=MAIN_MENU)
        
    return ConversationHandler.END


# ── Массовое изменение графика отдела ─────────────────────────────────────────
async def dept_schedule_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    await update.message.reply_text(
        "🏢 Выберите отдел, в котором хотите изменить график всех сотрудников:",
        reply_markup=departments_reply_keyboard(context)
    )
    return ASK_EDIT_SCHEDULE_DEPT

async def dept_schedule_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    dept = update.message.text.strip()
    
    conn = get_db()
    rows = conn.execute("SELECT COUNT(*) as cnt FROM workers WHERE lower(position) = lower(?)", (dept,)).fetchone()
    conn.close()
    
    if not rows or rows["cnt"] == 0:
        await update.message.reply_text("❌ Отдел не найден. Попробуйте еще раз или выберите из меню.")
        return ASK_EDIT_SCHEDULE_DEPT
        
    context.user_data["dept_to_change_schedule"] = dept
    await update.message.reply_text(
        f"⚙️ Выберите новый график для сотрудников отдела *{dept}* (всего {rows['cnt']} чел.):",
        parse_mode="Markdown",
        reply_markup=SCHEDULE_KEYBOARD
    )
    return ASK_DEPT_SCHEDULE_VAL

async def dept_schedule_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    sched_val = update.message.text.strip().upper()
    dept = context.user_data.get("dept_to_change_schedule")
    
    if sched_val not in ("A", "B"):
        await update.message.reply_text("❌ Пожалуйста, выберите A или B:", reply_markup=SCHEDULE_KEYBOARD)
        return ASK_DEPT_SCHEDULE_VAL
        
    conn = get_db()
    conn.execute(
        "UPDATE workers SET schedule = ? WHERE lower(position) = lower(?) AND is_active = 1",
        (sched_val, dept)
    )
    conn.commit()
    conn.close()
    
    await update.message.reply_text(
        f"✅ График всех активных сотрудников отдела *{dept}* успешно изменен на *{sched_val}*!",
        parse_mode="Markdown",
        reply_markup=MAIN_MENU
    )
    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    chat_type = update.effective_chat.type
    user_id = update.effective_user.id
    if is_admin(user_id) and chat_type == "private":
        await update.message.reply_text("Действие отменено.", reply_markup=MAIN_MENU)
    else:
        await update.message.reply_text("Действие отменено.", reply_markup=menu_for_user(user_id, chat_type))
    return ConversationHandler.END

async def import_workers_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    await update.message.reply_text(
        "📎 Пожалуйста, отправьте файл Excel (.xlsx) со списком сотрудников.\n\n"
        "Файл должен содержать заголовки: name (или ФИО), position (или должность), и т.д. "
        "А также telegram_id (или ID), если хотите связать аккаунты напрямую.",
        reply_markup=CANCEL_KEYBOARD
    )
    return ASK_IMPORT_FILE

async def import_workers_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    
    doc = update.message.document
    if not doc:
        await update.message.reply_text("❌ Пожалуйста, отправьте файл типа документ (.xlsx).")
        return ASK_IMPORT_FILE
        
    if not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("❌ Формат файла не поддерживается. Пожалуйста, отправьте файл .xlsx")
        return ASK_IMPORT_FILE
        
    await update.message.reply_text("⏳ Получение файла и импорт данных, пожалуйста, подождите...")
    
    try:
        tg_file = await context.bot.get_file(doc.file_id)
        local_path = "workers_temp_import.xlsx"
        await tg_file.download_to_drive(local_path)
        
        workers = read_excel(local_path)
        if not workers:
            await update.message.reply_text("⚠️ Не обнаружено записей в файле или структура не распознана. Проверьте заголовки.", reply_markup=MAIN_MENU)
            if os.path.exists(local_path):
                os.remove(local_path)
            return ConversationHandler.END
            
        success_count = 0
        for w in workers:
            try:
                upsert_worker(
                    telegram_id=w["telegram_id"],
                    last_name=w["last_name"],
                    first_name=w["first_name"],
                    position=w["position"],
                    group_id=w["group_id"],
                    schedule=w["schedule"],
                    needs_daily_fact=w["needs_daily_fact"],
                    object_id=w.get("object_id", "Основной")
                )
                success_count += 1
            except Exception as ex:
                logging.error(f"Error importing worker {w}: {ex}")
                
        # Clean up temp file
        if os.path.exists(local_path):
            os.remove(local_path)
            
        asyncio.create_task(async_sync_gsheets_background())
        await update.message.reply_text(
            f"✅ Импорт успешно завершен!\n"
            f"Загружено/обновлено сотрудников: {success_count}.",
            reply_markup=MAIN_MENU
        )
        return ConversationHandler.END
        
    except Exception as e:
        logging.error(f"Excel import error: {e}")
        await update.message.reply_text(f"❌ Произошла ошибка во время импорта: {e}", reply_markup=MAIN_MENU)
        return ConversationHandler.END


# ══════════════════════════════════════════════════════════════════════════════
# Диалог 📋 Список и редактирование сотрудников
# ══════════════════════════════════════════════════════════════════════════════

async def list_workers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    rows = get_all_workers()
    if not rows:
        await update.message.reply_text("В базе пока нет сотрудников.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
    kbd, _ = positions_keyboard(rows)
    await update.message.reply_text("Выберите отдел:", reply_markup=kbd)
    return ASK_LIST_DEPARTMENT

async def list_workers_department(update: Update, context: ContextTypes.DEFAULT_TYPE):
    position = update.message.text.strip()
    rows = get_workers_by_position(position)
    if not rows:
        await update.message.reply_text(f"Отдел «{position}» не найден.", reply_markup=MAIN_MENU)
        return ConversationHandler.END

    context.user_data["list_position"] = position
    context.user_data["list_rows"] = [dict(r) for r in rows]

    group_names = get_all_group_names()
    lines = [f"Сотрудники отдела «{position}»:\n"]
    for i, row in enumerate(rows, 1):
        schedule_str = ", ".join(SCHEDULES.get(row["schedule"], SCHEDULE_A))
        fact = "да" if row["needs_daily_fact"] else "нет"
        gname = group_names.get(row["group_id"], str(row["group_id"]))
        lines.append(f"{i}. {row['last_name']} {row['first_name']}\n   График: {row['schedule']} ({schedule_str})\n   Группа: {gname} | Факт дня: {fact}")

    # Решение проблемы 11: Объединение уведомления со списком сотрудников
    lines.append("\n👉 Выберите сотрудника по номеру из списка:")
    await update.message.reply_text("\n".join(lines), reply_markup=numbered_workers_keyboard(rows))
    return ASK_LIST_WORKER

async def list_workers_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    rows = context.user_data.get("list_rows", [])
    num_str = raw.split(".")[0].strip()
    if not num_str.isdigit():
        await update.message.reply_text("Выберите сотрудника по номеру из списка.")
        return ASK_LIST_WORKER

    idx = int(num_str) - 1
    if idx < 0 or idx >= len(rows):
        await update.message.reply_text("Номер не найден. Попробуйте ещё раз.")
        return ASK_LIST_WORKER

    worker = rows[idx]
    context.user_data["edit_worker"] = worker
    context.user_data["edit_worker_idx"] = idx

    schedule_str = ", ".join(SCHEDULES.get(worker["schedule"], SCHEDULE_A))
    fact = "да" if worker["needs_daily_fact"] else "нет"
    gname = await get_group_name_async(context.bot, worker["group_id"])
    active_str = "Активен" if worker["is_active"] else "В отпуске / на больничном"
    object_name = worker.get("object_id", "Основной")
    info = f"👤 {worker['last_name']} {worker['first_name']}\nОбъект: {object_name}\nОтдел: {worker['position']}\nГрафик: {worker['schedule']} ({schedule_str})\nГруппа: {gname}\nФакт дня: {fact}\nСтатус работы: {active_str}\n\nЧто хотите сделать?"

    # Добавление кнопки Истории еженедельных оценок (проблема 6)
    kbd = ReplyKeyboardMarkup(
        [
            ["📅 История за неделю", "✏️ Номер в списке"],
            ["✏️ Изменить фамилию", "✏️ Изменить имя"],
            ["✏️ Изменить отдел", "✏️ Изменить график"],
            ["✏️ Изменить группу", "✏️ Изменить объект"],
            ["✏️ Факт дня", "✏️ Статус работы"],
            ["🔼 Вверх в списке", "🔽 Вниз в списке"],
            ["❌ Отмена"]
        ],
        resize_keyboard=True,
    )
    await update.message.reply_text(info, reply_markup=kbd)
    return ASK_EDIT_FIELD

async def list_workers_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    action = update.message.text.strip()
    worker = context.user_data.get("edit_worker")
    rows = context.user_data.get("list_rows", [])
    idx = context.user_data.get("edit_worker_idx", 0)

    if not worker:
        await update.message.reply_text("Ошибка состояния. Начните сначала.", reply_markup=MAIN_MENU)
        return ConversationHandler.END

    if action == "📅 История за неделю":
        history_text = get_worker_history_last_week(worker["telegram_id"])
        await update.message.reply_text(history_text, reply_markup=MAIN_MENU)
        context.user_data.clear()
        return ConversationHandler.END

    if action == "✏️ Номер в списке":
        await update.message.reply_text(
            f"Введите новый порядковый номер для {worker['last_name']} {worker['first_name']}\n"
            f"(сейчас он под номером {idx + 1}, всего в отделе {len(rows)} сотрудников):",
            reply_markup=CANCEL_KEYBOARD
        )
        return ASK_EDIT_SORT_ORDER

    if action in ("🔼 Вверх в списке", "🔽 Вниз в списке"):
        target_idx = idx - 1 if action == "🔼 Вверх в списке" else idx + 1
        if target_idx < 0 or target_idx >= len(rows):
            await update.message.reply_text("Сотрудник уже на краю списка.", reply_markup=MAIN_MENU)
            return ConversationHandler.END
        swap_sort_order(worker["telegram_id"], rows[target_idx]["telegram_id"])
        await update.message.reply_text(f"Порядок изменен для {worker['last_name']}.", reply_markup=MAIN_MENU)
        context.user_data.clear()
        return ConversationHandler.END

    field_map = {
        "✏️ Изменить фамилию": ("last_name", "Введите новую фамилию:"),
        "✏️ Изменить имя": ("first_name", "Введите новое имя:"),
        "✏️ Изменить отдел": ("position", "Введите новое название отдела:"),
        "✏️ Изменить группу": ("group_id", "Введите новый ID группы Telegram (0 = по умолчанию):"),
        "✏️ Изменить объект": ("object_id", "Введите название объекта для сотрудника (например: Основной, Северный):"),
        "✏️ Изменить график": ("schedule", None),
        "✏️ Факт дня": ("needs_daily_fact", None),
        "✏️ Статус работы": ("is_active", None),
    }

    if action not in field_map:
        await update.message.reply_text("Выберите действие кнопкой.", reply_markup=MAIN_MENU)
        return ConversationHandler.END

    field, prompt = field_map[action]
    context.user_data["edit_field"] = field

    if field == "position":
        await update.message.reply_text(
            "Выберите существующий отдел из списка ниже или введите новое название отдела:",
            reply_markup=departments_reply_keyboard(context)
        )
        return ASK_EDIT_VALUE
    if field == "schedule":
        await update.message.reply_text("Выберите новый график:", reply_markup=SCHEDULE_KEYBOARD)
        return ASK_EDIT_SCHEDULE
    if field == "needs_daily_fact":
        await update.message.reply_text("Нужен ли ежедневный факт дня?", reply_markup=YES_NO_KEYBOARD)
        return ASK_EDIT_DAILY_FACT
    if field == "is_active":
        await update.message.reply_text(
            "Выберите новый статус работы:",
            reply_markup=ReplyKeyboardMarkup([["Активен", "В отпуске / на больничном"], ["❌ Отмена"]], resize_keyboard=True)
        )
        return ASK_EDIT_STATUS_WORK
    if field == "group_id":
        await update.message.reply_text(prompt, reply_markup=CANCEL_KEYBOARD)
        return ASK_EDIT_GROUP_VALUE

    await update.message.reply_text(prompt, reply_markup=CANCEL_KEYBOARD)
    return ASK_EDIT_VALUE

async def edit_value_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    value = update.message.text.strip()
    worker = context.user_data.get("edit_worker")
    field = context.user_data.get("edit_field")
    
    if field == "position":
        update_worker_field(worker["telegram_id"], "position", value)
        
        new_dept = value
        new_dept_workers = get_workers_by_position(new_dept)
        num_workers = len(new_dept_workers)
        
        context.user_data["move_new_dept"] = new_dept
        context.user_data["move_new_dept_workers"] = [dict(w) for w in new_dept_workers]
        context.user_data["move_worker_id"] = worker["telegram_id"]
        context.user_data["move_worker_name"] = f"{worker['last_name']} {worker['first_name']}"
        
        kbd = ReplyKeyboardMarkup([["Оставить в конце"], ["❌ Отмена"]], resize_keyboard=True)
        await update.message.reply_text(
            f"✅ Сотрудник {worker['last_name']} {worker['first_name']} успешно переведен в отдел «{new_dept}».\n"
            f"Сейчас он находится в конце списка (под номером {num_workers}).\n\n"
            f"Вы можете изменить его порядковый номер в новом отделе.\n"
            f"Введите порядковый номер (число от 1 до {num_workers}) или нажмите кнопку «Оставить в конце»:",
            reply_markup=kbd
        )
        return ASK_MOVE_POSITION_ORDER
    else:
        update_worker_field(worker["telegram_id"], field, value)
        await update.message.reply_text("Данные обновлены.", reply_markup=MAIN_MENU)
        context.user_data.clear()
        return ConversationHandler.END

async def edit_group_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if not raw.lstrip("-").isdigit():
        await update.message.reply_text("Введите числовой ID:")
        return ASK_EDIT_GROUP_VALUE
    worker = context.user_data.get("edit_worker")
    final_id = DEFAULT_GROUP_ID if int(raw) == 0 else int(raw)
    update_worker_field(worker["telegram_id"], "group_id", final_id)
    gname = await fetch_and_save_group_name(context.bot, final_id)
    await update.message.reply_text(f"Группа обновлена на: {gname}", reply_markup=MAIN_MENU)
    context.user_data.clear()
    return ConversationHandler.END

async def edit_schedule_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip().upper()
    if raw not in SCHEDULES: return ASK_EDIT_SCHEDULE
    worker = context.user_data.get("edit_worker")
    update_worker_field(worker["telegram_id"], "schedule", raw)
    await update.message.reply_text(f"График изменен на {raw}", reply_markup=MAIN_MENU)
    context.user_data.clear()
    return ConversationHandler.END

async def edit_daily_fact_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip().lower()
    if raw not in ("да", "нет"): return ASK_EDIT_DAILY_FACT
    worker = context.user_data.get("edit_worker")
    update_worker_field(worker["telegram_id"], "needs_daily_fact", 1 if raw == "да" else 0)
    await update.message.reply_text(f"Параметр Факт Дня изменен на: {raw}", reply_markup=MAIN_MENU)
    context.user_data.clear()
    return ConversationHandler.END

async def edit_status_work_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip().lower()
    if raw == "❌ отмена":
        await update.message.reply_text("Изменение отменено.", reply_markup=MAIN_MENU)
        context.user_data.clear()
        return ConversationHandler.END
    if raw not in ("активен", "в отпуске / на больничном"):
        return ASK_EDIT_STATUS_WORK
    worker = context.user_data.get("edit_worker")
    val = 1 if raw == "активен" else 0
    update_worker_field(worker["telegram_id"], "is_active", val)
    await update.message.reply_text(f"Статус работы изменен на: {raw}", reply_markup=MAIN_MENU)
    context.user_data.clear()
    return ConversationHandler.END

async def edit_sort_order_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == "❌ Отмена":
        await update.message.reply_text("Изменение отменено.", reply_markup=MAIN_MENU)
        context.user_data.clear()
        return ConversationHandler.END
        
    if not raw.isdigit():
        await update.message.reply_text("Пожалуйста, введите положительное целое число для порядкового номера:")
        return ASK_EDIT_SORT_ORDER
        
    target_num = int(raw)
    worker = context.user_data.get("edit_worker")
    rows = context.user_data.get("list_rows", [])
    
    if not worker or not rows:
        await update.message.reply_text("Ошибка сессии. Начните сначала.", reply_markup=MAIN_MENU)
        context.user_data.clear()
        return ConversationHandler.END
        
    num_workers = len(rows)
    if target_num < 1 or target_num > num_workers:
        await update.message.reply_text(f"Недопустимый порядковый номер. Пожалуйста, введите число от 1 до {num_workers}:")
        return ASK_EDIT_SORT_ORDER
        
    target_idx = target_num - 1 # 0-based target index
    
    position = worker["position"]
    all_dept_workers = [dict(r) for r in get_workers_by_position(position)]
    
    target_worker_id = worker["telegram_id"]
    worker_to_move = None
    for idx, w in enumerate(all_dept_workers):
        if w["telegram_id"] == target_worker_id:
            worker_to_move = all_dept_workers.pop(idx)
            break
            
    if worker_to_move is not None:
        all_dept_workers.insert(target_idx, worker_to_move)
        
        conn = get_db()
        for i, w in enumerate(all_dept_workers):
            conn.execute("UPDATE workers SET sort_order = ? WHERE telegram_id = ?", (i, w["telegram_id"]))
        conn.commit()
        conn.close()
        
        name = f"{worker['last_name']} {worker['first_name']}"
        await update.message.reply_text(
            f"✅ Порядковый номер сотрудника {name} успешно изменен на {target_num}.",
            reply_markup=MAIN_MENU
        )
    else:
        await update.message.reply_text("Сотрудник не найден в текущем списке отдела.", reply_markup=MAIN_MENU)
        
    context.user_data.clear()
    return ConversationHandler.END


async def edit_move_position_order_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    new_dept = context.user_data.get("move_new_dept")
    worker_id = context.user_data.get("move_worker_id")
    worker_name = context.user_data.get("move_worker_name")
    
    if not new_dept or not worker_id:
        await update.message.reply_text("Ошибка сессии. Изменения применились без сортировки.", reply_markup=MAIN_MENU)
        context.user_data.clear()
        return ConversationHandler.END
        
    new_dept_workers = [dict(w) for w in get_workers_by_position(new_dept)]
        
    if raw == "❌ Отмена" or raw.lower() == "отмена":
        await update.message.reply_text(
            f"Сотрудник {worker_name} оставлен в конце списка отдела «{new_dept}».",
            reply_markup=MAIN_MENU
        )
        context.user_data.clear()
        return ConversationHandler.END
        
    if raw == "Оставить в конце" or raw.lower() == "оставить в конце":
        await update.message.reply_text(
            f"Сотрудник {worker_name} сохранен на последней позиции в отделе «{new_dept}».",
            reply_markup=MAIN_MENU
        )
        context.user_data.clear()
        return ConversationHandler.END
        
    if not raw.isdigit():
        await update.message.reply_text(
            f"Пожалуйста, введите целое положительное число (от 1 до {len(new_dept_workers)}) "
            f"или нажмите кнопку «Оставить в конце»:"
        )
        return ASK_MOVE_POSITION_ORDER
        
    target_num = int(raw)
    num_workers = len(new_dept_workers)
    if target_num < 1 or target_num > num_workers:
        await update.message.reply_text(
            f"Недопустимый порядковый номер. Пожалуйста, введите число от 1 до {num_workers}:"
        )
        return ASK_MOVE_POSITION_ORDER
        
    target_idx = target_num - 1
    
    worker_to_move = None
    other_workers = []
    for w in new_dept_workers:
        if w["telegram_id"] == worker_id:
            worker_to_move = w
        else:
            other_workers.append(w)
            
    if worker_to_move is not None:
        other_workers.insert(target_idx, worker_to_move)
        
        conn = get_db()
        for i, w in enumerate(other_workers):
            conn.execute("UPDATE workers SET sort_order = ? WHERE telegram_id = ?", (i, w["telegram_id"]))
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            f"✅ Порядковый номер сотрудника {worker_name} в отделе «{new_dept}» успешно изменен на {target_num}.",
            reply_markup=MAIN_MENU
        )
    else:
        await update.message.reply_text(
            f"Сотрудник сохранен в отделе «{new_dept}» в конце списка.",
            reply_markup=MAIN_MENU
        )
        
    context.user_data.clear()
    return ConversationHandler.END


# ══════════════════════════════════════════════════════════════════════════════
# Диалог ➕ Добавление сотрудника
# ══════════════════════════════════════════════════════════════════════════════

async def add_worker_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    context.user_data.clear()
    await update.message.reply_text("Введите Telegram ID сотрудника:", reply_markup=CANCEL_KEYBOARD)
    return ASK_WORKER_ID

async def add_worker_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if not raw.lstrip("-").isdigit():
        await update.message.reply_text("Введите числовой ID:")
        return ASK_WORKER_ID
    worker_id = int(raw)
    context.user_data["new_worker_id"] = worker_id
    
    pending = get_pending_unregistered_user(worker_id)
    if pending:
        context.user_data["pending_last_name"] = pending["last_name"]
        context.user_data["pending_first_name"] = pending["first_name"]
        
        await update.message.reply_text(
            f"Найден временный отчет сотрудника!\n"
            f"ФИО: {pending['last_name']} {pending['first_name']}\n\n"
            f"Подтвердите фамилию (нажмите на кнопку ниже) или введите новую:",
            reply_markup=ReplyKeyboardMarkup([[pending["last_name"]], ["❌ Отмена"]], resize_keyboard=True)
        )
    else:
        await update.message.reply_text("Введите фамилию:", reply_markup=CANCEL_KEYBOARD)
    return ASK_LASTNAME

async def add_worker_lastname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    val = update.message.text.strip()
    context.user_data["last_name"] = val
    
    pending_first = context.user_data.get("pending_first_name")
    if pending_first:
        await update.message.reply_text(
            "Подтвердите имя (нажмите на кнопку ниже) или введите новое:",
            reply_markup=ReplyKeyboardMarkup([[pending_first], ["❌ Отмена"]], resize_keyboard=True)
        )
    else:
        await update.message.reply_text("Введите имя:", reply_markup=CANCEL_KEYBOARD)
    return ASK_FIRSTNAME

async def add_worker_firstname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["first_name"] = update.message.text.strip()
    await update.message.reply_text(
        "Выберите существующий отдел из списка ниже или введите новое название отдела:",
        reply_markup=departments_reply_keyboard(context)
    )
    return ASK_POSITION

async def add_worker_position(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["position"] = update.message.text.strip()
    await update.message.reply_text("Введите ID группы Telegram (или 0 для группы по умолчанию):", reply_markup=CANCEL_KEYBOARD)
    return ASK_GROUP

async def add_worker_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if not raw.lstrip("-").isdigit(): return ASK_GROUP
    context.user_data["group_id"] = DEFAULT_GROUP_ID if int(raw) == 0 else int(raw)
    await update.message.reply_text("Выберите график отчетов (A или B):", reply_markup=SCHEDULE_KEYBOARD)
    return ASK_SCHEDULE

async def add_worker_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip().upper()
    if raw not in SCHEDULES: return ASK_SCHEDULE
    context.user_data["schedule"] = raw
    await update.message.reply_text("Нужно ли присылать ежедневный факт дня? (Да/Нет)", reply_markup=YES_NO_KEYBOARD)
    return ASK_NEEDS_DAILY_FACT

async def add_worker_needs_daily_fact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip().lower()
    if raw not in ("да", "нет"): return ASK_NEEDS_DAILY_FACT

    next_order = get_next_sort_order(context.user_data["position"])
    upsert_worker(
        telegram_id=context.user_data["new_worker_id"],
        last_name=context.user_data["last_name"],
        first_name=context.user_data["first_name"],
        position=context.user_data["position"],
        group_id=context.user_data["group_id"],
        schedule=context.user_data["schedule"],
        needs_daily_fact=(raw == "да"),
        sort_order=next_order,
    )
    delete_pending_unregistered_user(context.user_data["new_worker_id"])
    await fetch_and_save_group_name(context.bot, context.user_data["group_id"])
    
    # Отправляем уведомление о новом сотруднике в группу
    target_group_id = context.user_data["group_id"] or DEFAULT_GROUP_ID
    try:
        notify_msg = f"👤 {context.user_data['last_name']} {context.user_data['first_name']} добавлен в систему, ID: {context.user_data['new_worker_id']}"
        await context.bot.send_message(chat_id=target_group_id, text=notify_msg)
        logger.info(f"Отправлено групповое уведомление о добавлении нового сотрудника в чат {target_group_id}")
    except Exception as e:
        logger.warning(f"Не удалось отправить уведомление о новом сотруднике в группу {target_group_id}: {e}")

    asyncio.create_task(async_sync_gsheets_background())
    await update.message.reply_text("Сотрудник успешно добавлен в базу!", reply_markup=MAIN_MENU)
    context.user_data.clear()
    return ConversationHandler.END


# ══════════════════════════════════════════════════════════════════════════════
# Диалог ➖ Удаление сотрудника
# ══════════════════════════════════════════════════════════════════════════════

async def delete_worker_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    rows = get_all_workers()
    if not rows:
        await update.message.reply_text("В базе нет сотрудников.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
    kbd, _ = positions_keyboard(rows)
    await update.message.reply_text("Выберите отдел сотрудника для удаления:", reply_markup=kbd)
    return ASK_REMOVE_DEPARTMENT

async def delete_worker_department(update: Update, context: ContextTypes.DEFAULT_TYPE):
    position = update.message.text.strip()
    rows = get_workers_by_position(position)
    if not rows: return ConversationHandler.END
    context.user_data["remove_rows"] = [dict(r) for r in rows]
    await update.message.reply_text("Выберите кого удалить:", reply_markup=numbered_workers_keyboard(rows))
    return ASK_REMOVE_WORKER

async def delete_worker_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    rows = context.user_data.get("remove_rows", [])
    num_str = raw.split(".")[0].strip()
    if not num_str.isdigit(): return ASK_REMOVE_WORKER
    idx = int(num_str) - 1
    if idx < 0 or idx >= len(rows): return ASK_REMOVE_WORKER

    worker = rows[idx]
    context.user_data["worker_to_delete"] = worker
    kbd = ReplyKeyboardMarkup([["Да, удалить", "Нет, отмена"]], resize_keyboard=True)
    await update.message.reply_text(
        f"⚠️ Вы уверены, что хотите удалить сотрудника {worker['last_name']} {worker['first_name']} (ID: {worker['telegram_id']})?",
        reply_markup=kbd
    )
    return ASK_CONFIRM_DELETE

async def delete_worker_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choice = update.message.text.strip()
    if choice == "Да, удалить":
        worker = context.user_data.get("worker_to_delete")
        if worker:
            delete_worker(worker["telegram_id"])
            asyncio.create_task(async_sync_gsheets_background())
            await update.message.reply_text(f"✅ Сотрудник {worker['last_name']} успешно удален.", reply_markup=MAIN_MENU)
    else:
        await update.message.reply_text("Удаление сотрудника отменено.", reply_markup=MAIN_MENU)
        
    context.user_data.clear()
    return ConversationHandler.END


# ══════════════════════════════════════════════════════════════════════════════
# Просмотр сотрудников отдела и Сводка
# ══════════════════════════════════════════════════════════════════════════════

async def department_workers_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    rows = get_all_workers()
    if not rows:
        await update.message.reply_text("В базе нет сотрудников.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
    kbd, _ = positions_keyboard(rows)
    await update.message.reply_text("Выберите отдел для просмотра:", reply_markup=kbd)
    return ASK_DEPARTMENT

async def department_workers_show(update: Update, context: ContextTypes.DEFAULT_TYPE):
    position = update.message.text.strip()
    rows = get_workers_by_position(position)
    if not rows:
        await update.message.reply_text("Сотрудники не найдены.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
    
    lines = [f"📋 Отдел: {position}"]
    for i, r in enumerate(rows, 1):
        lines.append(f"{i}. {r['last_name']} {r['first_name']}")
    await update.message.reply_text("\n".join(lines), reply_markup=MAIN_MENU)
    return ConversationHandler.END

async def send_summary_now(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return
    now = now_local()
    date_str = now.strftime("%Y-%m-%d")
    summary_text = generate_daily_summary_text(date_str)
    
    parts = split_message(summary_text)
    for part in parts[:-1]:
        await update.message.reply_text(part)
    if parts:
        await update.message.reply_text(parts[-1], reply_markup=MAIN_MENU)
    
    if SUMMARY_CHAT_ID and SUMMARY_CHAT_ID != update.effective_chat.id:
        try:
            for part in split_message(summary_text):
                await context.bot.send_message(chat_id=SUMMARY_CHAT_ID, text=part)
        except Exception as e:
            logger.error(f"Не удалось отправить сводку в чат {SUMMARY_CHAT_ID}: {e}")


async def myreports(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    conn = get_db()
    try:
        worker = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (user_id,)).fetchone()
        if not worker:
            await update.message.reply_text("Вы не зарегистрированы в системе. Обратитесь к администратору.")
            return
            
        now = now_local()
        thirty_days_ago = (now - dt_module.timedelta(days=30)).strftime("%Y-%m-%d")
        seven_days_ago = (now - dt_module.timedelta(days=7)).strftime("%Y-%m-%d")
        
        # Извлекаем отчеты за последние 30 дней для этого сотрудника
        reports_30 = conn.execute(
            "SELECT * FROM reports WHERE telegram_id = ? AND report_date >= ? ORDER BY report_date DESC, received_at DESC",
            (user_id, thirty_days_ago)
        ).fetchall()
        
        # Получаем дату первого отчета
        join_row = conn.execute(
            "SELECT MIN(report_date) as start_date FROM reports WHERE telegram_id = ?",
            (user_id,)
        ).fetchone()
    finally:
        conn.close()

    # Сначала считаем статистику за 30 дней
    not_working_dates = set()
    submitted_statuses = set()
    submitted_facts = set()
    total_lates = 0
    total_remarks = 0
    
    reports_7 = []
    
    for r in reports_30:
        r_dict = dict(r)
        rep_date = r_dict["report_date"]
        rep_type = r_dict["report_type"]
        
        if rep_date >= seven_days_ago:
            reports_7.append(r)
            
        if rep_type == "not_working":
            not_working_dates.add(rep_date)
        elif rep_type == "status":
            slot = r_dict["slot_time"]
            submitted_statuses.add((rep_date, slot))
            if r_dict["is_late"]:
                total_lates += 1
            if not r_dict["is_ok"]:
                total_remarks += 1
        elif rep_type == "daily_fact":
            submitted_facts.add(rep_date)
            if r_dict["is_late"]:
                total_lates += 1
            if not r_dict["is_ok"]:
                total_remarks += 1

    start_date_str = thirty_days_ago
    if join_row and join_row["start_date"]:
        if join_row["start_date"] > thirty_days_ago:
            start_date_str = join_row["start_date"]
            
    # Расчет ожидаемых отчетов
    start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_dt = now.date()
    
    expected_status_count = 0
    expected_fact_count = 0
    
    current_date_it = start_dt
    while current_date_it <= end_dt:
        date_it_str = current_date_it.strftime("%Y-%m-%d")
        if date_it_str not in not_working_dates:
            slots = SCHEDULES.get(worker["schedule"], SCHEDULE_A)
            if date_it_str == now.strftime("%Y-%m-%d"):
                current_mins = now.hour * 60 + now.minute
                for slot in slots:
                    hour, minute = map(int, slot.split(":"))
                    slot_mins = hour * 60 + minute
                    if current_mins > slot_mins + LATE_THRESHOLD_MIN or (date_it_str, slot) in submitted_statuses:
                        expected_status_count += 1
                if worker["needs_daily_fact"]:
                    if date_it_str in submitted_facts:
                        expected_fact_count += 1
                    else:
                        last_slot = slots[-1]
                        hour, minute = map(int, last_slot.split(":"))
                        slot_mins = hour * 60 + minute
                        if current_mins > slot_mins + 60:
                            expected_fact_count += 1
            else:
                expected_status_count += len(slots)
                if worker["needs_daily_fact"]:
                    expected_fact_count += 1
        current_date_it += dt_module.timedelta(days=1)
        
    total_expected = expected_status_count + expected_fact_count
    total_submitted = len(submitted_statuses) + len(submitted_facts)
    
    percent_submitted = 100.0
    if total_expected > 0:
        percent_submitted = (total_submitted / total_expected) * 100.0
        
    stats_header = (
        f"📊 *Ваша дисциплина за последние 30 дней:*\n"
        f"📈 *Сдано отчетов:* {percent_submitted:.1f}% ({total_submitted} из {total_expected})\n"
        f"⏰ *Опозданий:* {total_lates}\n"
        f"⚠️ *Замечаний:* {total_remarks}\n"
        f"──────────────────────────\n"
    )

    if not reports_7:
        await update.message.reply_text(
            stats_header + "\nУ вас нет отчетов за последние 7 дней для детальной истории.",
            parse_mode="Markdown"
        )
        return

    lines = [
        stats_header,
        f"📋 *Ваши отчеты за последние 7 дней ({worker['last_name']} {worker['first_name']}):*\n"
    ]
    current_date = None
    for r in reports_7:
        r_dict = dict(r)
        rep_date = r_dict["report_date"]
        if rep_date != current_date:
            current_date = rep_date
            lines.append(f"\n📅 *{current_date}:*")
        
        slot_time = r_dict["slot_time"]
        rep_type_str = f"Статус {slot_time}" if r_dict["report_type"] == "status" else "Факт дня (Итог)"
        is_ok_str = "✅ ОК" if r_dict["is_ok"] else "⚠️ НЕ ОК"
        comment_str = f"AI: {r_dict['format_comment']}" if r_dict["format_comment"] else "Проверка не требуется"
        
        lines.append(
            f"• {r_dict['received_at']} - {rep_type_str} | {is_ok_str}\n"
            f"  {comment_str}\n"
            f"  Текст: \"{r_dict['raw_text']}\""
        )
    
    full_text = "\n".join(lines)
    for part in split_message(full_text):
        await update.message.reply_text(part, parse_mode="Markdown")


async def status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    conn = get_db()
    try:
        worker = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (user_id,)).fetchone()
        if not worker:
            await update.message.reply_text("Вы не зарегистрированы в системе. Обратитесь к администратору.")
            return
        
        now = now_local()
        date_str = now.strftime("%Y-%m-%d")
        
        reports = conn.execute(
            "SELECT * FROM reports WHERE telegram_id = ? AND report_date = ?",
            (user_id, date_str)
        ).fetchall()
    finally:
        conn.close()

    status_reps = {}
    fact_reps = []
    not_working = None
    
    for r in reports:
        r_dict = dict(r)
        if r_dict["report_type"] == "status":
            slot = r_dict["slot_time"]
            status_reps[slot] = r_dict
        elif r_dict["report_type"] == "daily_fact":
            fact_reps.append(r_dict)
        elif r_dict["report_type"] == "not_working":
            not_working = r_dict

    name = f"{worker['last_name']} {worker['first_name']}"
    lines = [
        f"📊 *Ваш статус на сегодня ({now.strftime('%d.%m.%Y')})*",
        f"👤 *Сотрудник:* {name}",
        f"💼 *Отдел/Должность:* {worker['position']}",
        f"📅 *График:* {worker['schedule']}",
        ""
    ]
    
    if not worker["is_active"]:
        lines.append("🏝 Вы отмечены как неактивный (в отпуске / на больничном).")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        return
        
    if not_working:
        reason = not_working["format_comment"] or "не указана"
        lines.append(f"🛌 *Сегодня выходной:* Вы отметили, что не работаете сегодня (Причина: {reason}).")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        return

    # Check status slots
    schedule_slots = SCHEDULES.get(worker["schedule"], SCHEDULE_A)
    status_segments = []
    issues_list = []
    
    submitted_count = 0
    expected_count = len(schedule_slots)
    
    for slot in schedule_slots:
        rep = status_reps.get(slot)
        if rep:
            submitted_count += 1
            status_icon = "✅" if rep["is_ok"] else "⚠️"
            late_icon = " ⏰" if rep["is_late"] else ""
            status_segments.append(f"• *{slot}*: {status_icon} Сдан{late_icon}")
            if not rep["is_ok"]:
                comment = rep["format_comment"] or "Есть замечание"
                if comment.startswith("не ОК, "):
                    comment = comment[len("не ОК, "):]
                elif comment.startswith("не ОК: "):
                    comment = comment[len("не ОК: "):]
                issues_list.append(f"  └ ⚠️ {slot} — {comment}")
        else:
            # Check if slot time has passed
            sh, sm = map(int, slot.split(":"))
            current_mins = now.hour * 60 + now.minute
            slot_mins = sh * 60 + sm
            if current_mins > slot_mins + LATE_THRESHOLD_MIN:
                status_segments.append(f"• *{slot}*: ❌ Пропущен")
            else:
                status_segments.append(f"• *{slot}*: ⏳ Ожидается")
                
    lines.append(f"⏱ *Отчёты за слоты ({submitted_count}/{expected_count}):*")
    lines.extend(status_segments)
    lines.append("")
    
    # Check daily_fact
    if worker["needs_daily_fact"]:
        if fact_reps:
            f_rep = fact_reps[-1]
            if f_rep["is_ok"]:
                fact_str = "✅ Сдан"
            else:
                comment = f_rep["format_comment"] or "Есть замечание"
                if comment.startswith("не ОК, "):
                    comment = comment[len("не ОК, "):]
                elif comment.startswith("не ОК: "):
                    comment = comment[len("не ОК: "):]
                fact_str = f"⚠️ Замечание ({comment})"
        else:
            fact_str = "❌ Не отправлен"
    else:
        fact_str = "⚪ Не требуется"
        
    lines.append(f"📋 *Итог дня (Факт дня):* {fact_str}")
    
    if issues_list:
        lines.append("")
        lines.append("⚠️ *Замечания по сегодняшним отчетам:*")
        lines.extend(issues_list)
        
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


def format_date_no_year(date_str) -> str:
    if date_str is None:
        return "-"
    date_str = str(date_str).strip()
    if not date_str or date_str == "-":
        return "-"
    try:
        # Try YYYY-MM-DD
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%d.%m")
    except Exception:
        try:
            # Try DD.MM.YYYY
            dt = datetime.strptime(date_str, "%d.%m.%Y")
            return dt.strftime("%d.%m")
        except Exception:
            parts = date_str.split('.')
            if len(parts) >= 2:
                return f"{parts[0]}.{parts[1]}"
            return date_str

def format_time_no_seconds(time_str) -> str:
    if time_str is None:
        return "-"
    time_str = str(time_str).strip()
    if not time_str or time_str == "-":
        return "-"
    try:
        # Try HH:MM:SS
        dt = datetime.strptime(time_str, "%H:%M:%S")
        return dt.strftime("%H:%M")
    except Exception:
        try:
            # Try HH:MM
            dt = datetime.strptime(time_str, "%H:%M")
            return dt.strftime("%H:%M")
        except Exception:
            parts = time_str.split(':')
            if len(parts) >= 2:
                return f"{parts[0]}:{parts[1]}"
            return time_str


# ── Шифрование чувствительных настроек ───────────────────────────────────────
# Ключи, которые нельзя хранить в settings открытым текстом (секреты, токены доступа).
ENCRYPTED_SETTING_KEYS = {"google_service_account"}

_fernet = None

def _get_fernet():
    """Возвращает Fernet-инстанс для шифрования/расшифровки настроек.

    Ключ берётся из переменной окружения SETTINGS_ENCRYPTION_KEY (сгенерировать через
    `python -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"`
    и сохранить в .env). Если ключ не задан — шифрование выключено (с предупреждением в лог),
    чтобы не потерять доступ к существующим данным при первом запуске без настройки.
    """
    global _fernet
    if _fernet is not None:
        return _fernet
    key = os.environ.get("SETTINGS_ENCRYPTION_KEY")
    if not key:
        logger.warning(
            "SETTINGS_ENCRYPTION_KEY не задан — чувствительные настройки (%s) "
            "будут храниться БЕЗ шифрования. Задайте переменную окружения для продакшена.",
            ", ".join(ENCRYPTED_SETTING_KEYS),
        )
        _fernet = False
        return _fernet
    try:
        from cryptography.fernet import Fernet
        _fernet = Fernet(key.encode())
    except Exception:
        logger.exception("Не удалось инициализировать шифрование настроек, ключ SETTINGS_ENCRYPTION_KEY некорректен")
        _fernet = False
    return _fernet


def get_setting(key: str, default: str = None) -> str:
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    if not row:
        return default
    value = row["value"]
    if key in ENCRYPTED_SETTING_KEYS:
        fernet = _get_fernet()
        if fernet:
            try:
                return fernet.decrypt(value.encode()).decode()
            except Exception:
                # Значение могло быть сохранено до включения шифрования — отдаём как есть,
                # но логируем, чтобы было видно, что есть незашифрованные legacy-данные.
                logger.warning("Настройка %s не расшифрована (возможно, сохранена без шифрования)", key)
                return value
    return value


def set_setting(key: str, value: str):
    stored_value = value
    if key in ENCRYPTED_SETTING_KEYS:
        fernet = _get_fernet()
        if fernet:
            stored_value = fernet.encrypt(value.encode()).decode()
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, stored_value)
    )
    conn.commit()
    conn.close()


def fetch_export_data(dept: str = None, only_facts: bool = False):
    conn = get_db()
    try:
        # 1. Fetch active workers that match dept
        if dept is not None:
            workers = conn.execute("SELECT * FROM workers WHERE lower(position) = lower(?)", (dept,)).fetchall()
        else:
            workers = conn.execute("SELECT * FROM workers").fetchall()
            
        conditions = []
        params = []
        
        if dept is not None:
            conditions.append("lower(w.position) = lower(?)")
            params.append(dept)
            
        if only_facts:
            conditions.append("r.report_type = 'daily_fact'")
            
        where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""
        
        # We fetch relevant reports
        query = f"""
            SELECT r.id, r.telegram_id, r.report_date, r.report_type, r.slot_time, r.is_ok, r.format_comment, r.received_at
            FROM reports r
            LEFT JOIN workers w ON r.telegram_id = w.telegram_id
            {where_clause}
        """
        reports = conn.execute(query, params).fetchall()
        
        # Fetch first daily_fact date for each worker to determine when the fact requirement started
        first_fact_rows = conn.execute(
            "SELECT telegram_id, MIN(report_date) as first_date FROM reports WHERE report_type = 'daily_fact' GROUP BY telegram_id"
        ).fetchall()
        first_fact_dates = {row["telegram_id"]: row["first_date"] for row in first_fact_rows}
    finally:
        conn.close()
    
    if not reports:
        return None
        
    # Find list of unique dates and sort them chronologically
    unique_dates = sorted(list(set(r["report_date"] for r in reports)))
    
    # Map report keys: (telegram_id, date, slot/type)
    reports_map = {}
    for r in reports:
        tid = r["telegram_id"]
        r_type = r["report_type"]
        r_date = r["report_date"]
        
        if r_type == "daily_fact":
            key = (tid, r_date, "Факт")
        elif r_type == "not_working":
            key = (tid, r_date, "not_working")
        else:
            slot = r["slot_time"] or ""
            key = (tid, r_date, slot)
            
        reports_map[key] = {
            "is_ok": r["is_ok"],
            "format_comment": r["format_comment"],
            "received_at": r["received_at"]
        }
        
    # Include all workers we fetched
    workers_dict = {}
    for w in workers:
        workers_dict[w["telegram_id"]] = {
            "telegram_id": w["telegram_id"],
            "last_name": w["last_name"],
            "first_name": w["first_name"],
            "position": w["position"] or "Не указано",
            "schedule": w["schedule"] or "A",
            "needs_daily_fact": bool(w["needs_daily_fact"]),
            "is_active": bool(w["is_active"]),
            "sort_order": w["sort_order"] or 0,
        }
        
    # Also include any worker from reports who is not in workers_dict (e.g. deleted worker)
    for r in reports:
        tid = r["telegram_id"]
        if tid not in workers_dict:
            workers_dict[tid] = {
                "telegram_id": tid,
                "last_name": "Удаленный",
                "first_name": f"Сотрудник (ID {tid})",
                "position": "Не указано",
                "schedule": "A",
                "needs_daily_fact": True,
                "is_active": True,
                "sort_order": 99999,
            }
            
    # Group display_workers by dynamic department
    workers_by_dept = {}
    for tid, w in workers_dict.items():
        # Exclude test/placeholder objects
        lastname_lower = (w["last_name"] or "").lower()
        firstname_lower = (w["first_name"] or "").lower()
        dept_lower = (w["position"] or "").lower()
        if any(x in lastname_lower or x in firstname_lower or x in dept_lower for x in ("отмена", "test", "тест")):
            continue
        
        dept_name = w["position"]
        if dept_name not in workers_by_dept:
            workers_by_dept[dept_name] = []
        workers_by_dept[dept_name].append(w)
        
    sorted_depts = sorted(workers_by_dept.keys())
    for dept_name in sorted_depts:
        workers_by_dept[dept_name] = sorted(
            workers_by_dept[dept_name],
            key=lambda item: (item["sort_order"], (item["last_name"] or "").lower(), (item["first_name"] or "").lower())
        )
        
    return unique_dates, reports_map, sorted_depts, workers_by_dept, first_fact_dates


async def generate_and_send_excel(update: Update, context: ContextTypes.DEFAULT_TYPE, dept: str = None, only_facts: bool = False):
    await update.message.reply_text("⏳ Формирую выгрузку отчетов в формате Excel (.xlsx)...")
    
    data = fetch_export_data(dept, only_facts)
    if not data:
        criteria_msg = "для выгрузки по данному критерию."
        if only_facts:
            criteria_msg = "по фактам дня."
        await update.message.reply_text(f"В базе данных пока нет ни одного отчета {criteria_msg}", reply_markup=MAIN_MENU)
        return
        
    unique_dates, reports_map, sorted_depts, workers_by_dept, first_fact_dates = data
        
    # Start Excel Generation
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.comments import Comment
    import openpyxl.utils as o_utils
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    
    # Style declarations
    header_font = Font(name=font_family, size=11, bold=True, color="000000")
    dept_font = Font(name=font_family, size=11, bold=True, color="000000")
    worker_font = Font(name=font_family, size=11, bold=False, color="000000")
    slot_font = Font(name=font_family, size=10, bold=False, color="444444")
    ok_font = Font(name=font_family, size=12, bold=False, color="000000")
    fail_font = Font(name=font_family, size=12, bold=False, color="000000")
    
    # Fills
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    dept_fill = PatternFill(start_color="D6C9F5", end_color="D6C9F5", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="000000")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Alignments
    center_align = Alignment(vertical="center", horizontal="center")
    left_align_wrap = Alignment(vertical="center", horizontal="left", wrap_text=True)
    
    # Headers
    ws.cell(row=1, column=1, value="Сотрудник").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = grid_border
    
    ws.cell(row=1, column=2, value="Время").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = center_align
    ws.cell(row=1, column=2).border = grid_border
    
    for c_idx, date in enumerate(unique_dates, start=3):
        col_letter = o_utils.get_column_letter(c_idx)
        cell = ws.cell(row=1, column=c_idx, value=format_date_no_year(date))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = grid_border
        ws.column_dimensions[col_letter].width = 8
        
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.row_dimensions[1].height = 25
    
    ws.freeze_panes = "C2"
    
    curr_row = 2
    
    for dept_name in sorted_depts:
        dept_workers = workers_by_dept[dept_name]
        if not dept_workers:
            continue
            
        # Write Department Section header row
        ws.row_dimensions[curr_row].height = 22
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=len(unique_dates) + 2)
        cell = ws.cell(row=curr_row, column=1, value=f"{dept_name}")
        cell.font = dept_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
        
        for c_idx in range(1, len(unique_dates) + 3):
            c = ws.cell(row=curr_row, column=c_idx)
            c.fill = dept_fill
            c.border = grid_border
            
        curr_row += 1
        
        for w in dept_workers:
            schedule_slots = SCHEDULES.get(w["schedule"], SCHEDULE_A)
            if only_facts:
                worker_rows = ["Факт"]
            else:
                worker_rows = list(schedule_slots)
                if w["needs_daily_fact"]:
                    worker_rows.append("Факт")
                    
            num_rows = len(worker_rows)
            if num_rows == 0:
                continue
                
            name_cell = ws.cell(row=curr_row, column=1, value=f"{w['last_name']} {w['first_name']}")
            name_cell.font = worker_font
            name_cell.alignment = left_align_wrap
            name_cell.border = grid_border
            
            if num_rows > 1:
                ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row + num_rows - 1, end_column=1)
                
            for sub_idx, slot in enumerate(worker_rows):
                r_idx = curr_row + sub_idx
                ws.row_dimensions[r_idx].height = 20
                
                # Column B (Time slot)
                slot_cell = ws.cell(row=r_idx, column=2, value=slot)
                slot_cell.font = slot_font
                slot_cell.alignment = center_align
                slot_cell.border = grid_border
                
                # Column C onwards
                for d_idx, date in enumerate(unique_dates, start=3):
                    cell = ws.cell(row=r_idx, column=d_idx)
                    cell.alignment = center_align
                    cell.border = grid_border
                    
                    not_working_key = (w["telegram_id"], date, "not_working")
                    if not_working_key in reports_map:
                        cell.value = ""
                        cell.fill = gray_fill
                        reason_str = reports_map[not_working_key]["format_comment"] or "Не работает"
                        cell.comment = Comment(reason_str, "Прораб-Бот")
                    else:
                        rep_key = (w["telegram_id"], date, slot)
                        if rep_key in reports_map:
                            rep = reports_map[rep_key]
                            is_ok = bool(rep["is_ok"])
                            
                            if is_ok:
                                cell.value = "☑"
                                cell.font = ok_font
                            else:
                                cell.value = "☐"
                                cell.font = fail_font
                                cell.fill = fail_fill
                                
                                comment_str = rep["format_comment"] or "В отчете есть замечания"
                                if comment_str.startswith("не ОК, "):
                                    comment_str = comment_str[len("не ОК, "):]
                                elif comment_str.startswith("не ОК: "):
                                    comment_str = comment_str[len("не ОК: "):]
                                    
                                cell.comment = Comment(comment_str, "Прораб-Бот")
                        else:
                            is_hyphen = False
                            if slot == "Факт":
                                cur_date_str = now_local().strftime("%Y-%m-%d")
                                tid = w["telegram_id"]
                                if tid in first_fact_dates:
                                    if date < first_fact_dates[tid]:
                                        is_hyphen = True
                                else:
                                    if date < cur_date_str:
                                        is_hyphen = True
                            
                            if is_hyphen:
                                cell.value = "-"
                            else:
                                cell.value = ""
                        
            # Format and set borders inside merged cell parts
            for sub_idx in range(num_rows):
                r_idx = curr_row + sub_idx
                for c_idx in range(1, len(unique_dates) + 3):
                    ws.cell(row=r_idx, column=c_idx).border = grid_border
                    
            curr_row += num_rows

    # ── Создание второго листа: Аналитика ──
    try:
        ws_anal = wb.create_sheet(title="Аналитика")
        ws_anal.views.sheetView[0].showGridLines = True
        
        # Styles for Analytics
        anal_header_font = Font(name=font_family, size=11, bold=True, color="000000")
        anal_cell_font = Font(name=font_family, size=11, color="000000")
        
        # Determine the weeks present in unique_dates
        import datetime as dt_mod
        week_keys = set()
        for d_str in unique_dates:
            try:
                dt_val = dt_mod.datetime.strptime(d_str, "%Y-%m-%d").date()
                year, week, weekday = dt_val.isocalendar()
                week_keys.add((year, week))
            except Exception:
                continue
                
        sorted_weeks = sorted(list(week_keys))
        
        # Headers
        ws_anal.cell(row=1, column=1, value="Сотрудник").font = anal_header_font
        ws_anal.cell(row=1, column=1).alignment = center_align
        ws_anal.cell(row=1, column=1).border = grid_border
        
        ws_anal.cell(row=1, column=2, value="Отдел").font = anal_header_font
        ws_anal.cell(row=1, column=2).alignment = center_align
        ws_anal.cell(row=1, column=2).border = grid_border
        
        ws_anal.cell(row=1, column=3, value="Объект").font = anal_header_font
        ws_anal.cell(row=1, column=3).alignment = center_align
        ws_anal.cell(row=1, column=3).border = grid_border

        ws_anal.cell(row=1, column=4, value="Ожидалось (всего)").font = anal_header_font
        ws_anal.cell(row=1, column=4).alignment = center_align
        ws_anal.cell(row=1, column=4).border = grid_border

        ws_anal.cell(row=1, column=5, value="Сдано (всего)").font = anal_header_font
        ws_anal.cell(row=1, column=5).alignment = center_align
        ws_anal.cell(row=1, column=5).border = grid_border

        ws_anal.cell(row=1, column=6, value="Опозданий (всего)").font = anal_header_font
        ws_anal.cell(row=1, column=6).alignment = center_align
        ws_anal.cell(row=1, column=6).border = grid_border

        ws_anal.cell(row=1, column=7, value="Замечаний (всего)").font = anal_header_font
        ws_anal.cell(row=1, column=7).alignment = center_align
        ws_anal.cell(row=1, column=7).border = grid_border

        ws_anal.cell(row=1, column=8, value="Успешность (общая)").font = anal_header_font
        ws_anal.cell(row=1, column=8).alignment = center_align
        ws_anal.cell(row=1, column=8).border = grid_border
        
        # Week headers and map them to columns
        col_idx = 9
        week_columns = {} # (year, week): col_idx
        for (year, week) in sorted_weeks:
            # Get monday and sunday
            d = dt_mod.date(year, 1, 4)
            d = d + dt_mod.timedelta(weeks=week - 1)
            mon = d - dt_mod.timedelta(days=d.weekday())
            sun = mon + dt_mod.timedelta(days=6)
            
            mon_str = mon.strftime("%d.%m")
            sun_str = sun.strftime("%d.%m")
            header_val = f"Нед. {week:02d} ({mon_str}-{sun_str})"
            
            cell = ws_anal.cell(row=1, column=col_idx, value=header_val)
            cell.font = anal_header_font
            cell.alignment = center_align
            cell.border = grid_border
            
            week_columns[(year, week)] = col_idx
            col_idx += 1
            
        # Add worker rows
        conn_anal = get_db()
        row_idx = 2
        
        green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
        gray_light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for dept_name in sorted_depts:
            for worker in workers_by_dept[dept_name]:
                ws_anal.cell(row=row_idx, column=1, value=f"{worker['last_name']} {worker['first_name']}").font = anal_cell_font
                ws_anal.cell(row=row_idx, column=1).border = grid_border
                
                ws_anal.cell(row=row_idx, column=2, value=worker['position']).font = anal_cell_font
                ws_anal.cell(row=row_idx, column=2).border = grid_border
                
                ws_anal.cell(row=row_idx, column=3, value=worker.get('object_id', 'Основной')).font = anal_cell_font
                ws_anal.cell(row=row_idx, column=3).border = grid_border
                
                # Calculate overall stats
                overall_expected = 0
                overall_submitted = 0
                overall_lates = 0
                overall_remarks = 0
                overall_percent = 100.0
                if unique_dates:
                    try:
                        overall_stats = calculate_worker_stats(worker, unique_dates[0], unique_dates[-1], conn_anal)
                        overall_expected = overall_stats["expected"]
                        overall_submitted = overall_stats["submitted"]
                        overall_lates = overall_stats["lates"]
                        overall_remarks = overall_stats["remarks"]
                        overall_percent = overall_stats["percent"]
                    except Exception as e:
                        logger.error(f"Error calculating overall stats for worker {worker['telegram_id']}: {e}")
                
                ws_anal.cell(row=row_idx, column=4, value=overall_expected).font = anal_cell_font
                ws_anal.cell(row=row_idx, column=4).border = grid_border
                ws_anal.cell(row=row_idx, column=4).alignment = center_align

                ws_anal.cell(row=row_idx, column=5, value=overall_submitted).font = anal_cell_font
                ws_anal.cell(row=row_idx, column=5).border = grid_border
                ws_anal.cell(row=row_idx, column=5).alignment = center_align

                ws_anal.cell(row=row_idx, column=6, value=overall_lates).font = anal_cell_font
                ws_anal.cell(row=row_idx, column=6).border = grid_border
                ws_anal.cell(row=row_idx, column=6).alignment = center_align

                ws_anal.cell(row=row_idx, column=7, value=overall_remarks).font = anal_cell_font
                ws_anal.cell(row=row_idx, column=7).border = grid_border
                ws_anal.cell(row=row_idx, column=7).alignment = center_align

                c_overall = ws_anal.cell(row=row_idx, column=8, value=overall_percent / 100.0 if overall_expected > 0 else 1.0)
                c_overall.font = anal_cell_font
                c_overall.border = grid_border
                c_overall.alignment = center_align
                c_overall.number_format = '0.0%'
                if overall_expected > 0:
                    if overall_percent >= 90:
                        c_overall.fill = green_fill
                    elif overall_percent >= 70:
                        c_overall.fill = yellow_fill
                    else:
                        c_overall.fill = red_fill
                else:
                    c_overall.fill = gray_light_fill

                # Calculate for each week
                for (year, week), col_c in week_columns.items():
                    # calculate start and end dates
                    d = dt_mod.date(year, 1, 4)
                    d = d + dt_mod.timedelta(weeks=week - 1)
                    mon = d - dt_mod.timedelta(days=d.weekday())
                    sun = mon + dt_mod.timedelta(days=6)
                    
                    mon_str = mon.strftime("%Y-%m-%d")
                    sun_str = sun.strftime("%Y-%m-%d")
                    
                    cell = ws_anal.cell(row=row_idx, column=col_c)
                    cell.border = grid_border
                    cell.font = anal_cell_font
                    cell.alignment = center_align
                    
                    try:
                        stats = calculate_worker_stats(worker, mon_str, sun_str, conn_anal)
                        if stats["expected"] > 0:
                            percent_val = stats["percent"] / 100.0
                            cell.value = percent_val
                            cell.number_format = '0.0%'
                            
                            # Heatmap color scale
                            if stats["percent"] >= 90:
                                cell.fill = green_fill
                            elif stats["percent"] >= 70:
                                cell.fill = yellow_fill
                            else:
                                cell.fill = red_fill
                        else:
                            cell.value = "-"
                            cell.fill = gray_light_fill
                    except Exception as e:
                        logger.error(f"Error calculating weekly stats for worker {worker['telegram_id']} in week {year}-W{week}: {e}")
                        cell.value = "Ошибка"
                        
                row_idx += 1
                
        conn_anal.close()
        
        # Auto-adjust column widths for Analytics sheet
        for col in ws_anal.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format == '0.0%' and isinstance(cell.value, float):
                    val_str = f"{cell.value * 100:.1f}%"
                max_len = max(max_len, len(val_str))
            col_letter = o_utils.get_column_letter(col[0].column)
            ws_anal.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    except Exception as ex:
        logger.error(f"Не удалось сформировать лист Аналитика в Excel: {ex}")
            
    # Save Workbook to dynamic ByteIO
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    prefix = "facts" if only_facts else "reports"
    
    if dept is None:
        bio.name = f"{prefix}_all_{timestamp}.xlsx"
        caption = "📊 Общая выгрузка фактов дня в формате Excel успешно сформирована!" if only_facts else "📊 Общая выгрузка всех отчетов в формате Excel успешно сформирована!"
    else:
        safe_dept = "".join(c for c in dept if c.isalnum() or c in (" ", "_", "-")).strip()
        bio.name = f"{prefix}_dept_{safe_dept}_{timestamp}.xlsx"
        caption = f"📊 Выгрузка фактов дня для отдела «{dept}» в формате Excel успешно сформирована!" if only_facts else f"📊 Выгрузка отчетов для отдела «{dept}» в формате Excel успешно сформирована!"
        
    await context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=bio,
        filename=bio.name,
        caption=caption,
        reply_markup=MAIN_MENU
    )


gsheets_sync_lock = asyncio.Lock()

async def async_sync_gsheets_background():
    """Асинхронно запускает фоновую синхронизацию с Google Таблицей, если она настроена, используя глобальную блокировку."""
    spreadsheet_id = get_setting("google_spreadsheet_id")
    service_account_str = get_setting("google_service_account")
    if not spreadsheet_id or not service_account_str:
        return
    
    async with gsheets_sync_lock:
        data = fetch_export_data(None, False)
        if not data:
            return
            
        cur_date_str = now_local().strftime("%Y-%m-%d")
        try:
            await asyncio.to_thread(
                run_gsheets_sync,
                spreadsheet_id,
                service_account_str,
                None,
                False,
                data,
                cur_date_str
            )
            logger.info("Автоматическая синхронизация с Google Таблицами успешно выполнена!")
        except Exception as e:
            logger.error(f"Ошибка при автоматической синхронизации Google Таблиц: {e}")


def run_gsheets_sync(spreadsheet_id: str, service_account_str: str, dept: str, only_facts: bool, data, cur_date_str: str):
    if gspread is None or Credentials is None:
        raise ImportError("Библиотеки gspread или google-auth не установлены.")
    
    unique_dates, reports_map, sorted_depts, workers_by_dept, first_fact_dates = data
    
    creds_dict = json.loads(service_account_str)
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_key(spreadsheet_id)

    ws = None
    existing_map = {}
    existing_notes_map = {}
    existing_cell_colors = {}
    existing_column_colors = {}
    existing_manual_merges = []

    COLOR_HEADER = {"red": 1.0, "green": 1.0, "blue": 1.0}
    COLOR_DEPT = {"red": 0.839, "green": 0.788, "blue": 0.961}
    COLOR_FAIL = {"red": 0.988, "green": 0.894, "blue": 0.894}
    COLOR_GRAY = {"red": 0.902, "green": 0.902, "blue": 0.902}
    COLOR_BORDER = {"red": 0.0, "green": 0.0, "blue": 0.0}

    try:
        ws = spreadsheet.worksheet("Сводка")
        try:
            ws_id = ws.id
            meta = spreadsheet.fetch_sheet_metadata(params={
                "fields": "sheets(properties(title,sheetId),merges,data(rowData(values(userEnteredValue,userEnteredFormat,note,dataValidation))))"
            })
            ws_meta = None
            for s in meta.get("sheets", []):
                if s.get("properties", {}).get("title") == "Сводка":
                    ws_meta = s
                    break
            
            if ws_meta:
                grid_data_list = ws_meta.get("data", [])
                if grid_data_list:
                    row_data = grid_data_list[0].get("rowData", [])
                    
                    col_date_map = {}
                    if len(row_data) > 0:
                        header_values = row_data[0].get("values", [])
                        for col_idx in range(2, len(header_values)):
                            if col_idx >= len(header_values):
                                break
                            val_dict = header_values[col_idx]
                            header_val = ""
                            if "userEnteredValue" in val_dict:
                                header_val = str(val_dict["userEnteredValue"].get("stringValue", "")).strip()
                            for d in unique_dates:
                                if format_date_no_year(d) == header_val:
                                    col_date_map[col_idx] = d
                                    break
                        
                        for col_idx, date_str in col_date_map.items():
                            if col_idx < len(header_values):
                                header_cell = header_values[col_idx]
                                bg = header_cell.get("userEnteredFormat", {}).get("backgroundColor")
                                if bg and bg != {"red": 1.0, "green": 1.0, "blue": 1.0} and bg != {}:
                                    existing_column_colors[date_str] = bg
                    
                    row_mapping = {}
                    current_worker_name = None
                    
                    for row_idx in range(1, len(row_data)):
                        row_cells = row_data[row_idx].get("values", [])
                        if len(row_cells) < 2:
                            continue
                        
                        cell_a = row_cells[0] if len(row_cells) > 0 else {}
                        cell_b = row_cells[1] if len(row_cells) > 1 else {}
                        
                        val_a = ""
                        if "userEnteredValue" in cell_a:
                            val_a = str(cell_a["userEnteredValue"].get("stringValue", "")).strip()
                            
                        val_b = ""
                        if "userEnteredValue" in cell_b:
                            val_b = str(cell_b["userEnteredValue"].get("stringValue", "")).strip()
                        
                        if val_a != "":
                            if val_b == "":
                                row_mapping[row_idx] = {"type": "dept", "dept": val_a}
                                current_worker_name = None
                            else:
                                current_worker_name = val_a
                                row_mapping[row_idx] = {"type": "worker", "worker": val_a, "slot": val_b}
                        else:
                            if current_worker_name and val_b != "":
                                row_mapping[row_idx] = {"type": "worker", "worker": current_worker_name, "slot": val_b}
                    
                    for row_idx in range(1, len(row_data)):
                        row_cells = row_data[row_idx].get("values", [])
                        row_info = row_mapping.get(row_idx)
                        if not row_info:
                            continue
                        
                        for col_idx in range(2, len(row_cells)):
                            if col_idx not in col_date_map:
                                continue
                            date_str = col_date_map[col_idx]
                            cell_data = row_cells[col_idx]
                            
                            cell_val = ""
                            if "userEnteredValue" in cell_data:
                                val_dict = cell_data["userEnteredValue"]
                                if "stringValue" in val_dict:
                                    cell_val = val_dict["stringValue"]
                                elif "boolValue" in val_dict:
                                    cell_val = "TRUE" if val_dict["boolValue"] else "FALSE"
                                elif "numberValue" in val_dict:
                                    cell_val = str(val_dict["numberValue"])
                            
                            cell_note = cell_data.get("note", "").strip()
                            bg_color = cell_data.get("userEnteredFormat", {}).get("backgroundColor")
                            
                            if row_info["type"] == "worker":
                                w_name = row_info["worker"]
                                slot = row_info["slot"]
                                
                                if cell_val != "":
                                    existing_map[(w_name, date_str, slot)] = cell_val
                                
                                if cell_note != "":
                                    existing_notes_map[(w_name, date_str, slot)] = cell_note
                                    
                                if bg_color and bg_color != {"red": 1.0, "green": 1.0, "blue": 1.0} and bg_color != {}:
                                    existing_cell_colors[(w_name, date_str, slot)] = bg_color
                            
                            elif row_info["type"] == "dept":
                                dept_name = row_info["dept"]
                                if bg_color and bg_color != COLOR_DEPT and bg_color != {"red": 1.0, "green": 1.0, "blue": 1.0} and bg_color != {}:
                                    existing_cell_colors[(dept_name, date_str, "dept_header")] = bg_color
                                    
                    for m in ws_meta.get("merges", []):
                        r_start = m.get("startRowIndex", 0)
                        r_end = m.get("endRowIndex", 0)
                        c_start = m.get("startColumnIndex", 0)
                        c_end = m.get("endColumnIndex", 0)
                        
                        if c_start >= 2:
                            info_start = row_mapping.get(r_start)
                            info_end = row_mapping.get(r_end - 1)
                            
                            if c_start in col_date_map and (c_end - 1) in col_date_map:
                                date_start = col_date_map[c_start]
                                date_end = col_date_map[c_end - 1]
                                
                                if info_start and info_end:
                                    if info_start["type"] == "worker" and info_end["type"] == "worker" and info_start["worker"] == info_end["worker"]:
                                        existing_manual_merges.append({
                                            "type": "worker",
                                            "worker": info_start["worker"],
                                            "slot_start": info_start["slot"],
                                            "slot_end": info_end["slot"],
                                            "date_start": date_start,
                                            "date_end": date_end
                                        })
                                    elif info_start["type"] == "dept" and info_end["type"] == "dept" and info_start["dept"] == info_end["dept"]:
                                        existing_manual_merges.append({
                                            "type": "dept",
                                            "dept": info_start["dept"],
                                            "date_start": date_start,
                                            "date_end": date_end
                                        })
        except Exception as ex:
            logger.warning(f"Could not parse existing worksheet: {ex}")
    except gspread.exceptions.WorksheetNotFound:
        pass
        
    if not ws:
        ws = spreadsheet.add_worksheet(title="Сводка", rows="1000", cols="50")
        
    ws_id = ws.id

    num_cols = len(unique_dates) + 2
    values = []
    
    header_row = ["Сотрудник", "Время"] + [format_date_no_year(d) for d in unique_dates]
    values.append(header_row)
    
    requests = []
    curr_row = 1
    new_dept_rows = {}
    new_worker_rows = {}
    
    for dept_name in sorted_depts:
        dept_workers = workers_by_dept[dept_name]
        if not dept_workers:
            continue
            
        dept_row = [dept_name] + [""] * (len(unique_dates) + 1)
        values.append(dept_row)
        new_dept_rows[dept_name] = curr_row
        
        requests.append({
            "mergeCells": {
                "range": {
                    "sheetId": ws_id,
                    "startRowIndex": curr_row,
                    "endRowIndex": curr_row + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 2
                },
                "mergeType": "MERGE_ALL"
            }
        })
        
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": ws_id,
                    "startRowIndex": curr_row,
                    "endRowIndex": curr_row + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": COLOR_DEPT,
                        "textFormat": {
                            "fontFamily": "Segoe UI",
                            "fontSize": 11,
                            "bold": True
                        },
                        "horizontalAlignment": "LEFT"
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
            }
        })
        
        curr_row += 1
        
        for w in dept_workers:
            schedule_slots = SCHEDULES.get(w["schedule"], SCHEDULE_A)
            if only_facts:
                worker_rows = ["Факт"]
            else:
                worker_rows = list(schedule_slots)
                if w["needs_daily_fact"]:
                    worker_rows.append("Факт")
                    
            num_rows = len(worker_rows)
            if num_rows == 0:
                continue
                
            if num_rows > 1:
                requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": curr_row,
                            "endRowIndex": curr_row + num_rows,
                            "startColumnIndex": 0,
                            "endColumnIndex": 1
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": ws_id,
                        "startRowIndex": curr_row,
                        "endRowIndex": curr_row + num_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": 1
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "horizontalAlignment": "LEFT",
                            "wrapStrategy": "WRAP"
                        }
                    },
                    "fields": "userEnteredFormat(horizontalAlignment,wrapStrategy)"
                }
            })
            
            for sub_idx, slot in enumerate(worker_rows):
                r_idx = curr_row + sub_idx
                worker_full_name = f"{w['last_name']} {w['first_name']}".strip()
                new_worker_rows[(worker_full_name, slot)] = r_idx
                row_vals = []
                
                if sub_idx == 0:
                    row_vals.append(worker_full_name)
                else:
                    row_vals.append("")
                    
                row_vals.append(slot)
                
                for d_idx, date in enumerate(unique_dates, start=2):
                    not_working_key = (w["telegram_id"], date, "not_working")
                    if not_working_key in reports_map:
                        row_vals.append("")
                        reason_str = reports_map[not_working_key]["format_comment"] or "Не работает"
                        requests.append({
                            "repeatCell": {
                                "range": {
                                    "sheetId": ws_id,
                                    "startRowIndex": r_idx,
                                    "endRowIndex": r_idx + 1,
                                    "startColumnIndex": d_idx,
                                    "endColumnIndex": d_idx + 1
                                },
                                "cell": {
                                    "userEnteredFormat": {
                                        "backgroundColor": COLOR_GRAY
                                    },
                                    "note": reason_str
                                },
                                "fields": "userEnteredFormat.backgroundColor,note"
                            }
                        })
                        requests.append({
                            "setDataValidation": {
                                "range": {
                                    "sheetId": ws_id,
                                    "startRowIndex": r_idx,
                                    "endRowIndex": r_idx + 1,
                                    "startColumnIndex": d_idx,
                                    "endColumnIndex": d_idx + 1
                                }
                            }
                        })
                    else:
                        rep_key = (w["telegram_id"], date, slot)
                        sheet_key = (worker_full_name, date, slot)
                        existing_val = existing_map.get(sheet_key, "")
                        
                        if existing_val != "":
                            if existing_val == "TRUE":
                                row_vals.append(True)
                            elif existing_val == "FALSE":
                                row_vals.append(False)
                            else:
                                row_vals.append(existing_val)
                        else:
                            if rep_key in reports_map:
                                rep = reports_map[rep_key]
                                is_ok = bool(rep["is_ok"])
                                if is_ok:
                                    row_vals.append(True)
                                else:
                                    row_vals.append(False)
                                    
                                    comment_str = rep["format_comment"] or "В отчете есть замечания"
                                    if comment_str.startswith("не ОК, "):
                                        comment_str = comment_str[len("не ОК, "):]
                                    elif comment_str.startswith("не ОК: "):
                                        comment_str = comment_str[len("не ОК: "):]
                                        
                                    requests.append({
                                        "repeatCell": {
                                            "range": {
                                                "sheetId": ws_id,
                                                "startRowIndex": r_idx,
                                                "endRowIndex": r_idx + 1,
                                                "startColumnIndex": d_idx,
                                                "endColumnIndex": d_idx + 1
                                            },
                                            "cell": {
                                                "userEnteredFormat": {
                                                    "backgroundColor": COLOR_FAIL
                                                },
                                                "note": comment_str
                                            },
                                            "fields": "userEnteredFormat.backgroundColor,note"
                                        }
                                    })
                            else:
                                is_hyphen = False
                                if slot == "Факт":
                                    if w["telegram_id"] in first_fact_dates:
                                        if date < first_fact_dates[w["telegram_id"]]:
                                            is_hyphen = True
                                    else:
                                        if date < cur_date_str:
                                            is_hyphen = True
                                
                                if is_hyphen:
                                    row_vals.append("-")
                                    requests.append({
                                        "setDataValidation": {
                                            "range": {
                                                "sheetId": ws_id,
                                                "startRowIndex": r_idx,
                                                "endRowIndex": r_idx + 1,
                                                "startColumnIndex": d_idx,
                                                "endColumnIndex": d_idx + 1
                                            }
                                        }
                                    })
                                else:
                                    row_vals.append("")
                        
                values.append(row_vals)
                
                requests.append({
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": r_idx,
                            "endRowIndex": r_idx + 1,
                            "startColumnIndex": 2,
                            "endColumnIndex": num_cols
                        },
                        "rule": {
                            "condition": {
                                "type": "BOOLEAN"
                            },
                            "showCustomUi": True
                        }
                    }
                })
                
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": r_idx,
                            "endRowIndex": r_idx + 1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 2
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER"
                            }
                        },
                        "fields": "userEnteredFormat.horizontalAlignment"
                    }
                })
                
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": r_idx,
                            "endRowIndex": r_idx + 1,
                            "startColumnIndex": 2,
                            "endColumnIndex": num_cols
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER"
                            }
                        },
                        "fields": "userEnteredFormat.horizontalAlignment"
                    }
                })
                
            curr_row += num_rows

    total_rows = len(values)
    
    # Получаем исходные размеры листа для полной очистки перед форматированием
    orig_rows = 1000
    orig_cols = 50
    try:
        orig_rows = ws.row_count
        orig_cols = ws.col_count
    except Exception:
        pass
        
    max_rows = max(orig_rows, total_rows)
    max_cols = max(orig_cols, num_cols)

    try:
        ws.update(values=values, range_name="A1")
    except TypeError:
        ws.update("A1", values)
        
    full_requests = [
        {
            "setDataValidation": {
                "range": {
                    "sheetId": ws_id,
                    "startRowIndex": 0,
                    "endRowIndex": max_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": max_cols
                }
            }
        },
        {
            "unmergeCells": {
                "range": {
                    "sheetId": ws_id,
                    "startRowIndex": 0,
                    "endRowIndex": max_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": max_cols
                }
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": ws_id,
                    "startRowIndex": 0,
                    "endRowIndex": max_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": max_cols
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                        "textFormat": {
                            "fontFamily": "Segoe UI",
                            "fontSize": 10,
                            "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0},
                            "bold": False,
                            "italic": False,
                            "underline": False
                        },
                        "verticalAlignment": "MIDDLE",
                        "horizontalAlignment": "LEFT",
                        "borders": {}
                    },
                    "note": ""
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,verticalAlignment,horizontalAlignment,borders),note"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": ws_id,
                    "startRowIndex": 0,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {
                            "fontFamily": "Segoe UI",
                            "fontSize": 10,
                            "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0}
                        },
                        "verticalAlignment": "MIDDLE",
                        "borders": {
                            "top": {"style": "SOLID", "color": COLOR_BORDER},
                            "bottom": {"style": "SOLID", "color": COLOR_BORDER},
                            "left": {"style": "SOLID", "color": COLOR_BORDER},
                            "right": {"style": "SOLID", "color": COLOR_BORDER}
                        }
                    }
                },
                "fields": "userEnteredFormat(textFormat,verticalAlignment,borders)"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": ws_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": COLOR_HEADER,
                        "textFormat": {
                            "fontFamily": "Segoe UI",
                            "fontSize": 11,
                            "bold": True
                        },
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
            }
        },
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": ws_id,
                    "gridProperties": {
                        "frozenRowCount": 1,
                        "frozenColumnCount": 2,
                        "hideGridlines": False
                    },
                    "index": 0
                },
                "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount,gridProperties.hideGridlines,index"
            }
        },
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": ws_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": 1
                },
                "properties": {
                    "pixelSize": 250
                },
                "fields": "pixelSize"
            }
        },
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": ws_id,
                    "dimension": "COLUMNS",
                    "startIndex": 1,
                    "endIndex": 2
                },
                "properties": {
                    "pixelSize": 90
                },
                "fields": "pixelSize"
            }
        },
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": ws_id,
                    "dimension": "COLUMNS",
                    "startIndex": 2,
                    "endIndex": num_cols
                },
                "properties": {
                    "pixelSize": 80
                },
                "fields": "pixelSize"
            }
        }
    ] + requests

    restore_requests = []
    new_date_cols = {d: idx for idx, d in enumerate(unique_dates, start=2)}
    
    # 1. Column colors
    for date_str, bg in existing_column_colors.items():
        if date_str in new_date_cols:
            col_idx = new_date_cols[date_str]
            restore_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": ws_id,
                        "startRowIndex": 1,
                        "endRowIndex": total_rows,
                        "startColumnIndex": col_idx,
                        "endColumnIndex": col_idx + 1
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": bg
                        }
                    },
                    "fields": "userEnteredFormat.backgroundColor"
                }
            })
            
    # 2. Cell colors
    for (w_name, date_str, slot), bg in existing_cell_colors.items():
        if slot == "dept_header":
            if w_name in new_dept_rows and date_str in new_date_cols:
                r_idx = new_dept_rows[w_name]
                col_idx = new_date_cols[date_str]
                restore_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": r_idx,
                            "endRowIndex": r_idx + 1,
                            "startColumnIndex": col_idx,
                            "endColumnIndex": col_idx + 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": bg
                            }
                        },
                        "fields": "userEnteredFormat.backgroundColor"
                    }
                })
        else:
            if (w_name, slot) in new_worker_rows and date_str in new_date_cols:
                r_idx = new_worker_rows[(w_name, slot)]
                col_idx = new_date_cols[date_str]
                restore_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": r_idx,
                            "endRowIndex": r_idx + 1,
                            "startColumnIndex": col_idx,
                            "endColumnIndex": col_idx + 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": bg
                            }
                        },
                        "fields": "userEnteredFormat.backgroundColor"
                    }
                })
                
    # 3. Cell notes
    for (w_name, date_str, slot), note in existing_notes_map.items():
        if (w_name, slot) in new_worker_rows and date_str in new_date_cols:
            r_idx = new_worker_rows[(w_name, slot)]
            col_idx = new_date_cols[date_str]
            restore_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": ws_id,
                        "startRowIndex": r_idx,
                        "endRowIndex": r_idx + 1,
                        "startColumnIndex": col_idx,
                        "endColumnIndex": col_idx + 1
                    },
                    "cell": {
                        "note": note
                    },
                    "fields": "note"
                }
            })
            
    # 4. Merges
    for m in existing_manual_merges:
        if m["type"] == "worker":
            w_name = m["worker"]
            s_start = m["slot_start"]
            s_end = m["slot_end"]
            d_start = m["date_start"]
            d_end = m["date_end"]
            
            if (w_name, s_start) in new_worker_rows and (w_name, s_end) in new_worker_rows and d_start in new_date_cols and d_end in new_date_cols:
                r_start = new_worker_rows[(w_name, s_start)]
                r_end = new_worker_rows[(w_name, s_end)] + 1
                c_start = new_date_cols[d_start]
                c_end = new_date_cols[d_end] + 1
                
                restore_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": r_start,
                            "endRowIndex": r_end,
                            "startColumnIndex": c_start,
                            "endColumnIndex": c_end
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
        elif m["type"] == "dept":
            dept_name = m["dept"]
            d_start = m["date_start"]
            d_end = m["date_end"]
            
            if dept_name in new_dept_rows and d_start in new_date_cols and d_end in new_date_cols:
                r_idx = new_dept_rows[dept_name]
                c_start = new_date_cols[d_start]
                c_end = new_date_cols[d_end] + 1
                
                restore_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_id,
                            "startRowIndex": r_idx,
                            "endRowIndex": r_idx + 1,
                            "startColumnIndex": c_start,
                            "endColumnIndex": c_end
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })

    ws_anal = None
    try:
        ws_anal = spreadsheet.worksheet("Аналитика")
    except gspread.exceptions.WorksheetNotFound:
        pass
        
    if not ws_anal:
        try:
            ws_anal = spreadsheet.add_worksheet(title="Аналитика", rows="1000", cols="50")
        except Exception:
            ws_anal = None
            
    anal_formatting_requests = []
    anal_requests = []
    values_anal = []
    total_anal_rows = 0
    num_anal_cols = 0
    
    if ws_anal:
        import datetime as dt_mod
        week_keys = set()
        for d_str in unique_dates:
            try:
                dt_val = dt_mod.datetime.strptime(d_str, "%Y-%m-%d").date()
                year, week, weekday = dt_val.isocalendar()
                week_keys.add((year, week))
            except Exception:
                continue
                
        sorted_weeks = sorted(list(week_keys))
        
        anal_header = [
            "Сотрудник", "Отдел", "Объект", 
            "Ожидалось (всего)", "Сдано (всего)", "Опозданий (всего)", "Замечаний (всего)", "Успешность (общая)"
        ]
        for (year, week) in sorted_weeks:
            d = dt_mod.date(year, 1, 4)
            d = d + dt_mod.timedelta(weeks=week - 1)
            mon = d - dt_mod.timedelta(days=d.weekday())
            sun = mon + dt_mod.timedelta(days=6)
            mon_str = mon.strftime("%d.%m")
            sun_str = sun.strftime("%d.%m")
            anal_header.append(f"Нед. {week:02d} ({mon_str}-{sun_str})")
            
        values_anal.append(anal_header)
        
        conn_anal = get_db()
        anal_curr_row = 1
        
        COLOR_SUCCESS_GREEN = {"red": 0.886, "green": 0.941, "blue": 0.851}
        COLOR_SUCCESS_YELLOW = {"red": 1.0, "green": 0.949, "blue": 0.8}
        COLOR_FAIL_RED = {"red": 0.988, "green": 0.894, "blue": 0.894}
        COLOR_LIGHT_GRAY = {"red": 0.949, "green": 0.949, "blue": 0.949}
        
        for dept_name in sorted_depts:
            for worker in workers_by_dept[dept_name]:
                row_vals = [
                    f"{worker['last_name']} {worker['first_name']}",
                    worker["position"],
                    worker.get("object_id", "Основной")
                ]
                
                overall_expected = 0
                overall_submitted = 0
                overall_lates = 0
                overall_remarks = 0
                overall_percent = 100.0
                if unique_dates:
                    try:
                        overall_stats = calculate_worker_stats(worker, unique_dates[0], unique_dates[-1], conn_anal)
                        overall_expected = overall_stats["expected"]
                        overall_submitted = overall_stats["submitted"]
                        overall_lates = overall_stats["lates"]
                        overall_remarks = overall_stats["remarks"]
                        overall_percent = overall_stats["percent"]
                    except Exception as e:
                        logger.error(f"Error calculating overall stats for worker {worker['telegram_id']}: {e}")
                
                row_vals.extend([
                    overall_expected,
                    overall_submitted,
                    overall_lates,
                    overall_remarks,
                    overall_percent / 100.0 if overall_expected > 0 else 1.0
                ])
                
                target_fill = COLOR_LIGHT_GRAY
                if overall_expected > 0:
                    if overall_percent >= 90:
                        target_fill = COLOR_SUCCESS_GREEN
                    elif overall_percent >= 70:
                        target_fill = COLOR_SUCCESS_YELLOW
                    else:
                        target_fill = COLOR_FAIL_RED
                        
                anal_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_anal.id,
                            "startRowIndex": anal_curr_row,
                            "endRowIndex": anal_curr_row + 1,
                            "startColumnIndex": 7,
                            "endColumnIndex": 8
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": target_fill,
                                "numberFormat": {
                                    "type": "PERCENT",
                                    "pattern": "0.0%"
                                },
                                "horizontalAlignment": "CENTER"
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,numberFormat,horizontalAlignment)"
                    }
                })
                
                col_offset = 8
                for (year, week) in sorted_weeks:
                    d = dt_mod.date(year, 1, 4)
                    d = d + dt_mod.timedelta(weeks=week - 1)
                    mon = d - dt_mod.timedelta(days=d.weekday())
                    sun = mon + dt_mod.timedelta(days=6)
                    
                    mon_str = mon.strftime("%Y-%m-%d")
                    sun_str = sun.strftime("%Y-%m-%d")
                    
                    try:
                        stats = calculate_worker_stats(worker, mon_str, sun_str, conn_anal)
                        if stats["expected"] > 0:
                            percent_val = stats["percent"] / 100.0
                            row_vals.append(percent_val)
                            
                            week_fill = COLOR_LIGHT_GRAY
                            if stats["percent"] >= 90:
                                week_fill = COLOR_SUCCESS_GREEN
                            elif stats["percent"] >= 70:
                                week_fill = COLOR_SUCCESS_YELLOW
                            else:
                                week_fill = COLOR_FAIL_RED
                        else:
                            row_vals.append("-")
                            week_fill = COLOR_LIGHT_GRAY
                    except Exception as e:
                        logger.error(f"Error weekly stats in gsheets for worker {worker['telegram_id']} in week {year}-W{week}: {e}")
                        row_vals.append("Ошибка")
                        week_fill = COLOR_LIGHT_GRAY
                        
                    anal_requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": ws_anal.id,
                                "startRowIndex": anal_curr_row,
                                "endRowIndex": anal_curr_row + 1,
                                "startColumnIndex": col_offset,
                                "endColumnIndex": col_offset + 1
                            },
                            "cell": {
                                "userEnteredFormat": {
                                    "backgroundColor": week_fill,
                                    "horizontalAlignment": "CENTER"
                                }
                            },
                            "fields": "userEnteredFormat(backgroundColor,horizontalAlignment)"
                        }
                    })
                    
                    if isinstance(row_vals[-1], float):
                        anal_requests.append({
                            "repeatCell": {
                                "range": {
                                    "sheetId": ws_anal.id,
                                    "startRowIndex": anal_curr_row,
                                    "endRowIndex": anal_curr_row + 1,
                                    "startColumnIndex": col_offset,
                                    "endColumnIndex": col_offset + 1
                                },
                                "cell": {
                                    "userEnteredFormat": {
                                        "numberFormat": {
                                            "type": "PERCENT",
                                            "pattern": "0.0%"
                                        }
                                    }
                                },
                                "fields": "userEnteredFormat.numberFormat"
                            }
                        })
                    
                    col_offset += 1
                    
                values_anal.append(row_vals)
                anal_curr_row += 1
                
        conn_anal.close()
        
        try:
            ws_anal.update(values=values_anal, range_name="A1")
        except TypeError:
            ws_anal.update("A1", values_anal)
            
        total_anal_rows = len(values_anal)
        num_anal_cols = len(anal_header)
        
        orig_anal_rows = 1000
        orig_anal_cols = 50
        try:
            orig_anal_rows = ws_anal.row_count
            orig_anal_cols = ws_anal.col_count
        except Exception:
            pass
            
        max_anal_rows = max(orig_anal_rows, total_anal_rows)
        max_anal_cols = max(orig_anal_cols, num_anal_cols)
        
        anal_formatting_requests = [
            {
                "setDataValidation": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "startRowIndex": 0,
                        "endRowIndex": max_anal_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": max_anal_cols
                    }
                }
            },
            {
                "unmergeCells": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "startRowIndex": 0,
                        "endRowIndex": max_anal_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": max_anal_cols
                    }
                }
            },
            {
                "repeatCell": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "startRowIndex": 0,
                        "endRowIndex": max_anal_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": max_anal_cols
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                            "textFormat": {
                                "fontFamily": "Segoe UI",
                                "fontSize": 10,
                                "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0},
                                "bold": False,
                                "italic": False,
                                "underline": False
                            },
                            "verticalAlignment": "MIDDLE",
                            "horizontalAlignment": "LEFT",
                            "borders": {}
                        },
                        "note": ""
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,verticalAlignment,horizontalAlignment,borders),note"
                }
            },
            {
                "repeatCell": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "startRowIndex": 0,
                        "endRowIndex": total_anal_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": num_anal_cols
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "textFormat": {
                                "fontFamily": "Segoe UI",
                                "fontSize": 10,
                                "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0}
                            },
                            "verticalAlignment": "MIDDLE",
                            "borders": {
                                "top": {"style": "SOLID", "color": COLOR_BORDER},
                                "bottom": {"style": "SOLID", "color": COLOR_BORDER},
                                "left": {"style": "SOLID", "color": COLOR_BORDER},
                                "right": {"style": "SOLID", "color": COLOR_BORDER}
                            }
                        }
                    },
                    "fields": "userEnteredFormat(textFormat,verticalAlignment,borders)"
                }
            },
            {
                "repeatCell": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "startRowIndex": 0,
                        "endRowIndex": 1,
                        "startColumnIndex": 0,
                        "endColumnIndex": num_anal_cols
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 0.95, "green": 0.95, "blue": 0.95},
                            "textFormat": {
                                "fontFamily": "Segoe UI",
                                "fontSize": 11,
                                "bold": True
                            },
                            "horizontalAlignment": "CENTER"
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": ws_anal.id,
                        "gridProperties": {
                            "frozenRowCount": 1,
                            "frozenColumnCount": 3,
                            "hideGridlines": False
                        },
                        "index": 1,
                        "title": "Аналитика"
                    },
                    "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount,gridProperties.hideGridlines,index,title"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "dimension": "COLUMNS",
                        "startIndex": 0,
                        "endIndex": 1
                    },
                    "properties": {
                        "pixelSize": 250
                    },
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "dimension": "COLUMNS",
                        "startIndex": 1,
                        "endIndex": 3
                    },
                    "properties": {
                        "pixelSize": 150
                    },
                    "fields": "pixelSize"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "dimension": "COLUMNS",
                        "startIndex": 3,
                        "endIndex": 8
                    },
                    "properties": {
                        "pixelSize": 120
                    },
                    "fields": "pixelSize"
                }
            },
            {
                "repeatCell": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "startRowIndex": 1,
                        "endRowIndex": total_anal_rows,
                        "startColumnIndex": 3,
                        "endColumnIndex": 8
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "horizontalAlignment": "CENTER"
                        }
                    },
                    "fields": "userEnteredFormat.horizontalAlignment"
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": ws_anal.id,
                        "dimension": "COLUMNS",
                        "startIndex": 8,
                        "endIndex": num_anal_cols
                    },
                    "properties": {
                        "pixelSize": 130
                    },
                    "fields": "pixelSize"
                }
            }
        ]

    full_requests = full_requests + restore_requests + anal_formatting_requests + anal_requests
    
    spreadsheet.batch_update({"requests": full_requests})
    
    try:
        ws.resize(rows=total_rows, cols=num_cols)
    except Exception as resize_ex:
        logger.warning(f"Failed to resize worksheet: {resize_ex}")
        
    if ws_anal and total_anal_rows > 0:
        try:
            ws_anal.resize(rows=total_anal_rows, cols=num_anal_cols)
        except Exception as resize_ex:
            logger.warning(f"Failed to resize Analytics worksheet: {resize_ex}")
    
    for sheet in spreadsheet.worksheets():
        if sheet.title in ("Лист1", "Sheet1") and sheet.id != ws_id and (not ws_anal or sheet.id != ws_anal.id):
            try:
                spreadsheet.del_worksheet(sheet)
            except Exception:
                pass


async def generate_and_send_gsheets(update: Update, context: ContextTypes.DEFAULT_TYPE, dept: str = None, only_facts: bool = False):
    spreadsheet_id = get_setting("google_spreadsheet_id")
    service_account_str = get_setting("google_service_account")
    
    if not spreadsheet_id or not service_account_str:
        await update.message.reply_text(
            "⚠️ **Настройки Google Sheets не найдены!**\n\n"
            "Пожалуйста, настройте интеграцию с помощью меню *⚙️ Настроить Google Таблицу*.",
            reply_markup=MAIN_MENU,
            parse_mode="Markdown"
        )
        return

    await update.message.reply_text("⏳ Формирую выгрузку отчетов в вашу Google Таблицу...")
    
    data = fetch_export_data(dept, only_facts)
    if not data:
        criteria_msg = "для выгрузки по данному критерию."
        if only_facts:
            criteria_msg = "по фактам дня."
        await update.message.reply_text(f"В базе данных пока нет ни одного отчета {criteria_msg}", reply_markup=MAIN_MENU)
        return
        
    cur_date_str = now_local().strftime("%Y-%m-%d")
    
    try:
        await asyncio.to_thread(
            run_gsheets_sync,
            spreadsheet_id,
            service_account_str,
            dept,
            only_facts,
            data,
            cur_date_str
        )
    except Exception as e:
        logger.exception("Failed to update Google Sheets")
        await update.message.reply_text(
            f"❌ **Ошибка при заполнении Google Таблицы:**\n`{str(e)}`",
            reply_markup=MAIN_MENU
        )
        return

    prefix = "фактов дня" if only_facts else "отчетов"
    caption = f"🟢 Общая выгрузка {prefix} в Google Таблицу успешно выполнена!" if dept is None else f"🟢 Выгрузка {prefix} для отдела «{dept}» в Google Таблицу успешно выполнена!"
    
    sheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    await update.message.reply_text(
        f"{caption}\n\n🔗 **Ссылка на таблицу:**\n{sheet_url}",
        reply_markup=MAIN_MENU,
        parse_mode="Markdown"
    )


async def export_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update):
        return ConversationHandler.END
        
    kbd = ReplyKeyboardMarkup([
        ["📊 Excel (.xlsx)", "🟢 Google Таблица"],
        ["⚙️ Настроить Google Таблицу"],
        ["❌ Отмена"]
    ], resize_keyboard=True)
    
    await update.message.reply_text(
        "Выберите формат экспорта отчетов:",
        reply_markup=kbd
    )
    return ASK_EXPORT_FORMAT


async def export_format_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choice = update.message.text.strip()
    if choice == "❌ Отмена":
        await update.message.reply_text("Выгрузка отменена.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    if choice == "📊 Excel (.xlsx)":
        context.user_data["export_format"] = "excel"
        kbd = ReplyKeyboardMarkup([
            ["📊 Общая выгрузка", "🏢 Выгрузка по отделу"],
            ["❌ Отмена"]
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Выберите тип выгрузки в Excel:",
            reply_markup=kbd
        )
        return ASK_EXPORT_TYPE
        
    if choice == "🟢 Google Таблица":
        context.user_data["export_format"] = "gsheets"
        
        spreadsheet_id = get_setting("google_spreadsheet_id")
        service_account_str = get_setting("google_service_account")
        
        if not spreadsheet_id or not service_account_str:
            await update.message.reply_text(
                "⚠️ **Интеграция с Google Таблицами не настроена!**\n\n"
                "Для работы интеграции:\n"
                "1. Настройте интеграцию с помощью кнопки *⚙️ Настроить Google Таблицу*.\n"
                "2. Укажите ID таблицы и JSON-ключ сервисного аккаунта.\n"
                "3. Дайте права Редактора аккаунту Google.",
                reply_markup=MAIN_MENU,
                parse_mode="Markdown"
            )
            return ConversationHandler.END
            
        kbd = ReplyKeyboardMarkup([
            ["📊 Общая выгрузка", "🏢 Выгрузка по отделу"],
            ["❌ Отмена"]
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Выберите тип выгрузки в Google Таблицу:",
            reply_markup=kbd
        )
        return ASK_EXPORT_TYPE
        
    if choice == "⚙️ Настроить Google Таблицу":
        spreadsheet_id = get_setting("google_spreadsheet_id", "Не задан")
        service_account_str = get_setting("google_service_account")
        email = "Не задан"
        if service_account_str:
            try:
                email = json.loads(service_account_str).get("client_email", "Не задан")
            except Exception:
                pass
                
        await update.message.reply_text(
            f"⚙️ **Текущие настройки Google Таблиц:**\n\n"
            f"🔗 **ID таблицы:** `{spreadsheet_id}`\n"
            f"📧 **Сервисный аккаунт:** `{email}`\n\n"
            f"Пришлите полную ссылку на вашу Google Таблицу или её ID, чтобы настроить или изменить интеграцию.\n\n"
            f"💡 *Вы можете нажать кнопку ниже для отмены.*",
            reply_markup=CANCEL_KEYBOARD,
            parse_mode="Markdown"
        )
        return ASK_GSHEETS_URL
        
    await update.message.reply_text("Пожалуйста, выберите один из вариантов на клавиатуре.")
    return ASK_EXPORT_FORMAT


async def export_gsheets_url_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text == "❌ Отмена":
        await update.message.reply_text("Настройка отменена.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", text)
    if match:
        spreadsheet_id = match.group(1)
    else:
        spreadsheet_id = text
        
    if not re.match(r"^[a-zA-Z0-9-_]{10,}$", spreadsheet_id):
        await update.message.reply_text(
            "⚠️ **Некорректный формат!**\n\n"
            "Пришлите полную ссылку на Google Таблицу (например: `https://docs.google.com/spreadsheets/d/...`) или её ID."
        )
        return ASK_GSHEETS_URL
        
    context.user_data["temp_spreadsheet_id"] = spreadsheet_id
    
    await update.message.reply_text(
        "🔗 **Ссылка на таблицу принята!**\n\n"
        "Теперь отправьте файл `service_account.json` вашего сервисного аккаунта Google (или скопируйте и пришлите его текст JSON).\n\n"
        "💡 *Если вы хотите отменить настройку, нажмите кнопку ниже.*",
        reply_markup=CANCEL_KEYBOARD,
        parse_mode="Markdown"
    )
    return ASK_GSHEETS_CREDS


async def export_gsheets_creds_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document:
        doc = update.message.document
        if not doc.file_name.endswith('.json'):
            await update.message.reply_text("⚠️ Пожалуйста, загрузите JSON-файл ключа сервисного аккаунта.")
            return ASK_GSHEETS_CREDS
        file = await doc.get_file()
        file_bytes = await file.download_as_bytearray()
        raw_json = file_bytes.decode('utf-8')
    else:
        raw_json = update.message.text.strip()
        
    if raw_json == "❌ Отмена":
        await update.message.reply_text("Настройка отменена.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    try:
        creds_dict = json.loads(raw_json)
        if creds_dict.get("type") != "service_account" or "client_email" not in creds_dict:
            raise ValueError("Файл не является корректным ключом сервисного аккаунта Google (отсутствует type: service_account или client_email).")
            
        set_setting("google_service_account", raw_json)
        
        temp_id = context.user_data.get("temp_spreadsheet_id")
        if temp_id:
            set_setting("google_spreadsheet_id", temp_id)
            context.user_data.pop("temp_spreadsheet_id", None)
            
        client_email = creds_dict.get("client_email")
        
        await update.message.reply_text(
            f"✅ **Интеграция с Google Таблицами успешно настроена!**\n\n"
            f"📧 **Важно:** Обязательно предоставьте доступ (права Редактора) следующему сервисному аккаунту:\n\n"
            f"`{client_email}`\n\n"
            f"Без этого шага бот не сможет записывать данные в таблицу!",
            reply_markup=MAIN_MENU,
            parse_mode="Markdown"
        )
        return ConversationHandler.END
    except Exception as e:
        logger.exception("Failed to parse service account JSON")
        await update.message.reply_text(
            f"❌ **Ошибка при разборе JSON-ключа:**\n`{str(e)}`\n\n"
            f"Пожалуйста, пришлите корректный JSON-файл ключа сервисного аккаунта."
        )
        return ASK_GSHEETS_CREDS


async def export_type_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choice = update.message.text.strip()
    if choice == "❌ Отмена":
        await update.message.reply_text("Выгрузка отменена.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    fmt = context.user_data.get("export_format", "excel")
    
    if choice == "📊 Общая выгрузка":
        if fmt == "excel":
            await generate_and_send_excel(update, context, dept=None)
        else:
            await generate_and_send_gsheets(update, context, dept=None)
        return ConversationHandler.END
        
    if choice == "🏢 Выгрузка по отделу":
        conn = get_db()
        rows = conn.execute("SELECT DISTINCT position FROM workers ORDER BY position").fetchall()
        conn.close()
        
        depts = [r["position"] for r in rows if r["position"] and r["position"] != "Не указано"]
        if not depts:
            await update.message.reply_text(
                "В базе данных нет зарегистрированных отделов у сотрудников.",
                reply_markup=MAIN_MENU
            )
            return ConversationHandler.END
            
        kbd_rows = [[d] for d in depts]
        kbd_rows.append(["❌ Отмена"])
        kbd = ReplyKeyboardMarkup(kbd_rows, resize_keyboard=True)
        
        await update.message.reply_text(
            "Выберите отдел для выгрузки:",
            reply_markup=kbd
        )
        return ASK_EXPORT_DEPARTMENT
        
    await update.message.reply_text("Пожалуйста, выберите один из вариантов на клавиатуре.")
    return ASK_EXPORT_TYPE


async def export_department_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choice = update.message.text.strip()
    if choice == "❌ Отмена":
        await update.message.reply_text("Выгрузка отменена.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    fmt = context.user_data.get("export_format", "excel")
    if fmt == "excel":
        await generate_and_send_excel(update, context, dept=choice)
    else:
        await generate_and_send_gsheets(update, context, dept=choice)
    return ConversationHandler.END


async def check_missing_reports_job(context: ContextTypes.DEFAULT_TYPE):
    now = now_local()

    # Напоминания о статусах нужны только в будние дни (пн-пт).
    # weekday(): 0=понедельник ... 5=субота, 6=воскресенье
    if now.weekday() >= 5:
        return

    date_str = now.strftime("%Y-%m-%d")
    current_mins = now.hour * 60 + now.minute
    
    conn = get_db()
    workers = conn.execute("SELECT * FROM workers WHERE is_active = 1").fetchall()
    
    reports = conn.execute(
        "SELECT telegram_id, slot_time FROM reports WHERE report_date = ? AND report_type = 'status'",
        (date_str,)
    ).fetchall()
    submitted_worker_slots = {(r["telegram_id"], r["slot_time"]) for r in reports}
    
    # Исключаем сотрудников, у которых сегодня выходной/не работают
    not_working_reports = conn.execute(
        "SELECT telegram_id FROM reports WHERE report_date = ? AND report_type = 'not_working'",
        (date_str,)
    ).fetchall()
    not_working_worker_ids = {r["telegram_id"] for r in not_working_reports}
    
    sent = conn.execute(
        "SELECT telegram_id, slot_time FROM sent_reminders WHERE report_date = ?",
        (date_str,)
    ).fetchall()
    sent_worker_slots = {(s["telegram_id"], s["slot_time"]) for s in sent}

    sent_pre = conn.execute(
        "SELECT telegram_id, slot_time FROM sent_pre_reminders WHERE report_date = ?",
        (date_str,)
    ).fetchall()
    sent_pre_worker_slots = {(s["telegram_id"], s["slot_time"]) for s in sent_pre}
    
    for w in workers:
        tid = w["telegram_id"]
        if tid in not_working_worker_ids:
            continue
            
        sched_slots = SCHEDULES.get(w["schedule"], SCHEDULE_A)
        
        for slot in sched_slots:
            try:
                sh, sm = map(int, slot.split(":"))
                slot_mins = sh * 60 + sm
            except Exception:
                continue
                
            # 1. Предварительное напоминание (за 10 минут до слота)
            if current_mins >= slot_mins - 12 and current_mins <= slot_mins - 6:
                if (tid, slot) not in submitted_worker_slots and (tid, slot) not in sent_pre_worker_slots:
                    conn.execute(
                        "INSERT OR IGNORE INTO sent_pre_reminders (telegram_id, report_date, slot_time) VALUES (?, ?, ?)",
                        (tid, date_str, slot)
                    )
                    conn.commit()
                    
                    try:
                        await context.bot.send_message(
                            chat_id=tid,
                            text=f"🔔 *Внимание!* Через 10 минут нужно отправить отчёт за *{slot}*.",
                            parse_mode="Markdown"
                        )
                    except Exception as e:
                        logger.warning(f"Ошибка личного предварительного уведомления {tid} за слот {slot}: {e}")

            # 2. Опоздание / Забытый отчет (через 30 минут после слота)
            if current_mins >= slot_mins + 25 and current_mins <= slot_mins + 55:
                if (tid, slot) not in submitted_worker_slots and (tid, slot) not in sent_worker_slots:
                    conn.execute(
                        "INSERT OR IGNORE INTO sent_reminders (telegram_id, report_date, slot_time) VALUES (?, ?, ?)",
                        (tid, date_str, slot)
                    )
                    conn.commit()
                    
                    try:
                        await context.bot.send_message(
                            chat_id=tid,
                            text=f"⏰ *Напоминание!* Вы забыли отправить отчет за слот *{slot}* в систему.\n\nПожалуйста, отправьте его прямо сейчас голосовым сообщением, кружком, видео или текстом.",
                            parse_mode="Markdown"
                        )
                    except Exception as e:
                        logger.warning(f"Ошибка личного уведомления о пропуске слота {tid}: {e}")
                    
                    if not is_quiet_mode_enabled():
                        group_id = w["group_id"] or DEFAULT_GROUP_ID
                        try:
                            await context.bot.send_message(
                                chat_id=group_id,
                                text=f"⚠️ *{w['last_name']} {w['first_name']}* не предоставил вовремя отчет за статус *{slot}*.",
                                parse_mode="Markdown"
                            )
                        except Exception as e:
                            logger.warning(f"Ошибка отправки предупреждения нарушителя в группу {group_id}: {e}")
                        
    conn.close()


async def remind_all_missing(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return
    
    now = now_local()
    current_minutes = now.hour * 60 + now.minute
    date_str = now.strftime("%Y-%m-%d")
    
    conn = get_db()
    workers = conn.execute("SELECT * FROM workers WHERE is_active = 1").fetchall()
    reports = conn.execute(
        "SELECT telegram_id, slot_time FROM reports WHERE report_date = ? AND report_type = 'status'",
        (date_str,)
    ).fetchall()
    conn.close()
    
    submitted_worker_slots = {(r["telegram_id"], r["slot_time"]) for r in reports}
    reminded_count = 0
    
    for w in workers:
        sched_slots = SCHEDULES.get(w["schedule"], SCHEDULE_A)
        tid = w["telegram_id"]
        overdue_slots = []
        
        for slot in sched_slots:
            hour, minute = map(int, slot.split(":"))
            slot_minutes = hour * 60 + minute
            if slot_minutes <= current_minutes:
                if (tid, slot) not in submitted_worker_slots:
                    overdue_slots.append(slot)
                    
        if overdue_slots:
            slots_str = ", ".join(overdue_slots)
            try:
                # Напоминание сотруднику
                await context.bot.send_message(
                    chat_id=tid,
                    text=f"⏰ Срочное напоминание! Вы пропустили отправку отчетов за слоты (сегменты): **{slots_str}**.\n\nПожалуйста, немедленно отправьте отчет в бот!",
                    parse_mode="Markdown"
                )
                reminded_count += 1
                
                # Сообщение в чат бригады
                if not is_quiet_mode_enabled():
                    group_id = w["group_id"] or DEFAULT_GROUP_ID
                    await context.bot.send_message(
                        chat_id=group_id,
                        text=f"⚠️ {w['last_name']} {w['first_name']} пропустил отправку отчетов за слоты: {slots_str}. Напоминание направлено в ЛС."
                    )
            except Exception as e:
                logger.warning(f"Ошибка ручной отправки напоминаний для {tid}: {e}")
                
    await update.message.reply_text(
        f"✅ Проверка завершена. Разослано напоминаний должникам: {reminded_count}.",
        reply_markup=MAIN_MENU
    )


# ══════════════════════════════════════════════════════════════════════════════
# Настройка расписания автоматической сводки (Время сводки)
# ══════════════════════════════════════════════════════════════════════════════

async def summary_time_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_admin(update): return ConversationHandler.END
    times = get_scheduled_times()
    times_str = ", ".join(times) if times else "не настроено"
    
    text = (
        f"⏰ Текущее автоматическое расписание сводки: {times_str}\n\n"
        f"Выберите действие:"
    )
    kbd = ReplyKeyboardMarkup([
        ["➕ Добавить время", "➖ Удалить время"],
        ["❌ Назад"]
    ], resize_keyboard=True)
    
    await update.message.reply_text(text, reply_markup=kbd)
    return ASK_REPORT_TIME

async def summary_time_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    action = update.message.text.strip()
    
    if action == "❌ Назад":
        await update.message.reply_text("Возврат в главное меню.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    if action == "➕ Добавить время":
        await update.message.reply_text(
            "Введите новое время в формате ЧЧ:ММ (например, 19:30):",
            reply_markup=CANCEL_KEYBOARD
        )
        return ASK_EDIT_SCHEDULE
        
    if action == "➖ Удалить время":
        times = get_scheduled_times()
        if not times:
            await update.message.reply_text("В расписании пока нет сохраненного времени сводок.", reply_markup=MAIN_MENU)
            return ConversationHandler.END
            
        kbd = ReplyKeyboardMarkup([[t] for t in times] + [["❌ Отмена"]], resize_keyboard=True)
        await update.message.reply_text(
            "Выберите время для удаления:",
            reply_markup=kbd
        )
        return ASK_ORDER_DEPARTMENT
        
    await update.message.reply_text("Пожалуйста, нажмите на одну из предложенных кнопок.")
    return ASK_REPORT_TIME

async def summary_time_add_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    parts = raw.split(":")
    if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
        await update.message.reply_text("Некорректный формат времени. Попробуйте еще раз (ЧЧ:ММ, например, 19:30):")
        return ASK_EDIT_SCHEDULE
        
    hour, minute = int(parts[0]), int(parts[1])
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        await update.message.reply_text("Неподходящие часы/минуты. Диапазоны: 00-23 и 00-59:")
        return ASK_EDIT_SCHEDULE
        
    time_str = f"{hour:02d}:{minute:02d}"
    times = get_scheduled_times()
    if time_str in times:
        await update.message.reply_text("Это время уже содержится в расписании сводки.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    times.append(time_str)
    save_scheduled_times(times)
    reschedule_summary_jobs(context.application)
    
    await update.message.reply_text(f"✅ Время {time_str} успешно внесено в расписание!", reply_markup=MAIN_MENU)
    return ConversationHandler.END

async def summary_time_del_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == "❌ Отмена":
        await update.message.reply_text("Удаление отменено.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    times = get_scheduled_times()
    if raw not in times:
        await update.message.reply_text("Пожалуйста, выберите существующий элемент из списка:")
        return ASK_ORDER_DEPARTMENT
        
    times.remove(raw)
    save_scheduled_times(times)
    reschedule_summary_jobs(context.application)
    
    await update.message.reply_text(f"✅ Время сводки {raw} успешно удалено из расписания!", reply_markup=MAIN_MENU)
    return ConversationHandler.END

async def conversation_timeout_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keys_to_clear = [
        "awaiting_not_working_confirm", "not_working_reason", "awaiting_not_working_reason",
        "depts_cache", "editing_comment_report_id", "editing_comment_chat_id", 
        "editing_comment_message_id", "editing_comment_original_text", "editing_comment_prompt_message_id",
        "add_worker_telegram_id", "add_worker_last_name", "add_worker_first_name", 
        "add_worker_position", "add_worker_group_id", "add_worker_schedule", "add_worker_needs_daily_fact",
        "del_worker_position", "del_worker_list", "del_worker_idx",
        "edit_worker_id", "edit_worker_field",
        "export_type", "export_dept", "export_format"
    ]
    for key in keys_to_clear:
        context.user_data.pop(key, None)
        
    logger.info("Conversation timed out. user_data cleared.")
    if update and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "⏳ Время ожидания ответа истекло. Диалог автоматически завершен.",
                reply_markup=menu_for_user(update.effective_user.id if update.effective_user else 0)
            )
        except Exception:
            pass

# ══════════════════════════════════════════════════════════════════════════════
# Самостоятельная регистрация сотрудников
# ══════════════════════════════════════════════════════════════════════════════

async def register_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_type = update.effective_chat.type
    if chat_type != "private":
        return ConversationHandler.END
        
    user_id = update.effective_user.id
    if is_admin(user_id):
        await update.message.reply_text("Привет! Выберите действие кнопкой ниже.", reply_markup=MAIN_MENU)
        return ConversationHandler.END
        
    worker = await run_db(get_worker, user_id)
    if worker:
        await update.message.reply_text(
            f"Привет! Отправьте видеоотчет, когда он будет готов.",
            reply_markup=menu_for_user(user_id, chat_type)
        )
        return ConversationHandler.END
        
    # Не зарегистрирован
    await update.message.reply_text(
        "👋 *Добро пожаловать в систему сдачи отчетов!*\n\n"
        "Вы не зарегистрированы в системе.\n"
        "Пожалуйста, введите вашу **Фамилию** для поиска в списке сотрудников:",
        parse_mode="Markdown",
        reply_markup=CANCEL_KEYBOARD
    )
    return ASK_REG_LAST_NAME

async def register_lastname_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user_id = update.effective_user.id
    if text == CANCEL_TEXT:
        await update.message.reply_text("Регистрация отменена.", reply_markup=menu_for_user(user_id))
        return ConversationHandler.END
        
    username = update.effective_user.username or "нет"
    first_name_tg = update.effective_user.first_name or ""
    last_name_tg = update.effective_user.last_name or ""
    
    workers = find_unregistered_workers_by_lastname(text)
    
    if len(workers) == 0:
        # Уведомляем администраторов
        last_name_tg_esc = html.escape(last_name_tg)
        first_name_tg_esc = html.escape(first_name_tg)
        username_esc = html.escape(username)
        text_esc = html.escape(text)
        admin_msg = (
            f"⚠️ <b>Неизвестный пользователь пытался зарегистрироваться:</b>\n\n"
            f"ФИО в TG: <b>{last_name_tg_esc} {first_name_tg_esc}</b>\n"
            f"Никнейм: @{username_esc}\n"
            f"Telegram ID: <code>{user_id}</code>\n"
            f"Введенная фамилия: <b>{text_esc}</b>"
        )
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=admin_msg, parse_mode="HTML")
            except Exception:
                pass
                
        # Также отправляем в SUMMARY_CHAT_ID
        if SUMMARY_CHAT_ID and SUMMARY_CHAT_ID not in ADMIN_IDS:
            try:
                await context.bot.send_message(chat_id=SUMMARY_CHAT_ID, text=admin_msg, parse_mode="HTML")
            except Exception:
                pass
                
        await update.message.reply_text(
            f"❌ Сотрудник с фамилией *{text}* не найден среди незарегистрированных в базе данных.\n\n"
            "Администраторы уведомлены о вашей попытке регистрации. Они свяжутся с вами или добавят в базу.",
            parse_mode="Markdown",
            reply_markup=menu_for_user(user_id)
        )
        return ConversationHandler.END
        
    elif len(workers) == 1:
        # Один кандидат найден
        candidate = workers[0]
        old_id = candidate["telegram_id"]
        
        try:
            bind_worker_id(old_id, user_id)
        except Exception as e:
            await update.message.reply_text(
                "❌ Произошла ошибка при регистрации. Пожалуйста, попробуйте позже.",
                reply_markup=menu_for_user(user_id)
            )
            return ConversationHandler.END
            
        w_fio = f"{candidate['last_name']} {candidate['first_name']}"
        await update.message.reply_text(
            f"🎉 *Регистрация успешна!*\n\n"
            f"Вы успешно привязаны к профилю: *{w_fio}* ({candidate['position']}).\n\n"
            "Теперь вы можете отправлять отчеты в этот чат.",
            parse_mode="Markdown",
            reply_markup=menu_for_user(user_id)
        )
        
        # Уведомляем администраторов
        await notify_admins_new_registration(
            bot=context.bot,
            w_fio=w_fio,
            position=candidate['position'],
            username=username,
            user_id=user_id
        )
                
        return ConversationHandler.END
        
    else:
        # Найдено несколько человек
        context.user_data["candidate_workers"] = [dict(w) for w in workers]
        
        # Строим клавиатуру с именами
        buttons = []
        for w in workers:
            buttons.append([f"{w['last_name']} {w['first_name']}"])
        buttons.append([CANCEL_TEXT])
        
        await update.message.reply_text(
            "🔍 Найдено несколько сотрудников с такой фамилией.\n"
            "Пожалуйста, выберите ваше имя на клавиатуре ниже:",
            reply_markup=ReplyKeyboardMarkup(buttons, resize_keyboard=True)
        )
        return ASK_REG_FIRST_NAME

async def register_firstname_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user_id = update.effective_user.id
    if text == CANCEL_TEXT:
        context.user_data.pop("candidate_workers", None)
        await update.message.reply_text("Регистрация отменена.", reply_markup=menu_for_user(user_id))
        return ConversationHandler.END
        
    candidates = context.user_data.get("candidate_workers", [])
    matched_candidate = None
    
    for c in candidates:
        candidate_fio = f"{c['last_name']} {c['first_name']}"
        if text.lower() == candidate_fio.lower():
            matched_candidate = c
            break
            
    if not matched_candidate:
        await update.message.reply_text("❌ Пожалуйста, выберите имя из списка на клавиатуре:")
        return ASK_REG_FIRST_NAME
        
    username = update.effective_user.username or "нет"
    old_id = matched_candidate["telegram_id"]
    
    try:
        bind_worker_id(old_id, user_id)
    except Exception as e:
        await update.message.reply_text(
            "❌ Произошла ошибка при регистрации. Пожалуйста, попробуйте позже.",
            reply_markup=menu_for_user(user_id)
        )
        context.user_data.pop("candidate_workers", None)
        return ConversationHandler.END
        
    w_fio = f"{matched_candidate['last_name']} {matched_candidate['first_name']}"
    await update.message.reply_text(
        f"🎉 *Регистрация успешна!*\n\n"
        f"Вы успешно привязаны к профилю: *{w_fio}* ({matched_candidate['position']}).\n\n"
        "Теперь вы можете отправлять отчеты в этот чат.",
        parse_mode="Markdown",
        reply_markup=menu_for_user(user_id)
    )
    
    # Уведомляем администраторов
    await notify_admins_new_registration(
        bot=context.bot,
        w_fio=w_fio,
        position=matched_candidate['position'],
        username=username,
        user_id=user_id
    )
            
    context.user_data.pop("candidate_workers", None)
    return ConversationHandler.END



# ══════════════════════════════════════════════════════════════════════════════
# Прием отчетов от сотрудников (Голос / Видео / Текст)
# ══════════════════════════════════════════════════════════════════════════════

async def handle_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Проверка, не редактирует ли администратор комментарий ИИ к отчету
    if is_admin(user_id) and context.user_data.get("editing_comment_report_id"):
        report_id = context.user_data["editing_comment_report_id"]
        original_chat_id = context.user_data.get("editing_comment_chat_id")
        original_msg_id = context.user_data.get("editing_comment_message_id")
        original_text = context.user_data.get("editing_comment_original_text", "")
        
        # Сброс состояния
        del context.user_data["editing_comment_report_id"]
        context.user_data.pop("editing_comment_chat_id", None)
        context.user_data.pop("editing_comment_message_id", None)
        context.user_data.pop("editing_comment_original_text", None)
        prompt_message_id = context.user_data.pop("editing_comment_prompt_message_id", None)
        
        # Попытка удалить сообщение-запрос и ответ администратора
        if prompt_message_id and original_chat_id:
            try:
                await context.bot.delete_message(chat_id=original_chat_id, message_id=prompt_message_id)
            except Exception:
                pass
        try:
            await update.message.delete()
        except Exception:
            pass
            
        new_comment = update.message.text.strip() if update.message.text else ""
        if not new_comment or new_comment.lower() in ("отмена", "❌ отмена"):
            return
            
        conn = get_db()
        report = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
        if not report:
            conn.close()
            return
            
        # Обновляем комментарий в БД
        new_action = f"Комментарий изменен администратором вручную: {new_comment}"
        conn.execute(
            "UPDATE reports SET format_comment = ?, required_action = ? WHERE id = ?",
            (new_comment, new_action, report_id)
        )
        # Получаем обновленное состояние
        report = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
        worker = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (report["telegram_id"],)).fetchone()
        conn.commit()
        conn.close()
        asyncio.create_task(async_sync_gsheets_background())
        
        worker_name = f"{worker['last_name']} {worker['first_name']}" if worker else f"ID {report['telegram_id']}"
        status_emoji = "✅" if report["is_ok"] == 1 else "⚠️"
        
        # Обновляем оригинальное сообщение в группе / чате
        if original_chat_id and original_msg_id:
            try:
                # Если исходное сообщение в укороченном формате
                if "Официальный отчет:" not in (original_text or ""):
                    updated_text = (
                        f"🔧 Оценка отчета изменена вручную администратором @{update.effective_user.username or user_id}:\n"
                        f"Сотрудник: {worker_name}\n"
                        f"Дата отчета: {report['report_date']}\n"
                        f"Статус: {report['slot_time'] or report['report_type']}\n"
                        f"Новый статус: {status_emoji} ({new_comment})"
                    )
                else:
                    updated_text = update_message_text_fields(original_text, report["is_ok"] == 1, new_comment)
                
                kbd = make_report_keyboard(report_id, report["report_type"] if report else None)
                await context.bot.edit_message_text(
                    chat_id=original_chat_id,
                    message_id=original_msg_id,
                    text=updated_text,
                    reply_markup=kbd
                )
            except Exception as e:
                logger.error(f"Ошибка при обновлении сообщения после редактирования комментария: {e}")
        return

    if is_admin(user_id):
        return

    lock = get_user_lock(user_id)
    await lock.acquire()

    worker = await run_db(get_worker, user_id)
    
    text_content = ""
    # Решение проблемы 2: Обеспечение гарантированного создания каталога tmp/ и очистки через try/finally
    tmp_path = None
    
    try:
        if update.message.text:
            text_content = update.message.text.strip()
        else:
            file_obj = None
            if update.message.voice: file_obj = update.message.voice
            elif update.message.video: file_obj = update.message.video
            elif update.message.video_note: file_obj = update.message.video_note

            if file_obj:
                await update.message.reply_text("🎙 Отчет получен и отправлен на транскрибацию ИИ, ожидайте оценки...")
                tg_file = await context.bot.get_file(file_obj.file_id)
                ext = "mp4" if update.message.video or update.message.video_note else "ogg"
                
                os.makedirs("tmp", exist_ok=True)
                tmp_path = f"tmp/file_{user_id}_{int(datetime.now().timestamp())}.{ext}"
                
                await tg_file.download_to_drive(tmp_path)
                text_content = await transcribe_audio_async(tmp_path)

        if not text_content:
            await update.message.reply_text("Ошибка: Не удалось распознать аудио или медиа отчета.")
            return

        if text_content.startswith("Ошибка распознавания аудио"):
            await update.message.reply_text("❌ При распознавании аудио произошла ошибка. Пожалуйста, отправьте текстовый отчет или попробуйте перезаписать.")
            return

        is_media = bool(update.message.voice or update.message.video or update.message.video_note)

        # Решение проблемы 3: Сохранение данных незарегистрированных сотрудников по TELEGRAM_ID (через SQLite)
        if not worker:
            user_info = {
                "first_name": update.effective_user.first_name or "",
                "last_name": update.effective_user.last_name or "",
                "username": update.effective_user.username or "",
                "timestamp": datetime.now().isoformat(),
                "text": text_content
            }
            save_pending_unregistered_user(
                telegram_id=user_id,
                first_name=user_info["first_name"],
                last_name=user_info["last_name"],
                username=user_info["username"],
                timestamp=user_info["timestamp"],
                text_content=user_info["text"]
            )
            
            # Оповещение администраторов
            admin_msg = (
                f"👤 Обнаружен отчет от незарегистрированного сотрудника!\n"
                f"TG ID: {user_id}\n"
                f"Имя в Telegram: {user_info['first_name']} {user_info['last_name']} (@{user_info['username']})\n"
                f"Текст:\n\"{text_content[:300]}\"\n\n"
                f"Вы можете добавить его в базу через меню, указав ID."
            )
            for admin_id in ADMIN_IDS:
                try:
                    admin_copied_msg_id = None
                    if is_media:
                        try:
                            admin_copied = await context.bot.copy_message(
                                chat_id=admin_id,
                                from_chat_id=update.effective_chat.id,
                                message_id=update.message.message_id
                            )
                            admin_copied_msg_id = admin_copied.message_id
                        except Exception as copy_err:
                            logger.error(f"Ошибка копирования медиа незарегистрированного пользователя администратору {admin_id}: {copy_err}")
                    
                    if admin_copied_msg_id:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text=admin_msg,
                            reply_to_message_id=admin_copied_msg_id
                        )
                    else:
                        await context.bot.send_message(chat_id=admin_id, text=admin_msg)
                except Exception:
                    pass

            await update.message.reply_text(
                "Ошибка: Вы не зарегистрированы в системе авторизации бота.\n"
                "Ваш отчет отправлен администраторам как временный.\n\n"
                "Нажмите кнопку ниже, чтобы зарегистрироваться:",
                reply_markup=ReplyKeyboardMarkup([["🔑 Начать регистрацию"]], resize_keyboard=True)
            )
            return

        # Проверка статуса "Не работаю сегодня"
        if context.user_data.get("awaiting_not_working_reason"):
            reason = text_content
            if reason.lower() in ("отмена", "❌ отмена"):
                context.user_data.pop("awaiting_not_working_reason", None)
                await update.message.reply_text(
                    "Отменено.",
                    reply_markup=menu_for_user(user_id, update.effective_chat.type)
                )
                return
            
            context.user_data["not_working_reason"] = reason
            context.user_data.pop("awaiting_not_working_reason", None)
            context.user_data["awaiting_not_working_confirm"] = True
            
            kbd = ReplyKeyboardMarkup([["Да, я уверен", "❌ Отмена"]], resize_keyboard=True)
            await update.message.reply_text(
                f"⚠️ Внимание! При переключении в статус 'Не работаю сегодня' ВСЕ ваши сегодняшние отчеты будут безвозвратно удалены.\n\n"
                f"Вы уверены, что хотите продолжить? Причина: {reason}",
                reply_markup=kbd
            )
            return

        if context.user_data.get("awaiting_not_working_confirm"):
            confirm = text_content
            if confirm == "Да, я уверен":
                reason = context.user_data.pop("not_working_reason", "Без причины")
                context.user_data.pop("awaiting_not_working_confirm", None)
                
                now = now_local()
                date_str = now.strftime("%Y-%m-%d")
                
                conn = get_db()
                conn.execute(
                    "DELETE FROM reports WHERE telegram_id = ? AND report_date = ?",
                    (user_id, date_str)
                )
                conn.commit()
                conn.close()
                
                await run_db(
                    save_report,
                    telegram_id=user_id,
                    report_date=date_str,
                    report_type="not_working",
                    slot_time=None,
                    received_at=now.strftime("%H:%M:%S"),
                    is_ok=True,
                    is_late=False,
                    format_comment=reason,
                    required_action="Не работает",
                    raw_text=f"Не работает сегодня. Причина: {reason}"
                )
                
                await update.message.reply_text(
                    f"✅ Статус 'Не работаю' успешно сохранен.\nПричина: {reason}",
                    reply_markup=menu_for_user(user_id, update.effective_chat.type)
                )
                
                w_name = f"{worker['last_name']} {worker['first_name']}"
                dest_chat = worker["group_id"] or DEFAULT_GROUP_ID
                notify_text = (
                    f"🛌 {w_name} сегодня не работает.\n"
                    f"Причина: {reason}"
                )
                try:
                    await context.bot.send_message(chat_id=dest_chat, text=notify_text)
                except Exception as e:
                    logger.error(f"Ошибка отправки уведомления в группу: {e}")
                    
                for admin_id in ADMIN_IDS:
                    try:
                        await context.bot.send_message(chat_id=admin_id, text=notify_text)
                    except Exception:
                        pass
                return
            else:
                context.user_data.pop("not_working_reason", None)
                context.user_data.pop("awaiting_not_working_confirm", None)
                await update.message.reply_text(
                    "Действие отменено.",
                    reply_markup=menu_for_user(user_id, update.effective_chat.type)
                )
                return

        if text_content in ("🛌 Не работаю сегодня", "Не работаю сегодня") or text_content.lower() == "не работаю сегодня":
            now = now_local()
            date_str = now.strftime("%Y-%m-%d")
            conn = get_db()
            existing_not_working = conn.execute(
                "SELECT * FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'not_working'",
                (user_id, date_str)
            ).fetchone()
            conn.close()
            
            if existing_not_working:
                await update.message.reply_text(
                    "У вас уже установлен статус «Не работаю сегодня» на сегодня.",
                    reply_markup=menu_for_user(user_id, update.effective_chat.type)
                )
                return
                
            context.user_data["awaiting_not_working_reason"] = True
            await update.message.reply_text(
                "Укажите, пожалуйста, причину, почему вы сегодня не работаете (например: заболел, отпуск, отпросился у прораба):",
                reply_markup=CANCEL_KEYBOARD
            )
            return

        # КРИТИЧЕСКАЯ ЗАЩИТА: текст кнопок клавиатуры (например "❌ Отмена", "Да, я уверен")
        # НИКОГДА не должен попадать в систему как настоящий отчет сотрудника.
        # Такое могло произойти, если человек нажал кнопку диалога ПОСЛЕ того, как этот диалог
        # уже завершился/истёк по таймауту — тогда нажатие "проваливалось" сюда и неправильно
        # классифицировалось ИИ как реальный (плохой) отчет, попадая в группу и в БД.
        # Отчетом считается ТОЛЬКО то, что сотрудник реально написал сам, или видео/голосовое — и ничего больше.
        if not is_media and text_content.strip() in (CANCEL_TEXT, "Да, я уверен"):
            await update.message.reply_text(
                "Это была кнопка диалога, который уже завершился — отчётом она не считается.\n"
                "Если хотите сдать отчёт, нажмите «📋 Инструкция по сдаче видео-статуса» или просто запишите видео.",
                reply_markup=menu_for_user(user_id, update.effective_chat.type)
            )
            return

        now = now_local()

        # Видео/голосовые отчеты теперь обрабатываются пачкой (см. enqueue_media_report_item):
        # если сотрудник пришлет несколько видео подряд, бот подождет немного, соберет их вместе
        # и сформирует ОДНО общее сообщение-оценку в группе вместо нескольких разрозненных.
        if is_media:
            await enqueue_media_report_item(user_id, context, update, text_content, now)
            return

        date_str = now.strftime("%Y-%m-%d")
        sched_list = SCHEDULES.get(worker["schedule"], SCHEDULE_A)
        
        # Определение типа отчёта по времени: факт дня физически не может прийти раньше
        # последнего слота статуса — днём тип всегда "статус". Вечером же тип не фиксируем
        # жёстко, а даём ИИ самому решить по смыслу (см. check_status: режим "авто") —
        # это нужно, чтобы отличить запоздавший статус по конкретному времени от итога дня.
        last_slot_time_str = sched_list[-1]
        last_hour, last_minute = map(int, last_slot_time_str.split(":"))
        last_slot_time = now.replace(hour=last_hour, minute=last_minute, second=0, microsecond=0)
        last_slot_limit = last_slot_time + dt_module.timedelta(minutes=LATE_THRESHOLD_MIN)

        report_type_override = "status" if now <= last_slot_limit else None

        # Анализ промптом Llama для определения типа отчета
        ai_res_pre = await check_status_async(text_content, report_type_override=report_type_override)
        report_type = ai_res_pre["report_type"]

        if report_type == "status":
            submitted_slots = await run_db(get_submitted_status_slots, user_id, date_str)
            nearest_slot, is_late = pick_target_status_slot(sched_list, now, submitted_slots)
        else:
            nearest_slot, is_late = None, False

        existing_report = await run_db(get_existing_report_row, user_id, date_str, report_type, nearest_slot)

        is_addon = False
        if existing_report:
            is_addon = True
            # Склеиваем предыдущий текст и новое текстовое дополнение.
            text_content = build_addon_text(existing_report["raw_text"], text_content, use_video_label=False)
            # Прогоняем КЛАССИФИКАЦИЮ и АНАЛИЗ заново для объединенного контента.
            # Тип отчета при дополнении не пересматриваем — он уже закреплен за этой записью.
            ai_res = await check_status_async(text_content, report_type_override=report_type)
            cleaned_text = await clean_report_async(text_content)
            report_id = existing_report["id"]
            
            # Обновляем существующий отчет в БД
            await run_db(
                update_report_text_and_ai,
                report_id=report_id,
                is_ok=ai_res["is_ok"],
                format_comment=ai_res["format_comment"],
                required_action=ai_res["required_action"],
                raw_text=text_content,
                received_at=now.strftime("%H:%M:%S")
            )
        else:
            ai_res = ai_res_pre
            cleaned_text = await clean_report_async(text_content)
            report_id = await run_db(
                save_report,
                telegram_id=user_id,
                report_date=date_str,
                report_type=ai_res["report_type"],
                slot_time=nearest_slot if ai_res["report_type"] == "status" else None,
                received_at=now.strftime("%H:%M:%S"),
                is_ok=ai_res["is_ok"],
                is_late=is_late if ai_res["report_type"] == "status" else 0,
                format_comment=ai_res["format_comment"],
                required_action=ai_res["required_action"],
                raw_text=text_content
            )

        w_name = f"{worker['last_name']} {worker['first_name']}"
        
        # Информирование сотрудника
        time_str = now.strftime("%H:%M")
        if ai_res["report_type"] == "status" and nearest_slot:
            sh, sm = map(int, nearest_slot.split(":"))
            slot_mins = sh * 60 + sm
            current_mins = now.hour * 60 + now.minute
            diff_mins = current_mins - slot_mins
            if diff_mins > 0:
                if diff_mins > LATE_THRESHOLD_MIN:
                    late_str = f" (опоздание {diff_mins} мин)"
                else:
                    late_str = f" (вовремя, +{diff_mins} мин)"
            else:
                late_str = f" (раньше на {abs(diff_mins)} мин)"
            info_suffix = f" за слот *{nearest_slot}* принят в *{time_str}*{late_str}"
        else:
            info_suffix = f" (Итог дня) принят в *{time_str}*"

        if ai_res["is_ok"]:
            if is_addon:
                await update.message.reply_text(
                    f"🔄 Дополнение к отчёту{info_suffix}. Успешно проверено ИИ и принято без замечаний! Спасибо.",
                    parse_mode="Markdown"
                )
            else:
                await update.message.reply_text(
                    f"✅ Отчёт{info_suffix} успешно проверен ИИ и принят без замечаний! Спасибо.",
                    parse_mode="Markdown"
                )
        else:
            if is_addon:
                await update.message.reply_text(
                    f"⚠️ Дополнение к отчёту{info_suffix}.\nОценка дополненного отчета: {ai_res['employee_message']}",
                    parse_mode="Markdown"
                )
            else:
                await update.message.reply_text(
                    f"⚠️ Отчёт{info_suffix}.\nОценка отчета: {ai_res['employee_message']}",
                    parse_mode="Markdown"
                )

        # Решение проблемы 7: Кнопка «Исправить оценку» во всех отчетах для администраторов или в группе
        dest_chat = worker["group_id"] or DEFAULT_GROUP_ID
        
        # Получаем красивое название группы
        gname = await get_group_name_async(context.bot, dest_chat)
        
        title_text = f"Дополнение к отчету (отчет обновлен): {w_name}" if is_addon else w_name

        orig_label = "🗣 Оригинальный текст (объединенный):" if is_addon else "🗣 Оригинальный текст:"
        notify_text = (
            f"{title_text}\n"
            f"{format_status_or_fact_line(ai_res['report_type'], nearest_slot if ai_res['report_type'] == 'status' else None, date_str)}\n"
            f"Оценка ИИ: {'ОК' if ai_res['is_ok'] else 'НЕ ОК'}\n"
            f"Комментарий ИИ: {ai_res['format_comment']}\n\n"
            f"📝 Официальный отчет:\n\"{cleaned_text}\"\n\n"
            f"{orig_label}\n\"{text_content}\""
        )
        
        inline_kbd = make_report_keyboard(report_id, ai_res["report_type"])

        # При дополнении удаляем СТАРОЕ текстовое сообщение-оценку в группе (если оно есть),
        # чтобы не плодить дубли. Пересланные медиа (видео/голосовые) НЕ трогаем — они остаются
        # в группе как и были, новое медиа (если есть) просто добавляется отдельным сообщением ниже.
        if is_addon and existing_report and existing_report["group_chat_id"] and existing_report["group_message_id"]:
            try:
                await context.bot.delete_message(
                    chat_id=existing_report["group_chat_id"],
                    message_id=existing_report["group_message_id"]
                )
            except Exception as e:
                logger.warning(f"Не удалось удалить старое сообщение-оценку {existing_report['group_message_id']} в чате {existing_report['group_chat_id']}: {e}")

        copied_msg_id = None
        if is_media:
            try:
                copied_msg = await context.bot.copy_message(
                    chat_id=dest_chat,
                    from_chat_id=update.effective_chat.id,
                    message_id=update.message.message_id
                )
                copied_msg_id = copied_msg.message_id
            except Exception as e:
                logger.error(f"Ошибка копирования медиа в чат {dest_chat}: {e}")

        sent_notify_msg = None
        try:
            if copied_msg_id:
                sent_notify_msg = await context.bot.send_message(
                    chat_id=dest_chat,
                    text=notify_text,
                    reply_markup=inline_kbd,
                    reply_to_message_id=copied_msg_id
                )
            else:
                sent_notify_msg = await context.bot.send_message(
                    chat_id=dest_chat,
                    text=notify_text,
                    reply_markup=inline_kbd
                )
            # Запоминаем id этого сообщения, чтобы следующее дополнение к этому отчету
            # могло его удалить и заменить новым (а не плодить дубли в группе).
            await run_db(set_report_group_message, report_id, dest_chat, sent_notify_msg.message_id)
        except Exception as e:
            # Предохранитель: если отправка в группу сломалась, дублируем всем админам
            logger.error(f"Ошибка отправки оценки в чат {dest_chat}: {e}")
            for admin_id in ADMIN_IDS:
                try:
                    admin_copied_msg_id = None
                    if is_media:
                        try:
                            admin_copied = await context.bot.copy_message(
                                chat_id=admin_id,
                                from_chat_id=update.effective_chat.id,
                                message_id=update.message.message_id
                            )
                            admin_copied_msg_id = admin_copied.message_id
                        except Exception:
                            pass
                    
                    if admin_copied_msg_id:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text=notify_text,
                            reply_markup=inline_kbd,
                            reply_to_message_id=admin_copied_msg_id
                        )
                    else:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text=notify_text,
                            reply_markup=inline_kbd
                        )
                except Exception:
                    pass
    
    finally:
        try:
            lock.release()
        except Exception:
            pass
        # Решение проблемы 2: безусловная чистка временного файла
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception as rm_e:
                logger.warning(f"Ошибка удаления файла {tmp_path}: {rm_e}")


# ══════════════════════════════════════════════════════════════════════════════
# Пачковая обработка видео/голосовых отчетов
# ══════════════════════════════════════════════════════════════════════════════
# Если сотрудник присылает несколько видео подряд (например, 3 видео вечером, из которых
# 2 на самом деле статус по конкретному времени, а 1 — итог дня), бот не обрабатывает их
# по одному, а ждет короткую паузу, собирает все пришедшие видео вместе и:
#   1) пересылает их все в группу по порядку (как и раньше);
#   2) классифицирует КАЖДОЕ видео отдельно (статус/факт — по смыслу, если вечер);
#   3) создает/дополняет правильные строки в БД для каждого видео отдельно;
#   4) публикует ОДНО общее сообщение под всеми видео с разбивкой "Видео 1 / 2 / 3"
#      и общим итогом, вместо нескольких разрозненных сообщений.

MEDIA_BATCH_BUFFERS: dict[int, dict] = {}
MEDIA_BATCH_DEBOUNCE_SECONDS = 8

# Если сотрудник присылает ЕЩЁ одно видео к УЖЕ существующему статусу в пределах этого окна
# (например, забыл что-то сказать и досылает видео через несколько минут) — старое видео и
# старый комментарий-оценка в группе удаляются, и публикуются ЗАНОВО все видео этого статуса
# вместе (например 2 видео) с одним общим обновленным комментарием под ними.
# Если видео приходит позже этого окна — старое видео в группе не трогается (остается историей),
# дополняется только текст и комментарий-оценка (как раньше).
MEDIA_MERGE_WINDOW_MINUTES = 20


async def enqueue_media_report_item(user_id: int, context: ContextTypes.DEFAULT_TYPE, update: Update, text_content: str, now: datetime):
    """Добавляет видео/голосовое сообщение в буфер пользователя и (пере)запускает таймер ожидания.
    Если в течение MEDIA_BATCH_DEBOUNCE_SECONDS придет еще одно видео — таймер сбрасывается,
    и все видео будут обработаны вместе одной пачкой."""
    buf = MEDIA_BATCH_BUFFERS.setdefault(user_id, {"items": [], "task": None})
    buf["items"].append({"update": update, "text_content": text_content, "now": now})

    old_task = buf.get("task")
    if old_task and not old_task.done():
        old_task.cancel()
    buf["task"] = asyncio.create_task(_flush_media_batch(user_id, context))


async def _flush_media_batch(user_id: int, context: ContextTypes.DEFAULT_TYPE):
    try:
        await asyncio.sleep(MEDIA_BATCH_DEBOUNCE_SECONDS)
    except asyncio.CancelledError:
        # Пришло новое видео, таймер был перезапущен — эта задача больше не нужна
        return

    buf = MEDIA_BATCH_BUFFERS.pop(user_id, None)
    if not buf or not buf["items"]:
        return

    lock = get_user_lock(user_id)
    await lock.acquire()
    try:
        await process_media_batch(user_id, buf["items"], context)
    except Exception:
        logger.exception(f"Ошибка обработки пачки видео-отчетов пользователя {user_id}")
    finally:
        try:
            lock.release()
        except Exception:
            pass


async def process_media_batch(user_id: int, items: list[dict], context: ContextTypes.DEFAULT_TYPE):
    worker = await run_db(get_worker, user_id)
    if not worker:
        return  # незарегистрированный сотрудник уже обработан на этапе приёма каждого видео

    sched_list = SCHEDULES.get(worker["schedule"], SCHEDULE_A)
    w_name = f"{worker['last_name']} {worker['first_name']}"
    dest_chat = worker["group_id"] or DEFAULT_GROUP_ID

    last_slot_time_str = sched_list[-1]
    last_hour, last_minute = map(int, last_slot_time_str.split(":"))

    results = []
    for idx, item in enumerate(items, start=1):
        text_content = item["text_content"]
        now = item["now"]
        upd = item["update"]
        date_str = now.strftime("%Y-%m-%d")

        last_slot_time = now.replace(hour=last_hour, minute=last_minute, second=0, microsecond=0)
        last_slot_limit = last_slot_time + dt_module.timedelta(minutes=LATE_THRESHOLD_MIN)

        # Днём факт дня в принципе невозможен — это всегда статус.
        # Вечером даём ИИ решить по смыслу каждого видео отдельно — так бот различает
        # видео "на статус" и видео "на факт", даже если все они присланы в одно и то же время.
        forced_type = "status" if now <= last_slot_limit else None

        ai_res_pre = await check_status_async(text_content, report_type_override=forced_type)
        report_type = ai_res_pre["report_type"]

        if report_type == "status":
            submitted_slots = await run_db(get_submitted_status_slots, user_id, date_str)
            slot_time, is_late = pick_target_status_slot(sched_list, now, submitted_slots)
        else:
            slot_time, is_late = None, False

        existing = await run_db(get_existing_report_row, user_id, date_str, report_type, slot_time)
        use_label = len(items) > 1 or existing is not None

        # Определяем, попадаем ли мы в окно "слияния с переотправкой видео":
        # сотрудник дополняет статус ВИДЕО в пределах MEDIA_MERGE_WINDOW_MINUTES после последнего
        # обновления этого же статуса — тогда старое видео и старый комментарий в группе удаляются,
        # и публикуются заново ВСЕ видео этого статуса вместе с одним обновленным комментарием.
        do_full_merge = False
        if existing and report_type == "status":
            try:
                prev_h, prev_m, prev_s = map(int, existing["received_at"].split(":"))
                elapsed_minutes = (now.hour * 60 + now.minute) - (prev_h * 60 + prev_m)
                do_full_merge = 0 <= elapsed_minutes <= MEDIA_MERGE_WINDOW_MINUTES
            except Exception:
                do_full_merge = False

        old_media_rows = []
        if do_full_merge:
            old_media_rows = await run_db(get_report_media, existing["id"])

        if existing:
            merged_raw = build_addon_text(existing["raw_text"], text_content, use_video_label=True)
            ai_res = await check_status_async(merged_raw, report_type_override=report_type)
            cleaned_text = await clean_report_async(merged_raw)
            raw_text_final = merged_raw
            report_id = existing["id"]
            await run_db(
                update_report_text_and_ai,
                report_id=report_id,
                is_ok=ai_res["is_ok"],
                format_comment=ai_res["format_comment"],
                required_action=ai_res["required_action"],
                raw_text=raw_text_final,
                received_at=now.strftime("%H:%M:%S")
            )
            is_addon_item = True
        else:
            ai_res = ai_res_pre
            raw_text_final = f"[Видео {idx}]: {text_content}" if use_label else text_content
            cleaned_text = await clean_report_async(raw_text_final)
            report_id = await run_db(
                save_report,
                telegram_id=user_id,
                report_date=date_str,
                report_type=report_type,
                slot_time=slot_time,
                received_at=now.strftime("%H:%M:%S"),
                is_ok=ai_res["is_ok"],
                is_late=is_late if report_type == "status" else 0,
                format_comment=ai_res["format_comment"],
                required_action=ai_res["required_action"],
                raw_text=raw_text_final
            )
            is_addon_item = False

        copied_msg_id = None
        if do_full_merge and old_media_rows:
            # Удаляем СТАРЫЕ видео из группы (и старый комментарий чуть ниже, через общий механизм)
            for m in old_media_rows:
                if m["group_message_id"]:
                    try:
                        await context.bot.delete_message(chat_id=dest_chat, message_id=m["group_message_id"])
                    except Exception as e:
                        logger.warning(f"Не удалось удалить старое видео {m['group_message_id']} в чате {dest_chat}: {e}")
            await run_db(delete_report_media_rows, report_id)

            # Пересылаем заново ВСЕ видео этого статуса по порядку: старые (из их исходных сообщений
            # у сотрудника) + новое.
            media_sources = [(m["source_chat_id"], m["source_message_id"]) for m in old_media_rows]
            media_sources.append((upd.effective_chat.id, upd.message.message_id))
            for pos, (src_chat, src_msg) in enumerate(media_sources, start=1):
                try:
                    copied_msg = await context.bot.copy_message(chat_id=dest_chat, from_chat_id=src_chat, message_id=src_msg)
                    await run_db(add_report_media, report_id, src_chat, src_msg, copied_msg.message_id, pos, now.strftime("%H:%M:%S"))
                    if pos == 1:
                        copied_msg_id = copied_msg.message_id
                except Exception as e:
                    logger.error(f"Ошибка повторной пересылки видео {pos} в чат {dest_chat}: {e}")
        else:
            # Обычный случай: пересылаем только ЭТО видео, старые (если были) не трогаем
            try:
                copied_msg = await context.bot.copy_message(
                    chat_id=dest_chat,
                    from_chat_id=upd.effective_chat.id,
                    message_id=upd.message.message_id
                )
                copied_msg_id = copied_msg.message_id
                next_pos = len(await run_db(get_report_media, report_id)) + 1
                await run_db(add_report_media, report_id, upd.effective_chat.id, upd.message.message_id, copied_msg_id, next_pos, now.strftime("%H:%M:%S"))
            except Exception as e:
                logger.error(f"Ошибка копирования видео {idx} в чат {dest_chat}: {e}")

        results.append({
            "idx": idx, "report_type": report_type, "slot_time": slot_time,
            "ai_res": ai_res, "report_id": report_id, "copied_msg_id": copied_msg_id,
            "is_addon": is_addon_item, "cleaned_text": cleaned_text, "raw_text": raw_text_final,
            "date_str": date_str,
        })

        # Личный фидбек сотруднику по каждому видео отдельно (как и раньше)
        time_str = now.strftime("%H:%M")
        suffix_tail = f" (видео {idx} из {len(items)})" if len(items) > 1 else ""
        if report_type == "status" and slot_time:
            info_suffix = f" за слот *{slot_time}*{suffix_tail}"
        else:
            info_suffix = f" (Итог дня){suffix_tail}"
        try:
            if ai_res["is_ok"]:
                await upd.message.reply_text(f"✅ Отчёт{info_suffix} принят без замечаний!", parse_mode="Markdown")
            else:
                await upd.message.reply_text(f"⚠️ Отчёт{info_suffix}.\n{ai_res['employee_message']}", parse_mode="Markdown")
        except Exception as e:
            logger.warning(f"Не удалось отправить личный фидбек по видео {idx} пользователю {user_id}: {e}")

    if not results:
        return

    # Формируем ОДНО общее сообщение-оценку в группу под всеми видео этой пачки
    if len(results) == 1:
        r = results[0]
        notify_text = (
            f"{w_name}\n"
            f"{format_status_or_fact_line(r['report_type'], r['slot_time'], r['date_str'])}\n"
            f"Оценка ИИ: {'ОК' if r['ai_res']['is_ok'] else 'НЕ ОК'}\n"
            f"Комментарий ИИ: {r['ai_res']['format_comment']}\n\n"
            f"📝 Официальный отчет:\n\"{r['cleaned_text']}\"\n\n"
            f"🗣 Оригинальный текст:\n\"{r['raw_text']}\""
        )
        inline_kbd = make_report_keyboard(r["report_id"], r["report_type"])
    else:
        lines = []
        for r in results:
            label = f"Статус {r['slot_time']}" if r["report_type"] == "status" else "Итог дня"
            icon = "✅" if r["ai_res"]["is_ok"] else "⚠️"
            lines.append(f"🎬 Видео {r['idx']} — {label}: {icon} {r['ai_res']['format_comment']}")
        overall_ok = all(r["ai_res"]["is_ok"] for r in results)
        overall_line = "Все видео приняты без замечаний." if overall_ok else "Есть замечания — см. видео выше."
        notify_text = (
            f"{w_name}\n"
            f"📦 Отчёт из {len(results)} видео за {results[0]['date_str']}:\n\n"
            + "\n".join(lines) +
            f"\n\nОбщий итог: {overall_line}"
        )
        kbd_rows = []
        for r in results:
            kbd_rows.append([
                InlineKeyboardButton(f"✏️ Видео {r['idx']}: комментарий", callback_data=f"edit_comment_{r['report_id']}"),
                InlineKeyboardButton(f"🔄 Видео {r['idx']}: оценка", callback_data=f"fix_toggle_{r['report_id']}")
            ])
        inline_kbd = InlineKeyboardMarkup(kbd_rows)

    # Если для какого-то из обновленных отчетов уже было своё предыдущее сообщение-оценка в группе —
    # удаляем его, чтобы не плодить дубли (это отдельно от удаления видео при слиянии внутри
    # MEDIA_MERGE_WINDOW_MINUTES — то уже сделано выше для каждого видео индивидуально).
    old_messages_to_delete = set()
    for r in results:
        if r["is_addon"]:
            row = await run_db(get_report_group_message, r["report_id"])
            if row and row["group_chat_id"] and row["group_message_id"]:
                old_messages_to_delete.add((row["group_chat_id"], row["group_message_id"]))
    for chat_id, msg_id in old_messages_to_delete:
        try:
            await context.bot.delete_message(chat_id=chat_id, message_id=msg_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить старое сообщение-оценку {msg_id} в чате {chat_id}: {e}")

    first_copied_msg_id = next((r["copied_msg_id"] for r in results if r["copied_msg_id"]), None)
    try:
        if first_copied_msg_id:
            sent_notify_msg = await context.bot.send_message(
                chat_id=dest_chat, text=notify_text, reply_markup=inline_kbd, reply_to_message_id=first_copied_msg_id
            )
        else:
            sent_notify_msg = await context.bot.send_message(chat_id=dest_chat, text=notify_text, reply_markup=inline_kbd)
        for r in results:
            await run_db(set_report_group_message, r["report_id"], dest_chat, sent_notify_msg.message_id)
    except Exception as e:
        logger.error(f"Ошибка отправки общей оценки по пачке из {len(results)} видео в чат {dest_chat}: {e}")
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=notify_text, reply_markup=inline_kbd)
            except Exception:
                pass


# ── Новые функции для Прораб-Бот ──

async def top_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id) and await run_db(get_worker, user_id) is None:
        await update.message.reply_text("Эта команда доступна только сотрудникам и администраторам.")
        return
        
    await update.message.reply_text("⏳ Расчитываю рейтинг лучших сотрудников за последние 30 дней...")
    
    now = now_local()
    start_dt = now - dt_module.timedelta(days=29)
    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = now.strftime("%Y-%m-%d")
    
    conn = get_db()
    try:
        workers = conn.execute("SELECT * FROM workers WHERE is_active = 1").fetchall()
        leaderboard = []
        
        for w in workers:
            lastname_lower = (w["last_name"] or "").lower()
            firstname_lower = (w["first_name"] or "").lower()
            dept_lower = (w["position"] or "").lower()
            if any(x in lastname_lower or x in firstname_lower or x in dept_lower for x in ("отмена", "test", "тест")):
                continue
                
            stats = calculate_worker_stats(w, start_str, end_str, conn)
            if stats["expected"] > 0:
                leaderboard.append((w, stats))
        
        # Сортировка по проценту, количеству сданных, отсутствию опозданий
        leaderboard.sort(key=lambda x: (x[1]["percent"], x[1]["submitted"], -x[1]["lates"]), reverse=True)
        
        top_5 = leaderboard[:5]
        if not top_5:
            await update.message.reply_text("Нет данных по сотрудникам за последний месяц.")
            return
            
        lines = [
            f"🏆 *Топ-5 лучших сотрудников по дисциплине* за последние 30 дней:\n"
            f"({start_dt.strftime('%d.%m.%Y')} - {now.strftime('%d.%m.%Y')})\n"
        ]
        
        medal_icons = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
        for idx, (w, stats) in enumerate(top_5):
            name = f"{w['last_name']} {w['first_name']}"
            icon = medal_icons[idx]
            lines.append(
                f"{icon} *{name}* ({w['position']})\n"
                f"   • Выполнение: *{stats['percent']:.1f}%*\n"
                f"   • Сдано отчетов: {stats['submitted']} из {stats['expected']}\n"
                f"   • Опозданий: {stats['lates']} | Замечаний: {stats['remarks']}\n"
            )
        
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    except Exception as e:
        logger.error(f"Ошибка при расчете команды /top: {e}")
        await update.message.reply_text("❌ Произошла ошибка при расчете рейтинга.")
    finally:
        conn.close()

async def set_object_group_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("Эта команда доступна только администраторам.")
        return
        
    args = context.args
    if not args:
        await update.message.reply_text(
            "📝 *Установка группы для объекта*:\n\n"
            "Вы можете вызвать эту команду в группе:\n"
            "`/set_object_group НазваниеОбъекта`\n"
            "или указав ID группы:\n"
            "`/set_object_group НазваниеОбъекта -100xxxxxxxxx`",
            parse_mode="Markdown"
        )
        return
        
    obj_name = args[0]
    if len(args) > 1:
        try:
            group_id = int(args[1])
        except ValueError:
            await update.message.reply_text("❌ Неверный формат ID группы. Это должно быть целое число.")
            return
    else:
        group_id = update.effective_chat.id
        if update.effective_chat.type == "private":
            await update.message.reply_text("❌ В личном чате необходимо указать ID группы вручную: `/set_object_group ИмяОбъекта ID`.")
            return
            
    save_object_group(obj_name, group_id)
    await update.message.reply_text(
        f"✅ Объект «*{obj_name}*» успешно привязан к группе с ID `{group_id}`!",
        parse_mode="Markdown"
    )

async def cmd_quiet_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Эта команда доступна только администраторам.")
        return
        
    current_status = is_quiet_mode_enabled()
    new_status = not current_status
    set_quiet_mode(new_status)
    
    if new_status:
        await update.message.reply_text(
            "🔇 *Тихий режим для группы ВКЛЮЧЕН*.\n\n"
            "Индивидуальные напоминания о пропусках отчетов больше не будут отправляться в группы. "
            "Нарушители будут получать уведомления только в личные сообщения.",
            parse_mode="Markdown"
        )
    else:
        await update.message.reply_text(
            "🔊 *Тихий режим для группы ВЫКЛЮЧЕН*.\n\n"
            "Напоминания о пропущенных отчетах будут дублироваться в группы бригад/объектов.",
            parse_mode="Markdown"
        )

async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Эта команда доступна только администраторам.")
        return
        
    now = now_local()
    # По умолчанию с начала текущего месяца до сегодняшнего дня
    start_date = dt_module.date(now.year, now.month, 1)
    end_date = now.date()
    
    args = context.args
    if len(args) >= 2:
        try:
            try:
                start_date = datetime.strptime(args[0], "%d.%m.%Y").date()
            except ValueError:
                start_date = datetime.strptime(args[0], "%Y-%m-%d").date()
                
            try:
                end_date = datetime.strptime(args[1], "%d.%m.%Y").date()
            except ValueError:
                end_date = datetime.strptime(args[1], "%Y-%m-%d").date()
        except Exception:
            await update.message.reply_text(
                "❌ Неверный формат дат.\n"
                "Используйте: `/stats ДД.ММ.ГГГГ ДД.ММ.ГГГГ` или `/stats` для статистики за текущий месяц.",
                parse_mode="Markdown"
            )
            return

    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")
    
    await update.message.reply_text(f"⏳ Рассчитываю статистику за период {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}...")
    
    conn = get_db()
    workers = conn.execute("SELECT * FROM workers ORDER BY position, last_name, first_name").fetchall()
    
    if not workers:
        conn.close()
        await update.message.reply_text("Сотрудники в базе данных не найдены.")
        return
        
    workers_by_dept = {}
    for w in workers:
        dept = w["position"] or "Не указано"
        workers_by_dept.setdefault(dept, []).append(w)
        
    lines = [f"📊 *Сводная статистика сдачи отчетов*"]
    lines.append(f"📅 Период: *{start_date.strftime('%d.%m.%Y')}* — *{end_date.strftime('%d.%m.%Y')}*\n")
    
    total_expected = 0
    total_submitted = 0
    total_lates = 0
    total_remarks = 0
    
    for dept, dept_workers in sorted(workers_by_dept.items()):
        lines.append(f"🏢 *Отдел: {dept}*")
        for w in dept_workers:
            stats = calculate_worker_stats(w, start_date_str, end_date_str, conn)
            exp = stats["expected"]
            sub = stats["submitted"]
            lates = stats["lates"]
            remarks = stats["remarks"]
            pct = stats["percent"]
            
            total_expected += exp
            total_submitted += sub
            total_lates += lates
            total_remarks += remarks
            
            w_name = f"{w['last_name']} {w['first_name']}"
            
            if exp == 0:
                emoji = "⚪️"
                pct_str = "нет слотов"
            else:
                pct_str = f"{pct}%"
                if pct >= 90: emoji = "🟢"
                elif pct >= 70: emoji = "🟡"
                else: emoji = "🔴"
                
            lines.append(
                f"  {emoji} {w_name}:\n"
                f"    Ожидалось: {exp} | Сдано: {sub}\n"
                f"    Опозданий: {lates} | Замечаний: {remarks}\n"
                f"    Успешность: *{pct_str}*\n"
            )
            
    conn.close()
    
    lines.append("───────────────────")
    if total_expected > 0:
        total_pct = round((total_submitted / total_expected) * 100)
    else:
        total_pct = 100
    lines.append(
        f"📈 *Итого по всем сотрудникам*:\n"
        f"  Ожидалось отчетов: {total_expected}\n"
        f"  Сдано всего: {total_submitted}\n"
        f"  Всего опозданий: {total_lates}\n"
        f"  Всего замечаний: {total_remarks}\n"
        f"  Средняя успешность: *{total_pct}%*"
    )
    
    full_text = "\n".join(lines)
    for part in split_message(full_text):
        await update.message.reply_text(part, parse_mode="Markdown")

async def daily_backup_callback(context: ContextTypes.DEFAULT_TYPE):
    logger.info("Запуск ежедневного резервного копирования базы данных...")
    now = now_local()
    date_str = now.strftime("%Y_%m_%d")
    backup_filename = f"backup_{date_str}.sql"
    backup_path = os.path.join(os.getcwd(), backup_filename)
    
    try:
        import sqlite3
        conn = sqlite3.connect(DB_PATH)
        with open(backup_path, "w", encoding="utf-8") as f:
            for line in conn.iterdump():
                f.write(f"{line}\n")
        conn.close()
        logger.info(f"Дамп базы данных успешно сохранен в {backup_path}")
        
        for admin_id in ADMIN_IDS:
            try:
                with open(backup_path, "rb") as backup_file:
                    await context.bot.send_document(
                        chat_id=admin_id,
                        document=backup_file,
                        filename=backup_filename,
                        caption=f"💾 Автоматическая резервная копия базы данных за {now.strftime('%d.%m.%Y %H:%M')}"
                    )
                logger.info(f"Бэкап успешно отправлен администратору {admin_id}")
            except Exception as ex:
                logger.error(f"Не удалось отправить бэкап администратору {admin_id}: {ex}")
                
    except Exception as e:
        logger.error(f"Критическая ошибка при создании бэкапа базы данных: {e}")
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=f"❌ Ошибка автоматического резервного копирования базы данных: {e}")
            except Exception:
                pass
    finally:
        if os.path.exists(backup_path):
            try:
                os.remove(backup_path)
                logger.info(f"Временный файл бэкапа {backup_path} удален")
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════════════
# Инициализация и запуск приложения
# ══════════════════════════════════════════════════════════════════════════════

async def post_init(application: Application):
    """Срабатывает после запуска приложения Telegram."""
    # Решение проблемы 8: автоматическое сохранение имени группы по умолчанию при старте
    try:
        chat = await application.bot.get_chat(DEFAULT_GROUP_ID)
        name = chat.title or str(DEFAULT_GROUP_ID)
        save_group_name(DEFAULT_GROUP_ID, name)
        logger.info(f"Группа по умолчанию кэширована: {name}")
    except Exception as e:
        logger.error(f"Не удалось получить название DEFAULT_GROUP_ID {DEFAULT_GROUP_ID}: {e}")

    # Решение проблемы 12: Восстановление автосводок из БД при старте / перезапуске
    reschedule_summary_jobs(application)

    # Периодическая фоновая проверка забытых/пропущенных отчетов (каждые 5 минут)
    job_queue = application.job_queue
    if job_queue:
        job_queue.run_repeating(check_missing_reports_job, interval=300, first=10)

def main():
    init_db()
    if not TOKEN:
        logger.error("Критическая ошибка: не задан TELEGRAM_TOKEN")
        return

    # Запуск бота с подключением к функции post_init для восстановления кеша и сводок
    application = Application.builder().token(TOKEN).post_init(post_init).build()

    # Точечные команды и кнопки меню (срабатывают моментально)
    # Команда /start теперь обрабатывается в registration_handler для поддержки саморегистрации
    application.add_handler(CommandHandler("myreports", myreports))
    application.add_handler(CommandHandler("status", status))
    application.add_handler(CommandHandler("threshold", set_threshold_command))
    application.add_handler(CommandHandler("top", top_command))
    application.add_handler(CommandHandler("set_object_group", set_object_group_command))
    application.add_handler(CommandHandler("stats", cmd_stats))
    application.add_handler(CommandHandler("quiet_mode", cmd_quiet_mode))
    application.add_handler(MessageHandler(filters.Regex("^🆔 ID чата$"), get_chat_id))
    application.add_handler(MessageHandler(filters.Regex("^📊 Сводка сейчас$"), send_summary_now))
    application.add_handler(MessageHandler(filters.Regex("^📣 Напомнить всем$"), remind_all_missing))
    
    # Callback-кнопки (для изменения оценок)
    application.add_handler(CallbackQueryHandler(handle_callback_query))

    # Диалоговые обработчики (ConversationHandlers)
    list_handler = ConversationHandler(
        entry_points=[
            CommandHandler("find", find_worker_command),
            MessageHandler(filters.Regex("^📋 Сотрудники$"), list_workers)
        ],
        states={
            ASK_LIST_DEPARTMENT: [MessageHandler(DIALOG_TEXT, list_workers_department)],
            ASK_LIST_WORKER: [MessageHandler(DIALOG_TEXT, list_workers_select)],
            ASK_EDIT_FIELD: [MessageHandler(DIALOG_TEXT, list_workers_action)],
            ASK_EDIT_VALUE: [MessageHandler(DIALOG_TEXT, edit_value_finish)],
            ASK_EDIT_GROUP_VALUE: [MessageHandler(DIALOG_TEXT, edit_group_finish)],
            ASK_EDIT_SCHEDULE: [MessageHandler(DIALOG_TEXT, edit_schedule_finish)],
            ASK_EDIT_DAILY_FACT: [MessageHandler(DIALOG_TEXT, edit_daily_fact_finish)],
            ASK_EDIT_STATUS_WORK: [MessageHandler(DIALOG_TEXT, edit_status_work_finish)],
            ASK_EDIT_SORT_ORDER: [MessageHandler(DIALOG_TEXT, edit_sort_order_finish)],
            ASK_MOVE_POSITION_ORDER: [MessageHandler(DIALOG_TEXT, edit_move_position_order_finish)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    add_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^➕ Добавить сотрудника$"), add_worker_start)],
        states={
            ASK_WORKER_ID: [MessageHandler(DIALOG_TEXT, add_worker_id)],
            ASK_LASTNAME: [MessageHandler(DIALOG_TEXT, add_worker_lastname)],
            ASK_FIRSTNAME: [MessageHandler(DIALOG_TEXT, add_worker_firstname)],
            ASK_POSITION: [MessageHandler(DIALOG_TEXT, add_worker_position)],
            ASK_GROUP: [MessageHandler(DIALOG_TEXT, add_worker_group)],
            ASK_SCHEDULE: [MessageHandler(DIALOG_TEXT, add_worker_schedule)],
            ASK_NEEDS_DAILY_FACT: [MessageHandler(DIALOG_TEXT, add_worker_needs_daily_fact)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    delete_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^➖ Удалить сотрудника$"), delete_worker_start)],
        states={
            ASK_REMOVE_DEPARTMENT: [MessageHandler(DIALOG_TEXT, delete_worker_department)],
            ASK_REMOVE_WORKER: [MessageHandler(DIALOG_TEXT, delete_worker_finish)],
            ASK_CONFIRM_DELETE: [MessageHandler(DIALOG_TEXT, delete_worker_confirm)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    view_dept_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^🏢 Сотрудники отдела$"), department_workers_start)],
        states={
            ASK_DEPARTMENT: [MessageHandler(DIALOG_TEXT, department_workers_show)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    summary_scheduler_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^⏰ Время сводки$"), summary_time_start)],
        states={
            ASK_REPORT_TIME: [MessageHandler(DIALOG_TEXT, summary_time_action)],
            ASK_EDIT_SCHEDULE: [MessageHandler(DIALOG_TEXT, summary_time_add_finish)],
            ASK_ORDER_DEPARTMENT: [MessageHandler(DIALOG_TEXT, summary_time_del_finish)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    export_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📥 Выгрузить отчеты$"), export_start)],
        states={
            ASK_EXPORT_FORMAT: [MessageHandler(DIALOG_TEXT, export_format_selected)],
            ASK_EXPORT_TYPE: [MessageHandler(DIALOG_TEXT, export_type_selected)],
            ASK_EXPORT_DEPARTMENT: [MessageHandler(DIALOG_TEXT, export_department_selected)],
            ASK_GSHEETS_URL: [MessageHandler(DIALOG_TEXT, export_gsheets_url_received)],
            ASK_GSHEETS_CREDS: [MessageHandler(filters.Document.ALL | filters.TEXT & ~filters.COMMAND, export_gsheets_creds_received)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    import_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📥 Импорт сотрудников$"), import_workers_start)],
        states={
            ASK_IMPORT_FILE: [MessageHandler(filters.Document.ALL, import_workers_file)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    summary_date_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^📅 Сводка за дату$"), summary_date_start),
            CommandHandler("summary_date", summary_date_start)
        ],
        states={
            ASK_SUMMARY_DATE: [MessageHandler(DIALOG_TEXT, summary_date_finish)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    dept_schedule_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^🔄 График отдела$"), dept_schedule_start),
            CommandHandler("dept_schedule", dept_schedule_start)
        ],
        states={
            ASK_EDIT_SCHEDULE_DEPT: [MessageHandler(DIALOG_TEXT, dept_schedule_value)],
            ASK_DEPT_SCHEDULE_VAL: [MessageHandler(DIALOG_TEXT, dept_schedule_finish)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    registration_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", register_start),
            MessageHandler(filters.Regex("^🔑 Начать регистрацию$"), register_start)
        ],
        states={
            ASK_REG_LAST_NAME: [MessageHandler(DIALOG_TEXT, register_lastname_received)],
            ASK_REG_FIRST_NAME: [MessageHandler(DIALOG_TEXT, register_firstname_received)],
            ConversationHandler.TIMEOUT: [MessageHandler(filters.ALL, conversation_timeout_callback)],
        },
        fallbacks=[MessageHandler(filters.Regex(f"^{CANCEL_TEXT}$"), cancel)],
        conversation_timeout=300,
    )

    # Регистрация диалогов
    application.add_handler(registration_handler)
    application.add_handler(list_handler)
    application.add_handler(add_handler)
    application.add_handler(delete_handler)
    application.add_handler(view_dept_handler)
    application.add_handler(summary_scheduler_handler)
    application.add_handler(export_handler)
    application.add_handler(import_handler)
    application.add_handler(summary_date_handler)
    application.add_handler(dept_schedule_handler)

    # Дополнительные хэндлеры для меню обычных сотрудников
    application.add_handler(MessageHandler(filters.Regex("^📋 Инструкция по сдаче видео-статуса$"), send_report_instruction))
    application.add_handler(MessageHandler(filters.Regex("^📊 Мой статус$"), status))
    application.add_handler(MessageHandler(filters.Regex("^📅 Мои отчеты за 30 дней$"), myreports))

    # Хэндлер для приема аудио/видео/текстовых отчетов сотрудников (регистрируется в самом конце)
    application.add_handler(MessageHandler(
        filters.VOICE | filters.VIDEO | filters.VIDEO_NOTE | filters.TEXT & ~filters.COMMAND, 
        handle_report
    ))

    logger.info("Бот успешно инициализирован и запущен...")
    application.run_polling()

if __name__ == "__main__":
    main()
