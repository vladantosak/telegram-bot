import os
import re
import sqlite3
import json
import hashlib
import itertools
import logging
from datetime import datetime
import datetime as dt_module
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

try:
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError:
    gspread = None
    Credentials = None

DB_PATH = os.environ.get("DB_PATH", "workers.db")
DEFAULT_GROUP_ID = int(os.environ.get("GROUP_ID", "-1003804380536"))
LOCAL_TZ = ZoneInfo("Europe/Chisinau")
LATE_THRESHOLD_MIN = 15  # used for reminders / "expected" counting - when a slot is considered due
# Grace period: how many minutes after a slot's anchor time a status is still accepted as
# "on time" (an early submission is never late, no matter how early — see
# pick_target_status_slot). Single, sole source of truth for this threshold across the
# project (report_handlers.py's slot attribution, bot.py's missed-status reminder, and
# db.py's own Сводка "still pending" gate below all import/reference this one constant),
# so "still within your window" and "you missed it" can't disagree, and there's exactly one
# place to change the grace period.
STATUS_LATE_TOLERANCE_MIN = 30

# Read ADMIN_IDS from environment variables
ADMIN_IDS_RAW = os.environ.get("ADMIN_IDS", "")
ADMIN_IDS = []
if ADMIN_IDS_RAW:
    for x in ADMIN_IDS_RAW.split(","):
        x = x.strip()
        if x.replace("-", "").isdigit():
            ADMIN_IDS.append(int(x))

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

SCHEDULE_A = ["10:00", "12:00", "15:00", "17:00"]
SCHEDULE_B = ["11:00", "13:00", "16:00", "18:00"]
SCHEDULES = {"A": SCHEDULE_A, "B": SCHEDULE_B}

def clean_position(position: str | None) -> str:
    """Strip a trailing '(...)' suffix some positions were saved with, e.g.
    'Сварщик (Industriala_Welders_Reports)' -> 'Сварщик' — the department
    is already shown separately, it shouldn't be duplicated inside position."""
    text = (position or "Не указано").strip()
    return re.sub(r"\s*\([^()]*\)\s*$", "", text).strip() or "Не указано"

def extract_issue_lines(format_comment: str) -> list[str]:
    """Turns an AI-generated format_comment (possibly multi-video, e.g. 'Видео 1: не ОК -
    на видео молчал; Видео 2: ОК - сказал что сделал') into a plain list of just the
    negative remarks, one per line — no 'Видео N:'/'не ОК -' labels, no clean videos."""
    parts = [p.strip() for p in (format_comment or "").split(";") if p.strip()]
    lines = []
    for part in parts:
        part = re.sub(r"^Видео\s*\d+\s*:\s*", "", part, flags=re.IGNORECASE).strip()
        low = part.lower()
        if low.startswith("не ок"):
            issue = re.sub(r"^не\s*ок\s*-?\s*", "", part, flags=re.IGNORECASE).strip()
            if issue:
                lines.append(issue[0].upper() + issue[1:])
        elif low.startswith("ок"):
            continue
        elif part:
            lines.append(part[0].upper() + part[1:])
    return lines


ENCRYPTED_SETTING_KEYS = {"google_service_account"}
_fernet = None

def _get_fernet():
    global _fernet
    if _fernet is not None:
        return _fernet
    key = os.environ.get("SETTINGS_ENCRYPTION_KEY")
    if not key:
        _fernet = False
        return _fernet
    try:
        from cryptography.fernet import Fernet
        _fernet = Fernet(key.encode())
    except Exception:
        _fernet = False
    return _fernet

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
    except Exception:
        pass
    return conn

def now_local() -> datetime:
    return datetime.now(LOCAL_TZ)

def backup_database_to_file(target_path: str):
    """Consistent snapshot of the live DB (safe under WAL mode / concurrent writers)."""
    source = sqlite3.connect(DB_PATH, timeout=30.0)
    dest = sqlite3.connect(target_path)
    try:
        source.backup(dest)
    finally:
        dest.close()
        source.close()

def restore_database_from_file(source_path: str):
    """Reverse of backup_database_to_file - copies a previously-taken snapshot back over the
    live database, using the same SQLite online-backup API (safe under concurrent access).
    Used to roll back a workers-database sync that failed partway through applying."""
    source = sqlite3.connect(source_path, timeout=30.0)
    dest = sqlite3.connect(DB_PATH)
    try:
        source.backup(dest)
    finally:
        dest.close()
        source.close()

async def run_db(func, *args, **kwargs):
    import asyncio
    return await asyncio.to_thread(func, *args, **kwargs)

def init_db():
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS workers (
            telegram_id INTEGER PRIMARY KEY,
            last_name TEXT NOT NULL,
            first_name TEXT NOT NULL,
            position TEXT NOT NULL DEFAULT 'Не указано',
            group_id INTEGER NOT NULL,
            schedule TEXT NOT NULL DEFAULT 'A',
            needs_daily_fact INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1,
            object_id TEXT NOT NULL DEFAULT 'Основной'
        )
        """
    )
    cols = {row["name"] for row in conn.execute("PRAGMA table_info(workers)").fetchall()}
    for col, definition in [
        ("schedule", "TEXT NOT NULL DEFAULT 'A'"),
        ("needs_daily_fact", "INTEGER NOT NULL DEFAULT 1"),
        ("sort_order", "INTEGER NOT NULL DEFAULT 0"),
        ("is_active", "INTEGER NOT NULL DEFAULT 1"),
        ("object_id", "TEXT NOT NULL DEFAULT 'Основной'"),
        ("registered_at", "TEXT"),
        ("last_remark_alert_count", "INTEGER NOT NULL DEFAULT 0"),
    ]:
        if col not in cols:
            conn.execute(f"ALTER TABLE workers ADD COLUMN {col} {definition}")

    conn.execute("CREATE TABLE IF NOT EXISTS objects (object_id TEXT PRIMARY KEY, group_id INTEGER NOT NULL DEFAULT 0)")
    cols_objects = {row["name"] for row in conn.execute("PRAGMA table_info(objects)").fetchall()}
    if "sort_order" not in cols_objects:
        conn.execute("ALTER TABLE objects ADD COLUMN sort_order INTEGER")
    conn.execute("CREATE TABLE IF NOT EXISTS groups (group_id INTEGER PRIMARY KEY, group_name TEXT NOT NULL)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER NOT NULL,
            report_date TEXT NOT NULL,
            report_type TEXT NOT NULL,
            slot_time TEXT,
            received_at TEXT NOT NULL,
            is_ok INTEGER NOT NULL,
            is_late INTEGER NOT NULL DEFAULT 0,
            format_comment TEXT,
            required_action TEXT,
            raw_text TEXT NOT NULL DEFAULT '',
            group_chat_id INTEGER,
            group_message_id INTEGER
        )
        """
    )
    cols_reports = {row["name"] for row in conn.execute("PRAGMA table_info(reports)").fetchall()}
    if "raw_text" not in cols_reports:
        conn.execute("ALTER TABLE reports ADD COLUMN raw_text TEXT NOT NULL DEFAULT ''")
    if "group_chat_id" not in cols_reports:
        conn.execute("ALTER TABLE reports ADD COLUMN group_chat_id INTEGER")
    if "group_message_id" not in cols_reports:
        conn.execute("ALTER TABLE reports ADD COLUMN group_message_id INTEGER")
    if "action_override" not in cols_reports:
        # Set when an admin manually types a custom "Требуемые действия" text (the "📋
        # Действия" button) - required_action then holds that exact text verbatim instead of
        # the auto-computed "ничего не предпринимать"/"сделано замечание.../делегировано..."
        # wording, until the admin edits it again.
        conn.execute("ALTER TABLE reports ADD COLUMN action_override INTEGER NOT NULL DEFAULT 0")
    if "batch_id" not in cols_reports:
        # Shared by every report/unrecognized_speech row that came from ONE submission
        # session (one process_media_batch call, one debounce-window burst of videos) - lets
        # a manually-resolved report (after a technical AI failure) know how many siblings
        # from the same session still need an admin's decision before the group message for
        # that type can be posted. Null for reports predating this feature or made through
        # paths that don't need session tracking (e.g. a single text report).
        conn.execute("ALTER TABLE reports ADD COLUMN batch_id TEXT")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS report_media (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER NOT NULL,
            source_chat_id INTEGER NOT NULL,
            source_message_id INTEGER NOT NULL,
            group_message_id INTEGER,
            position INTEGER NOT NULL,
            added_at TEXT NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_report_media_report ON report_media(report_id)")
    conn.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sent_reminders (
            telegram_id INTEGER,
            report_date TEXT,
            slot_time TEXT,
            PRIMARY KEY (telegram_id, report_date, slot_time)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sent_pre_reminders (
            telegram_id INTEGER,
            report_date TEXT,
            slot_time TEXT,
            PRIMARY KEY (telegram_id, report_date, slot_time)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS pending_reason_requests (
            telegram_id INTEGER,
            report_date TEXT,
            slot_time TEXT,
            requested_at TEXT,
            PRIMARY KEY (telegram_id, report_date, slot_time)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS missed_status_reasons (
            telegram_id INTEGER,
            report_date TEXT,
            slot_time TEXT,
            reason TEXT,
            PRIMARY KEY (telegram_id, report_date, slot_time)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS pending_unregistered_users (
            telegram_id INTEGER PRIMARY KEY,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            username TEXT NOT NULL,
            timestamp TEXT NOT NULL,
            text_content TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sheet_cell_sync (
            telegram_id INTEGER,
            report_date TEXT,
            slot_label TEXT,
            checked INTEGER,
            note TEXT,
            color_tag TEXT,
            PRIMARY KEY (telegram_id, report_date, slot_label)
        )
        """
    )
    # Tracks every admin DM sent for a report stuck at report_type='unrecognized_speech'
    # (transcription-quality gate failed) - lets the bot find and disable the "Пометить как
    # Статус/Факт" buttons in EVERY admin's chat once any one admin has resolved it, so two
    # admins can't both process the same report.
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS speech_review_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER NOT NULL,
            admin_chat_id INTEGER NOT NULL,
            message_id INTEGER NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_speech_review_report ON speech_review_messages(report_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_reports_date ON reports(report_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_reports_worker_date ON reports(telegram_id, report_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_workers_pos ON workers(position)")

    conn.commit()
    conn.close()

def get_all_departments() -> list[str]:
    """Union of formally-registered departments (objects table) and any department
    name already in use on a worker record, so nothing existing is ever hidden."""
    conn = get_db()
    from_objects = {r["object_id"] for r in conn.execute("SELECT object_id FROM objects").fetchall()}
    from_workers = {r["object_id"] for r in conn.execute("SELECT DISTINCT object_id FROM workers").fetchall() if r["object_id"]}
    conn.close()
    names = from_objects | from_workers | {"Основной"}
    return sorted(names)

def add_department(object_id: str):
    object_id = object_id.strip()
    conn = get_db()
    conn.execute(
        "INSERT INTO objects (object_id, group_id) VALUES (?, 0) ON CONFLICT(object_id) DO NOTHING",
        (object_id,)
    )
    conn.commit()
    conn.close()

def get_departments_ordered() -> tuple[list[str], list[str]]:
    """Returns (ordered_department_names, newly_appended_names). Any department that exists
    (per get_all_departments - including "Основной", which doesn't necessarily have its own
    objects row) but has no sort_order yet is assigned the next available rank, appended to
    the end, and persisted immediately - covers both a brand new department and the one-time
    backfill the first time this feature is used."""
    all_names = get_all_departments()
    conn = get_db()
    rows = conn.execute(
        "SELECT object_id, sort_order FROM objects WHERE sort_order IS NOT NULL ORDER BY sort_order"
    ).fetchall()
    ordered = [r["object_id"] for r in rows if r["object_id"] in all_names]
    known = set(ordered)
    new_names = sorted(n for n in all_names if n not in known)
    if new_names:
        next_rank = max((r["sort_order"] for r in rows), default=-1) + 1
        for i, name in enumerate(new_names):
            conn.execute(
                "INSERT INTO objects (object_id, group_id, sort_order) VALUES (?, 0, ?) "
                "ON CONFLICT(object_id) DO UPDATE SET sort_order=excluded.sort_order",
                (name, next_rank + i)
            )
        conn.commit()
        ordered.extend(new_names)
    conn.close()
    return ordered, new_names

def save_departments_order(ordered_names: list[str]):
    conn = get_db()
    for i, name in enumerate(ordered_names):
        conn.execute(
            "INSERT INTO objects (object_id, group_id, sort_order) VALUES (?, 0, ?) "
            "ON CONFLICT(object_id) DO UPDATE SET sort_order=excluded.sort_order",
            (name, i)
        )
    conn.commit()
    conn.close()

def count_workers_in_department(object_id: str) -> int:
    conn = get_db()
    row = conn.execute("SELECT COUNT(*) as c FROM workers WHERE object_id = ?", (object_id,)).fetchone()
    conn.close()
    return row["c"] if row else 0

def delete_department(object_id: str) -> bool:
    """Refuses to delete a department that still has workers assigned to it."""
    if count_workers_in_department(object_id) > 0:
        return False
    conn = get_db()
    conn.execute("DELETE FROM objects WHERE object_id = ?", (object_id,))
    conn.commit()
    conn.close()
    return True

def get_worker(telegram_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM workers WHERE telegram_id = ?", (telegram_id,)).fetchone()
    conn.close()
    return row

def get_all_workers():
    conn = get_db()
    rows = conn.execute("SELECT * FROM workers ORDER BY object_id, sort_order, last_name, first_name").fetchall()
    conn.close()
    return rows

def get_workers_by_position(position: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM workers WHERE lower(position) = lower(?) ORDER BY object_id, sort_order, last_name, first_name",
        (position,),
    ).fetchall()
    conn.close()
    return rows

def get_workers_by_object_id(object_id: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM workers WHERE lower(object_id) = lower(?) ORDER BY sort_order, last_name, first_name",
        (object_id,),
    ).fetchall()
    conn.close()
    return rows

def find_unregistered_workers_by_lastname(last_name: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM workers WHERE telegram_id < 0 AND LOWER(last_name) = LOWER(?)",
        (last_name.strip(),)
    ).fetchall()
    conn.close()
    return rows

def find_registered_workers_by_lastname(last_name: str):
    """Workers matching the name that are already bound to a real Telegram account —
    used to tell 'already registered' apart from 'not in the database at all'."""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM workers WHERE telegram_id > 0 AND LOWER(last_name) = LOWER(?)",
        (last_name.strip(),)
    ).fetchall()
    conn.close()
    return rows

def bind_worker_id(old_id: int, new_id: int, registered_at: str | None = None):
    conn = get_db()
    try:
        conn.execute(
            "UPDATE workers SET telegram_id = ?, registered_at = ? WHERE telegram_id = ?",
            (new_id, registered_at, old_id)
        )
        conn.execute("UPDATE reports SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE sent_reminders SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE sent_pre_reminders SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE pending_reason_requests SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE missed_status_reasons SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.execute("UPDATE pending_unregistered_users SET telegram_id = ? WHERE telegram_id = ?", (new_id, old_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

def save_pending_unregistered_user(telegram_id: int, first_name: str, last_name: str, username: str, timestamp: str, text_content: str):
    conn = get_db()
    conn.execute(
        """
        INSERT OR REPLACE INTO pending_unregistered_users (telegram_id, first_name, last_name, username, timestamp, text_content)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (telegram_id, first_name, last_name, username, timestamp, text_content)
    )
    conn.commit()
    conn.close()

def get_pending_unregistered_user(telegram_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM pending_unregistered_users WHERE telegram_id = ?", (telegram_id,)).fetchone()
    conn.close()
    return row

def delete_pending_unregistered_user(telegram_id: int):
    conn = get_db()
    conn.execute("DELETE FROM pending_unregistered_users WHERE telegram_id = ?", (telegram_id,))
    conn.commit()
    conn.close()

def upsert_worker(telegram_id: int, last_name: str, first_name: str, position: str, group_id: int, schedule: str, needs_daily_fact: bool, sort_order: int = 0, is_active: int = 1, object_id: str = 'Основной'):
    conn = get_db()
    conn.execute(
        """
        INSERT INTO workers (telegram_id, last_name, first_name, position, group_id, schedule, needs_daily_fact, sort_order, is_active, object_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(telegram_id) DO UPDATE SET
            last_name=excluded.last_name,
            first_name=excluded.first_name,
            position=excluded.position,
            group_id=excluded.group_id,
            schedule=excluded.schedule,
            needs_daily_fact=excluded.needs_daily_fact,
            sort_order=excluded.sort_order,
            is_active=excluded.is_active,
            object_id=excluded.object_id
        """,
        (telegram_id, last_name, first_name, position, group_id, schedule, int(needs_daily_fact), sort_order, is_active, object_id),
    )
    conn.commit()
    conn.close()

def get_object_group(object_id: str) -> int:
    conn = get_db()
    row = conn.execute("SELECT group_id FROM objects WHERE object_id = ?", (object_id,)).fetchone()
    conn.close()
    return row["group_id"] if row else 0

def save_object_group(object_id: str, group_id: int):
    conn = get_db()
    conn.execute(
        """
        INSERT INTO objects (object_id, group_id) VALUES (?, ?)
        ON CONFLICT(object_id) DO UPDATE SET group_id=excluded.group_id
        """,
        (object_id, group_id)
    )
    conn.commit()
    conn.close()

def get_setting(key: str, default: str = None) -> str:
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    if not row:
        return default
    value = row["value"]
    if key in ENCRYPTED_SETTING_KEYS:
        fernet = _get_fernet()
        if fernet:
            try:
                return fernet.decrypt(value.encode()).decode()
            except Exception:
                return value
    return value

def set_setting(key: str, value: str):
    stored_value = value
    if key in ENCRYPTED_SETTING_KEYS:
        fernet = _get_fernet()
        if fernet:
            stored_value = fernet.encrypt(value.encode()).decode()
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, stored_value)
    )
    conn.commit()
    conn.close()

def delete_worker(telegram_id: int) -> bool:
    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM report_media WHERE report_id IN (SELECT id FROM reports WHERE telegram_id = ?)",
            (telegram_id,)
        )
        conn.execute("DELETE FROM reports WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM sent_reminders WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM sent_pre_reminders WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM pending_reason_requests WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM missed_status_reasons WHERE telegram_id = ?", (telegram_id,))
        conn.execute("DELETE FROM pending_unregistered_users WHERE telegram_id = ?", (telegram_id,))
        cur = conn.execute("DELETE FROM workers WHERE telegram_id = ?", (telegram_id,))
        conn.commit()
        deleted = cur.rowcount > 0
    except Exception:
        conn.rollback()
        deleted = False
    finally:
        conn.close()
    return deleted

def save_report(telegram_id: int, report_date: str, report_type: str, slot_time: str | None, received_at: str, is_ok: bool, is_late: bool, format_comment: str, required_action: str, raw_text: str = "", batch_id: str | None = None) -> int:
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO reports (telegram_id, report_date, report_type, slot_time, received_at, is_ok, is_late, format_comment, required_action, raw_text, batch_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (telegram_id, report_date, report_type, slot_time, received_at, int(is_ok), int(is_late), format_comment, required_action, raw_text, batch_id),
    )
    conn.commit()
    inserted_id = cur.lastrowid
    conn.close()
    return inserted_id

def update_report_text_and_ai(report_id: int, is_ok: bool, format_comment: str, required_action: str, raw_text: str, received_at: str):
    conn = get_db()
    conn.execute(
        """
        UPDATE reports
        SET is_ok = ?, format_comment = ?, required_action = ?, raw_text = ?, received_at = ?
        WHERE id = ?
        """,
        (int(is_ok), format_comment, required_action, raw_text, received_at, report_id)
    )
    conn.commit()
    conn.close()

def resolve_unrecognized_report(report_id: int, report_type: str, slot_time: str | None, is_ok: bool, is_late: bool, format_comment: str, required_action: str):
    """Converts a report_type='unrecognized_speech' row into a normal status/daily_fact row
    once an admin has manually decided which one it is, after watching the original video."""
    conn = get_db()
    conn.execute(
        """
        UPDATE reports
        SET report_type = ?, slot_time = ?, is_ok = ?, is_late = ?, format_comment = ?, required_action = ?
        WHERE id = ?
        """,
        (report_type, slot_time, int(is_ok), int(is_late), format_comment, required_action, report_id)
    )
    conn.commit()
    conn.close()

def count_unresolved_batch_siblings(batch_id: str, telegram_id: int) -> int:
    """How many reports sharing this batch_id (the same submission session) are still stuck
    at report_type='unrecognized_speech' for this worker - used to decide whether a just-
    resolved report was the LAST one pending, and the session's group message(s) can finally
    be posted."""
    if not batch_id:
        return 0
    conn = get_db()
    row = conn.execute(
        "SELECT COUNT(*) as c FROM reports WHERE batch_id = ? AND telegram_id = ? AND report_type = 'unrecognized_speech'",
        (batch_id, telegram_id)
    ).fetchone()
    conn.close()
    return row["c"] if row else 0

def get_resolved_batch_reports_by_type(batch_id: str, telegram_id: int, report_type: str) -> list:
    """Reports from the same submission session that have ALREADY been resolved (by an
    admin, one at a time) to the given type - the set that gets combined into one group
    message once every sibling in the session has been labeled."""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM reports WHERE batch_id = ? AND telegram_id = ? AND report_type = ? ORDER BY id",
        (batch_id, telegram_id, report_type)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def merge_resolved_batch_reports(report_ids: list[int], required_action: str | None = None) -> int:
    """Combines multiple already-saved `reports` rows that were just resolved to the SAME
    type (status or daily_fact) into ONE row - mirrors how process_media_batch already
    combines multiple same-type videos from one auto-classified session into a single row
    with a "Видео N: ..." joined format_comment/raw_text, so a manually-resolved session ends
    up represented identically (one row, one editable report card) instead of N disconnected
    ones. The first id becomes the primary (keeps its own row, its already-computed slot_time/
    is_late/received_at); the rest have their report_media reassigned to it and are then
    deleted. required_action, if given, is stored on the merged row too (the caller computes
    it from the combined verdict - db.py has no access to the project's fixed action-text
    constants). Returns the primary report_id - a no-op returning report_ids[0] if there's
    only one id (nothing to merge)."""
    if len(report_ids) <= 1:
        return report_ids[0] if report_ids else None

    conn = get_db()
    rows = [conn.execute("SELECT * FROM reports WHERE id = ?", (rid,)).fetchone() for rid in report_ids]
    rows = [dict(r) for r in rows if r]
    if not rows:
        conn.close()
        return report_ids[0]

    primary = rows[0]
    raw_text_parts = [f"[Видео {i}]: {r['raw_text']}" for i, r in enumerate(rows, start=1)]
    combined_raw_text = "\n".join(raw_text_parts)
    # Each row's own format_comment already carries its "ОК - ..."/"не ОК - ..." prefix
    # (baked in by ai.py's check_status, same as every non-batch report) - just number them,
    # don't add a second prefix on top or parse_video_comments would see a doubled-up label.
    comment_parts = [f"Видео {i}: {r['format_comment'] or ('ОК' if r['is_ok'] else 'не ОК')}" for i, r in enumerate(rows, start=1)]
    combined_format_comment = "; ".join(comment_parts)
    # Matches process_media_batch's own multi-video status/fact aggregation convention
    # exactly (ANY video OK => overall row marked is_ok) - required_action, computed by the
    # caller from the real error count, is what actually drives "needs correction" downstream.
    combined_is_ok = any(bool(r["is_ok"]) for r in rows)

    if required_action is not None:
        conn.execute(
            "UPDATE reports SET raw_text = ?, format_comment = ?, is_ok = ?, required_action = ? WHERE id = ?",
            (combined_raw_text, combined_format_comment, int(combined_is_ok), required_action, primary["id"])
        )
    else:
        conn.execute(
            "UPDATE reports SET raw_text = ?, format_comment = ?, is_ok = ? WHERE id = ?",
            (combined_raw_text, combined_format_comment, int(combined_is_ok), primary["id"])
        )
    for other in rows[1:]:
        conn.execute("UPDATE report_media SET report_id = ? WHERE report_id = ?", (primary["id"], other["id"]))
        conn.execute("DELETE FROM reports WHERE id = ?", (other["id"],))
    conn.commit()
    conn.close()
    return primary["id"]

def count_consecutive_unrecognized_reports(telegram_id: int) -> int:
    """How many of this worker's MOST RECENT reports (by insertion order) are, one after
    another, report_type='unrecognized_speech' - stops counting at the first report of any
    other type. Used to detect a run of back-to-back failed recognitions (bad mic/recording
    conditions) without needing a separate persisted streak counter."""
    conn = get_db()
    rows = conn.execute(
        "SELECT report_type FROM reports WHERE telegram_id = ? ORDER BY id DESC LIMIT 20",
        (telegram_id,)
    ).fetchall()
    conn.close()
    count = 0
    for r in rows:
        if r["report_type"] == "unrecognized_speech":
            count += 1
        else:
            break
    return count

def save_speech_review_message(report_id: int, admin_chat_id: int, message_id: int):
    conn = get_db()
    conn.execute(
        "INSERT INTO speech_review_messages (report_id, admin_chat_id, message_id) VALUES (?, ?, ?)",
        (report_id, admin_chat_id, message_id)
    )
    conn.commit()
    conn.close()

def get_speech_review_messages(report_id: int) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT admin_chat_id, message_id FROM speech_review_messages WHERE report_id = ?",
        (report_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def set_report_media_group_message(media_id: int, group_message_id: int):
    conn = get_db()
    conn.execute("UPDATE report_media SET group_message_id = ? WHERE id = ?", (group_message_id, media_id))
    conn.commit()
    conn.close()

def count_remarks(telegram_id: int) -> int:
    """Total number of video-status reports with a remark (is_ok = 0) for this worker, all-time."""
    conn = get_db()
    row = conn.execute(
        "SELECT COUNT(*) as c FROM reports WHERE telegram_id = ? AND report_type = 'status' AND is_ok = 0",
        (telegram_id,)
    ).fetchone()
    conn.close()
    return row["c"] if row else 0

def count_effective_remarks(telegram_id: int) -> int:
    """Cumulative, never-reset count of this worker's status/daily_fact reports that are
    either content-not-ok OR late - the basis for the "Требуемые действия" escalation shown
    in the group message ("сделано замечание" below 3, "делегировано отделу контроля" once
    it reaches 3 - sticky, stays "делегировано" on every report after that too). Deliberately
    separate from count_remarks (content-only, status-only) since a late-but-content-ok
    report is a NEW category of "не ОК" for THIS display only, per product decision - it does
    NOT change the stored is_ok value or feed the existing count_remarks/remark-alert-
    threshold admin notification, both of which stay exactly as they were."""
    conn = get_db()
    row = conn.execute(
        "SELECT COUNT(*) as c FROM reports WHERE telegram_id = ? AND report_type IN ('status', 'daily_fact') "
        "AND (is_ok = 0 OR is_late = 1)",
        (telegram_id,)
    ).fetchone()
    conn.close()
    return row["c"] if row else 0

def set_report_action_override(report_id: int, text: str):
    conn = get_db()
    conn.execute(
        "UPDATE reports SET required_action = ?, action_override = 1 WHERE id = ?",
        (text, report_id)
    )
    conn.commit()
    conn.close()

def get_recent_remarks(telegram_id: int, limit: int = 5) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT report_date, slot_time, format_comment FROM reports "
        "WHERE telegram_id = ? AND report_type = 'status' AND is_ok = 0 "
        "ORDER BY report_date DESC, id DESC LIMIT ?",
        (telegram_id, limit)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def check_and_update_remark_alert_threshold(telegram_id: int, threshold: int = 3) -> int | None:
    """Returns the new total remark count if the worker just crossed another multiple of
    `threshold` since the last alert, otherwise None. Updates the bookkeeping either way
    is not needed on None; only advances last_remark_alert_count when an alert should fire,
    so a worker can never be re-notified for the same tier."""
    conn = get_db()
    row = conn.execute(
        "SELECT last_remark_alert_count FROM workers WHERE telegram_id = ?", (telegram_id,)
    ).fetchone()
    if row is None:
        conn.close()
        return None
    last_alerted = row["last_remark_alert_count"] or 0

    total = conn.execute(
        "SELECT COUNT(*) as c FROM reports WHERE telegram_id = ? AND report_type = 'status' AND is_ok = 0",
        (telegram_id,)
    ).fetchone()["c"]

    if total >= last_alerted + threshold:
        new_tier = (total // threshold) * threshold
        conn.execute("UPDATE workers SET last_remark_alert_count = ? WHERE telegram_id = ?", (new_tier, telegram_id))
        conn.commit()
        conn.close()
        return total

    conn.close()
    return None

def set_report_group_message(report_id: int, chat_id: int, message_id: int):
    conn = get_db()
    conn.execute(
        "UPDATE reports SET group_chat_id = ?, group_message_id = ? WHERE id = ?",
        (chat_id, message_id, report_id)
    )
    conn.commit()
    conn.close()

def get_report_group_message(report_id: int):
    conn = get_db()
    row = conn.execute("SELECT group_chat_id, group_message_id FROM reports WHERE id = ?", (report_id,)).fetchone()
    conn.close()
    return row

def get_submitted_status_slots(telegram_id: int, report_date: str) -> set:
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT slot_time FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'status'",
        (telegram_id, report_date)
    ).fetchall()
    conn.close()
    return {r["slot_time"] for r in rows}

def has_pre_reminder_sent(telegram_id: int, report_date: str, slot_time: str) -> bool:
    conn = get_db()
    row = conn.execute(
        "SELECT 1 FROM sent_pre_reminders WHERE telegram_id = ? AND report_date = ? AND slot_time = ?",
        (telegram_id, report_date, slot_time)
    ).fetchone()
    conn.close()
    return row is not None

def mark_pre_reminder_sent(telegram_id: int, report_date: str, slot_time: str):
    conn = get_db()
    conn.execute(
        "INSERT OR IGNORE INTO sent_pre_reminders (telegram_id, report_date, slot_time) VALUES (?, ?, ?)",
        (telegram_id, report_date, slot_time)
    )
    conn.commit()
    conn.close()

def has_pending_reason_request(telegram_id: int, report_date: str, slot_time: str) -> bool:
    conn = get_db()
    row = conn.execute(
        "SELECT 1 FROM pending_reason_requests WHERE telegram_id = ? AND report_date = ? AND slot_time = ?",
        (telegram_id, report_date, slot_time)
    ).fetchone()
    conn.close()
    return row is not None

def create_pending_reason_request(telegram_id: int, report_date: str, slot_time: str, requested_at: str):
    conn = get_db()
    conn.execute(
        "INSERT OR IGNORE INTO pending_reason_requests (telegram_id, report_date, slot_time, requested_at) VALUES (?, ?, ?, ?)",
        (telegram_id, report_date, slot_time, requested_at)
    )
    conn.commit()
    conn.close()

def get_pending_reason_requests(telegram_id: int) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT report_date, slot_time FROM pending_reason_requests WHERE telegram_id = ? ORDER BY requested_at",
        (telegram_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def resolve_pending_reason_requests(telegram_id: int, reason: str):
    conn = get_db()
    pending = conn.execute(
        "SELECT report_date, slot_time FROM pending_reason_requests WHERE telegram_id = ?",
        (telegram_id,)
    ).fetchall()
    for p in pending:
        conn.execute(
            "INSERT INTO missed_status_reasons (telegram_id, report_date, slot_time, reason) VALUES (?, ?, ?, ?) "
            "ON CONFLICT(telegram_id, report_date, slot_time) DO UPDATE SET reason=excluded.reason",
            (telegram_id, p["report_date"], p["slot_time"], reason)
        )
    conn.execute("DELETE FROM pending_reason_requests WHERE telegram_id = ?", (telegram_id,))
    conn.commit()
    conn.close()
    return [dict(p) for p in pending]

def get_missed_status_reason(telegram_id: int, report_date: str, slot_time: str) -> str | None:
    conn = get_db()
    row = conn.execute(
        "SELECT reason FROM missed_status_reasons WHERE telegram_id = ? AND report_date = ? AND slot_time = ?",
        (telegram_id, report_date, slot_time)
    ).fetchone()
    conn.close()
    return row["reason"] if row else None

def get_existing_report_row(telegram_id: int, report_date: str, report_type: str, slot_time: str | None = None):
    conn = get_db()
    if report_type == "status":
        row = conn.execute(
            "SELECT * FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'status' AND slot_time = ?",
            (telegram_id, report_date, slot_time)
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT * FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'daily_fact'",
            (telegram_id, report_date)
        ).fetchone()
    conn.close()
    return row

def add_report_media(report_id: int, source_chat_id: int, source_message_id: int, group_message_id: int | None, position: int, added_at: str):
    conn = get_db()
    conn.execute(
        "INSERT INTO report_media (report_id, source_chat_id, source_message_id, group_message_id, position, added_at) VALUES (?, ?, ?, ?, ?, ?)",
        (report_id, source_chat_id, source_message_id, group_message_id, position, added_at)
    )
    conn.commit()
    conn.close()

def get_report_media(report_id: int):
    conn = get_db()
    rows = conn.execute("SELECT * FROM report_media WHERE report_id = ? ORDER BY position", (report_id,)).fetchall()
    conn.close()
    return rows

def is_message_already_processed(source_chat_id: int, source_message_id: int) -> bool:
    """True if this exact Telegram message was already turned into a report (report_media
    row exists for it) — guards against the same update being redelivered and processed
    twice, e.g. if the bot restarts mid-flight before Telegram's update offset is advanced."""
    conn = get_db()
    row = conn.execute(
        "SELECT 1 FROM report_media WHERE source_chat_id = ? AND source_message_id = ? LIMIT 1",
        (source_chat_id, source_message_id)
    ).fetchone()
    conn.close()
    return row is not None

def delete_report_media_rows(report_id: int):
    conn = get_db()
    conn.execute("DELETE FROM report_media WHERE report_id = ?", (report_id,))
    conn.commit()
    conn.close()

def cancel_not_working(telegram_id: int, report_date: str):
    conn = get_db()
    conn.execute("DELETE FROM reports WHERE telegram_id = ? AND report_date = ? AND report_type = 'not_working'", (telegram_id, report_date))
    conn.commit()
    conn.close()

def save_group_name(group_id: int, group_name: str):
    conn = get_db()
    conn.execute(
        "INSERT INTO groups (group_id, group_name) VALUES (?, ?) ON CONFLICT(group_id) DO UPDATE SET group_name=excluded.group_name",
        (group_id, group_name),
    )
    conn.commit()
    conn.close()

def calculate_worker_stats(worker, start_date_str: str, end_date_str: str, conn) -> dict:
    join_row = conn.execute("SELECT MIN(report_date) as start_date FROM reports WHERE telegram_id = ?", (worker["telegram_id"],)).fetchone()
    adjusted_start_date_str = start_date_str
    if join_row and join_row["start_date"] and join_row["start_date"] > start_date_str:
        adjusted_start_date_str = join_row["start_date"]

    reports = conn.execute(
        "SELECT * FROM reports WHERE telegram_id = ? AND report_date >= ? AND report_date <= ?",
        (worker["telegram_id"], adjusted_start_date_str, end_date_str)
    ).fetchall()

    not_working_dates = set()
    submitted_statuses = set()
    submitted_facts = set()
    total_lates = 0
    total_remarks = 0

    for r in reports:
        rep_date = r["report_date"]
        rep_type = r["report_type"]
        if rep_type == "not_working":
            not_working_dates.add(rep_date)
        elif rep_type == "status":
            submitted_statuses.add((rep_date, r["slot_time"]))
            if r["is_late"]: total_lates += 1
            if not r["is_ok"]: total_remarks += 1
        elif rep_type == "daily_fact":
            submitted_facts.add(rep_date)
            if r["is_late"]: total_lates += 1
            if not r["is_ok"]: total_remarks += 1

    start_dt = datetime.strptime(adjusted_start_date_str, "%Y-%m-%d").date()
    end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    expected_status_count = 0
    expected_fact_count = 0
    current_date_it = start_dt
    now = now_local()
    now_date_str = now.strftime("%Y-%m-%d")
    current_mins = now.hour * 60 + now.minute

    while current_date_it <= end_dt:
        date_it_str = current_date_it.strftime("%Y-%m-%d")
        if date_it_str not in not_working_dates:
            slots = SCHEDULES.get(worker["schedule"], SCHEDULE_A)
            if date_it_str == now_date_str:
                for slot in slots:
                    hour, minute = map(int, slot.split(":"))
                    if current_mins > hour * 60 + minute + LATE_THRESHOLD_MIN or (date_it_str, slot) in submitted_statuses:
                        expected_status_count += 1
                if worker["needs_daily_fact"]:
                    if date_it_str in submitted_facts:
                        expected_fact_count += 1
                    else:
                        last_slot = slots[-1]
                        hour, minute = map(int, last_slot.split(":"))
                        if current_mins > hour * 60 + minute + 60:
                            expected_fact_count += 1
            else:
                expected_status_count += len(slots)
                if worker["needs_daily_fact"]:
                    expected_fact_count += 1
        current_date_it += dt_module.timedelta(days=1)

    total_expected = expected_status_count + expected_fact_count
    total_submitted = len(submitted_statuses) + len(submitted_facts)
    missed_count = max(0, total_expected - total_submitted)
    percent_submitted = (total_submitted / total_expected * 100.0) if total_expected > 0 else 100.0

    return {
        "expected": total_expected,
        "submitted": total_submitted,
        "lates": total_lates,
        "remarks": total_remarks,
        "percent": percent_submitted,
        "missed": missed_count,
    }

def is_quiet_mode_enabled() -> bool:
    try:
        conn = get_db()
        row = conn.execute("SELECT value FROM settings WHERE key = 'quiet_mode_enabled'").fetchone()
        conn.close()
        if row:
            return row["value"] == "1"
    except Exception:
        pass
    return False

def set_quiet_mode(enabled: bool):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES ('quiet_mode_enabled', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        ("1" if enabled else "0",)
    )
    conn.commit()
    conn.close()

def is_ai_error_simulation_enabled() -> bool:
    """Temporary QA toggle (/simulate_ai_error) - while on, check_status_async/transcribe_
    audio_async (report_handlers.py) raise AITechnicalError immediately instead of calling
    the real LLM, so the whole "technical failure -> manual admin review" pipeline can be
    tested on demand without waiting for a real rate limit. Affects EVERY worker's videos
    while enabled - meant to be turned off again right after testing."""
    try:
        conn = get_db()
        row = conn.execute("SELECT value FROM settings WHERE key = 'ai_error_simulation_enabled'").fetchone()
        conn.close()
        if row:
            return row["value"] == "1"
    except Exception:
        pass
    return False

def set_ai_error_simulation(enabled: bool):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES ('ai_error_simulation_enabled', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        ("1" if enabled else "0",)
    )
    conn.commit()
    conn.close()

def is_missed_reason_request_enabled() -> bool:
    """Whether a worker who misses a status slot is asked to explain why (and blocked from
    sending a new video until they do) - disabled by default for now, per product decision,
    toggled from the Settings menu."""
    try:
        conn = get_db()
        row = conn.execute("SELECT value FROM settings WHERE key = 'missed_reason_request_enabled'").fetchone()
        conn.close()
        if row:
            return row["value"] == "1"
    except Exception:
        pass
    return False

def set_missed_reason_request_enabled(enabled: bool):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES ('missed_reason_request_enabled', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        ("1" if enabled else "0",)
    )
    conn.commit()
    conn.close()

def get_scheduled_times() -> list[str]:
    try:
        val = get_setting("scheduled_summary_times")
        if val:
            return json.loads(val)
    except Exception:
        pass
    return ["19:00"]

def save_scheduled_times(times: list[str]):
    set_setting("scheduled_summary_times", json.dumps(times))

def get_worker_target_group(worker) -> int:
    try:
        w_dict = dict(worker)
    except (TypeError, ValueError):
        return DEFAULT_GROUP_ID
    if "object_id" in w_dict and w_dict["object_id"]:
        obj_group = get_object_group(w_dict["object_id"])
        if obj_group and obj_group != 0:
            return obj_group
    return w_dict.get("group_id") or DEFAULT_GROUP_ID

def get_group_name(group_id: int) -> str:
    conn = get_db()
    row = conn.execute("SELECT group_name FROM groups WHERE group_id = ?", (group_id,)).fetchone()
    conn.close()
    return row["group_name"] if row else str(group_id)

def _get_group_name_row(group_id: int):
    conn = get_db()
    row = conn.execute("SELECT group_name FROM groups WHERE group_id = ?", (group_id,)).fetchone()
    conn.close()
    return row

async def get_group_name_async(bot, group_id: int) -> str:
    row = await run_db(_get_group_name_row, group_id)
    if row:
        return row["group_name"]
    try:
        chat = await bot.get_chat(group_id)
        name = chat.title or str(group_id)
        await run_db(save_group_name, group_id, name)
        return name
    except Exception:
        return str(group_id)

async def fetch_and_save_group_name(bot, group_id: int) -> str:
    try:
        chat = await bot.get_chat(group_id)
        name = chat.title or str(group_id)
    except Exception:
        name = str(group_id)
    await run_db(save_group_name, group_id, name)
    return name

def get_all_group_names() -> dict:
    conn = get_db()
    rows = conn.execute("SELECT group_id, group_name FROM groups").fetchall()
    conn.close()
    res = {row["group_id"]: row["group_name"] for row in rows}
    if DEFAULT_GROUP_ID not in res:
        res[DEFAULT_GROUP_ID] = str(DEFAULT_GROUP_ID)
    return res

def get_next_sort_order(position: str) -> int:
    conn = get_db()
    row = conn.execute("SELECT MAX(sort_order) as max_order FROM workers WHERE lower(position) = lower(?)", (position,)).fetchone()
    conn.close()
    if row and row["max_order"] is not None:
        return row["max_order"] + 1
    return 0

def update_worker_field(telegram_id: int, field_name: str, value):
    conn = get_db()
    allowed_fields = {"last_name", "first_name", "position", "group_id", "schedule", "needs_daily_fact", "sort_order", "is_active", "object_id"}
    if field_name in allowed_fields:
        conn.execute(f"UPDATE workers SET {field_name} = ? WHERE telegram_id = ?", (value, telegram_id))
        conn.commit()
    conn.close()

def get_violators_threshold() -> int:
    try:
        val = get_setting("violators_threshold")
        if val is not None:
            return int(val)
    except Exception:
        pass
    return 3

def save_violators_threshold(val: int):
    set_setting("violators_threshold", str(val))

def generate_and_send_excel(*args, **kwargs):
    pass

def generate_and_send_gsheets(*args, **kwargs):
    pass

def fetch_export_data():
    return get_all_workers()

def export_reports_to_excel() -> bytes:
    import openpyxl
    from openpyxl import Workbook
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    
    headers = [
        "ID отчета", "ID сотрудника", "Фамилия", "Имя", "Должность (Отдел)",
        "Дата отчета", "Тип отчета", "Время отчёта", "Время получения",
        "Оценка (ОК)", "Опоздание", "Замечания", "Действия", "Оригинальный отчёт"
    ]
    ws.append(headers)
    
    conn = get_db()
    reports = conn.execute("SELECT * FROM reports ORDER BY report_date DESC, id DESC").fetchall()
    
    # Cache worker profiles
    workers_cache = {}
    
    for r in reports:
        t_id = r["telegram_id"]
        if t_id not in workers_cache:
            w_row = conn.execute("SELECT last_name, first_name, position, object_id FROM workers WHERE telegram_id = ?", (t_id,)).fetchone()
            if w_row:
                position_display = f"{clean_position(w_row['position'])} ({w_row['object_id'] or 'Основной'})"
                workers_cache[t_id] = (w_row["last_name"], w_row["first_name"], position_display)
            else:
                workers_cache[t_id] = ("Неизвестно", "Неизвестно", "Неизвестно")

        last_name, first_name, position = workers_cache[t_id]
        
        ws.append([
            r["id"],
            r["telegram_id"],
            last_name,
            first_name,
            position,
            r["report_date"],
            r["report_type"],
            r["slot_time"] or "",
            r["received_at"],
            "Да" if r["is_ok"] else "Нет",
            "Да" if r["is_late"] else "Нет",
            r["format_comment"] or "",
            r["required_action"] or "",
            r["raw_text"] or ""
        ])
    
    conn.close()
    
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def export_workers_to_excel() -> bytes:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    headers = [
        "Telegram ID", "Фамилия", "Имя", "Должность",
        "ID чата (уведомления)", "График (А или Б)", "Итоговый отчет за день",
        "Статус", "Номер в списке", "Объект"
    ]
    hints = [
        "не менять", "фамилия", "имя", "должность",
        "чат id", "А или Б", "Да/Нет",
        "Работает/Отпуск/Больничный", "число", "объект"
    ]
    ws.append(headers)
    ws.append(hints)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    hint_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    hint_font = Font(italic=True, color="555555")
    warn_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    for col in range(1, len(headers) + 1):
        c1 = ws.cell(row=1, column=col)
        c1.fill = header_fill
        c1.font = header_font
        c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2 = ws.cell(row=2, column=col)
        c2.fill = hint_fill
        c2.font = hint_font
        c2.alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1).fill = warn_fill

    ws.freeze_panes = "A3"
    for i, width in enumerate([14, 16, 14, 22, 18, 12, 16, 20, 10, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # get_all_workers() already orders by object_id first (then sort_order/last_name/
    # first_name), so same-object rows are already contiguous - no re-sorting needed here,
    # just detect the boundary between one object's rows and the next.
    workers = get_all_workers()
    separator_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    prev_object_id = None
    for i, w in enumerate(workers):
        if i > 0 and w["object_id"] != prev_object_id:
            ws.append([None] * len(headers))
            sep_row = ws.max_row
            for col in range(1, len(headers) + 1):
                ws.cell(row=sep_row, column=col).fill = separator_fill

        ws.append([
            w["telegram_id"],
            w["last_name"],
            w["first_name"],
            w["position"],
            w["group_id"],
            w["schedule"],
            "Да" if w["needs_daily_fact"] else "Нет",
            "Работает" if w["is_active"] else "Отпуск",
            w["sort_order"],
            w["object_id"],
        ])
        prev_object_id = w["object_id"]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

EXCEL_HEADER_ALIASES = {
    "telegram id": "telegram_id",
    "фамилия": "last_name",
    "имя": "first_name",
    "должность": "position",
    "должность / отдел": "position",
    "id чата (уведомления)": "group_id",
    "id чата": "group_id",
    "график (а или б)": "schedule",
    "график": "schedule",
    "итоговый отчет за день": "needs_daily_fact",
    "итоговый отчёт за день": "needs_daily_fact",
    "статус": "is_active",
    "номер в списке": "sort_order",
    "объект": "object_id",
}

def _normalize_worker_row(row_dict: dict) -> dict | None:
    """Shared row-normalization rules between read_excel (existing bulk-import feature) and
    parse_workers_excel_bytes (the diff-based sync feature) - one place to fix a parsing rule
    so both stay in sync instead of silently drifting apart. Returns None if the row has no
    usable Telegram ID (caller should skip it)."""
    if "telegram_id" not in row_dict or row_dict["telegram_id"] is None:
        return None
    try:
        row_dict["telegram_id"] = int(row_dict["telegram_id"])
    except (ValueError, TypeError):
        return None

    row_dict["last_name"] = str(row_dict.get("last_name", "") or "").strip()
    row_dict["first_name"] = str(row_dict.get("first_name", "") or "").strip()
    row_dict["position"] = clean_position(str(row_dict.get("position", "Не указано") or "Не указано"))

    try:
        row_dict["group_id"] = int(row_dict.get("group_id", DEFAULT_GROUP_ID) or DEFAULT_GROUP_ID)
    except (ValueError, TypeError):
        row_dict["group_id"] = DEFAULT_GROUP_ID

    row_dict["schedule"] = str(row_dict.get("schedule", "A") or "A").strip().upper()
    if row_dict["schedule"] not in SCHEDULES:
        row_dict["schedule"] = "A"

    ndf = row_dict.get("needs_daily_fact")
    row_dict["needs_daily_fact"] = ndf in (True, 1, "1", "Да", "да", "yes", "YES", "True", "true")

    row_dict["object_id"] = str(row_dict.get("object_id", "Основной") or "Основной").strip()

    is_act = row_dict.get("is_active")
    if isinstance(is_act, str):
        not_working_labels = {"отпуск", "больничный", "нет", "no", "false", "0", "не работает"}
        row_dict["is_active"] = is_act.strip().lower() not in not_working_labels
    elif is_act in (False, 0):
        row_dict["is_active"] = False
    else:
        row_dict["is_active"] = True

    try:
        row_dict["sort_order"] = int(row_dict.get("sort_order", 0) or 0)
    except (ValueError, TypeError):
        row_dict["sort_order"] = 0

    return row_dict

def read_excel(file_path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    raw_headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    headers = [EXCEL_HEADER_ALIASES.get(h.lower(), h) for h in raw_headers]

    data_rows = rows[1:]
    if data_rows and data_rows[0]:
        first_cell = data_rows[0][0]
        if isinstance(first_cell, str) and not first_cell.strip().lstrip("-").isdigit():
            data_rows = data_rows[1:]  # skip the "не менять / фамилия / имя / ..." hint row

    workers = []
    for row in data_rows:
        if not any(row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(row) and h:
                row_dict[h] = row[i]

        normalized = _normalize_worker_row(row_dict)
        if normalized:
            workers.append(normalized)

    return workers

REQUIRED_WORKERS_COLUMNS = [
    ("telegram_id", "Telegram ID"),
    ("last_name", "Фамилия"),
    ("first_name", "Имя"),
    ("position", "Должность"),
    ("group_id", "ID чата (уведомления)"),
    ("schedule", "График (А или Б)"),
    ("object_id", "Объект"),
]

def parse_workers_excel_bytes(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Like read_excel, but reads from in-memory bytes (no temp file - avoids the filename-
    collision race read_excel's caller has when two admins import at once) and validates the
    file BEFORE returning any rows: every required column present (matched by header NAME,
    not position - an admin can reorder columns freely) and no duplicate Telegram ID within
    the file itself. Returns (rows, errors) - a non-empty errors list means the file was
    rejected outright and rows is always []; nothing about the database is touched here."""
    import io
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return [], [f"Не удалось открыть файл как Excel (.xlsx): {e}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Файл пустой."]

    raw_headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    headers = [EXCEL_HEADER_ALIASES.get(h.lower(), h) for h in raw_headers]

    missing_columns = [label for key, label in REQUIRED_WORKERS_COLUMNS if key not in headers]
    if missing_columns:
        return [], ["Отсутствует колонка «" + "», «".join(missing_columns) + "»"]

    data_rows = rows[1:]
    if data_rows and data_rows[0]:
        first_cell = data_rows[0][0]
        if isinstance(first_cell, str) and not first_cell.strip().lstrip("-").isdigit():
            data_rows = data_rows[1:]

    seen_ids = set()
    dup_ids = []
    workers = []
    for row in data_rows:
        if not any(row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(row) and h:
                row_dict[h] = row[i]

        raw_id = row_dict.get("telegram_id")
        if raw_id is None:
            continue
        try:
            tg_id = int(raw_id)
        except (ValueError, TypeError):
            continue

        if tg_id in seen_ids and tg_id not in dup_ids:
            dup_ids.append(tg_id)
        seen_ids.add(tg_id)

        normalized = _normalize_worker_row(row_dict)
        if normalized:
            workers.append(normalized)

    if dup_ids:
        return [], ["Обнаружен дублирующийся Telegram ID внутри файла: " + ", ".join(str(i) for i in dup_ids)]

    return workers, []

_WORKERS_DIFF_FIELDS = [
    ("last_name", "Фамилия"),
    ("first_name", "Имя"),
    ("position", "Должность"),
    ("group_id", "ID чата"),
    ("schedule", "График"),
    ("needs_daily_fact", "Итоговый отчёт за день"),
    ("is_active", "Статус"),
    ("sort_order", "Номер в списке"),
    ("object_id", "Объект"),
]

def _format_worker_field_value(key: str, value) -> str:
    if key == "needs_daily_fact":
        return "Да" if value else "Нет"
    if key == "is_active":
        return "Работает" if value else "Отпуск"
    return str(value)

def _diff_worker_fields(old_row, new_row) -> list[tuple[str, str, str]]:
    diffs = []
    for key, label in _WORKERS_DIFF_FIELDS:
        old_val = old_row[key]
        new_val = new_row[key]
        if key in ("needs_daily_fact", "is_active"):
            old_val, new_val = bool(old_val), bool(new_val)
        if old_val != new_val:
            diffs.append((label, _format_worker_field_value(key, old_val), _format_worker_field_value(key, new_val)))
    return diffs

def compute_workers_diff(new_rows: list[dict]) -> dict:
    """Compares freshly-parsed Excel rows against the current `workers` table (the source of
    truth - Google Sheets is a one-way mirror of it, so comparisons are never made against
    Sheets) by Telegram ID. "missing" workers (in the DB but not the uploaded file) are only
    ever reported here, never auto-deleted - deletion happens only via a separate, explicit
    admin confirmation (see apply_workers_sync)."""
    current = {w["telegram_id"]: dict(w) for w in get_all_workers()}
    new_by_id = {w["telegram_id"]: w for w in new_rows}

    new_workers = []
    changed_workers = []
    for tg_id, new_row in new_by_id.items():
        if tg_id not in current:
            new_workers.append(new_row)
        else:
            diffs = _diff_worker_fields(current[tg_id], new_row)
            if diffs:
                changed_workers.append({"telegram_id": tg_id, "old": current[tg_id], "new": new_row, "diffs": diffs})

    missing_workers = [w for tg_id, w in current.items() if tg_id not in new_by_id]

    return {"new": new_workers, "changed": changed_workers, "missing": missing_workers}

def apply_workers_sync(new_workers: list[dict], changed_workers: list[dict], missing_ids_to_delete: list[int]) -> dict:
    """Applies a previously-computed workers diff - upserts new+changed rows via the same
    upsert_worker the existing bulk-import feature already uses, and fully removes any
    worker in missing_ids_to_delete via the existing delete_worker (the same thorough
    cleanup of their reports/media/reminders as the manual "🗑 Удалить сотрудника" admin
    action - not a separate, narrower deletion path). The caller is responsible for taking a
    backup first and restoring it if this raises."""
    for w in new_workers + [item["new"] for item in changed_workers]:
        upsert_worker(
            telegram_id=w["telegram_id"], last_name=w["last_name"], first_name=w["first_name"],
            position=w["position"], group_id=w["group_id"], schedule=w["schedule"],
            needs_daily_fact=w["needs_daily_fact"], sort_order=w["sort_order"],
            is_active=1 if w["is_active"] else 0, object_id=w["object_id"]
        )
    deleted_count = 0
    for tg_id in missing_ids_to_delete:
        if delete_worker(tg_id):
            deleted_count += 1
    return {"added": len(new_workers), "changed": len(changed_workers), "deleted": deleted_count}

# Known Сводка cell background colors the bot itself paints (see sync_gsheets_task) - used
# to tell "this color is one the bot could have set" apart from a manually chosen highlight.
_KNOWN_CELL_COLOR_TAGS = {
    "yellow": {"red": 1.0, "green": 0.95, "blue": 0.6},
    "pink": {"red": 0.96, "green": 0.78, "blue": 0.78},
    "grey": {"red": 0.85, "green": 0.85, "blue": 0.85},
}

def _color_to_tag(color: dict | None) -> str:
    """Turns a Sheets backgroundColor object into "none"/"yellow"/"pink"/"grey", or a
    "custom:r,g,b" tag for any other color a human chose by hand, so it can be told apart
    from the bot's own palette and restored exactly if it needs to be preserved."""
    if not color:
        return "none"
    for tag, known in _KNOWN_CELL_COLOR_TAGS.items():
        if all(abs(color.get(k, 0) - known[k]) < 0.01 for k in ("red", "green", "blue")):
            return tag
    r, g, b = color.get("red", 1), color.get("green", 1), color.get("blue", 1)
    return f"custom:{r:.3f},{g:.3f},{b:.3f}"

def _color_tag_to_dict(tag: str | None) -> dict | None:
    if not tag or tag == "none":
        return None
    if tag in _KNOWN_CELL_COLOR_TAGS:
        return _KNOWN_CELL_COLOR_TAGS[tag]
    if tag.startswith("custom:"):
        try:
            r, g, b = (float(x) for x in tag[len("custom:"):].split(","))
            return {"red": r, "green": g, "blue": b}
        except (ValueError, IndexError):
            return None
    return None

def _resolve_manual_preserving_cell(sync_map: dict, old_entry: dict | None, key: tuple,
                                     fresh_checked: bool, fresh_note: str | None, fresh_color_tag: str) -> tuple:
    """Core of the "не перезаписывать ручные изменения" fix. Compares what the sheet
    currently shows for this cell (old_entry, read fresh from the sheet right before this
    sync) against what the bot itself last wrote there (sync_map, our own persistent
    record in sheet_cell_sync). If the two match, nothing has happened to this cell by hand
    since our last write, so it's safe to apply the newly computed value/note/color. If they
    differ, someone (отдел контроля) changed the checkbox, note or highlight color directly
    in Google Sheets since then - their edit is kept exactly as-is and adopted as the new
    baseline, instead of being silently reverted on this sync.
    Returns (final_checked, final_note, final_color_tag)."""
    sheet_checked = old_entry.get("bool") if old_entry else None
    sheet_note = (old_entry.get("note") if old_entry else None) or ""
    sheet_color_tag = (old_entry.get("color_tag") if old_entry else None) or "none"

    sync_entry = sync_map.get(key)
    if sync_entry is None:
        # First time this cell has ever been tracked (new deploy, or a brand new cell) -
        # nothing to compare against yet, so there's nothing manual to protect.
        unchanged_since_last_sync = True
    else:
        unchanged_since_last_sync = (
            bool(sync_entry.get("checked")) == bool(sheet_checked)
            and (sync_entry.get("note") or "") == sheet_note
            and (sync_entry.get("color_tag") or "none") == sheet_color_tag
        )

    if unchanged_since_last_sync:
        return fresh_checked, fresh_note, fresh_color_tag
    return (sheet_checked if sheet_checked is not None else fresh_checked), (sheet_note or None), sheet_color_tag

def get_sheet_cell_sync_map(telegram_id: int, start_date: str, end_date: str) -> dict:
    """What the bot itself last wrote to each Сводка cell for this worker - the baseline
    used to tell a genuine data update apart from a manual edit made directly in the sheet
    (see _process_worker in sync_gsheets_task)."""
    conn = get_db()
    rows = conn.execute(
        "SELECT report_date, slot_label, checked, note, color_tag FROM sheet_cell_sync "
        "WHERE telegram_id = ? AND report_date >= ? AND report_date <= ?",
        (telegram_id, start_date, end_date)
    ).fetchall()
    conn.close()
    return {(r["report_date"], r["slot_label"]): dict(r) for r in rows}

def sync_gsheets_task() -> tuple[bool, str | None]:
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as e:
        return False, f"Отсутствуют необходимые библиотеки Python: {e}"

    spreadsheet_id = get_setting("google_spreadsheet_id")
    creds_str = get_setting("google_service_account")
    if not spreadsheet_id or not creds_str:
        return False, "Настройки Google Таблицы или сервисного аккаунта не заданы."

    # BUG FIX: conn/conn_sum are opened partway through the try block below and closed at a
    # specific point in the normal flow: if an exception fires anywhere between open and
    # that point (e.g. a bad query, or calculate_worker_stats raising outside its own
    # per-worker guard), the function fell straight to `except Exception` and the
    # connection was never closed. Pre-declared here so the finally block below can safely
    # close whichever of them actually got opened, on every exit path.
    conn = None
    conn_sum = None
    try:
        creds_dict = json.loads(creds_str)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        
        try:
            sheet = client.open_by_key(spreadsheet_id)
        except Exception as ex:
            email = creds_dict.get("client_email", "неизвестный email")
            return False, f"Не удалось открыть таблицу по ID. Убедитесь, что ID правильный и таблица открыта для редактирования (Поделиться) сервисному аккаунту: {email}. Ошибка: {ex}"
        
        # Helper to update worksheet compatibly with older/newer gspread versions
        def safe_update(ws, data):
            ws.clear()
            try:
                ws.update("A1", data)
            except Exception:
                try:
                    ws.update(data, "A1")
                except Exception:
                    ws.update(range_name="A1", values=data)

        # 1. Sync Workers to "Сотрудники"
        workers = get_all_workers()

        # Apply the admin-configured department display order (⚙️ Настройки -> "📊 Порядок
        # отделов в таблице"). get_all_workers() already orders by object_id (alphabetical),
        # sort_order, last_name, first_name - re-sort here by the custom department rank
        # instead, keeping the same secondary ordering within each department. This single
        # re-sort flows through to every tab below (Сотрудники, Аналитика, Сводка) since they
        # all just iterate this same `workers` list - no need to touch each one separately.
        # Every full sync already rebuilds every tab from scratch, so this alone is enough to
        # reorder rows that already exist - no separate "move existing rows" step is needed.
        dept_order, _ = get_departments_ordered()
        dept_rank = {name: i for i, name in enumerate(dept_order)}
        workers = sorted(
            workers,
            key=lambda w: (
                dept_rank.get(w["object_id"] or "Основной", len(dept_order)),
                w["sort_order"], w["last_name"], w["first_name"]
            )
        )

        headers_workers = [
            "Отдел", "ФИО сотрудника", "Должность", "Telegram ID",
            "ID Группы", "График", "Факт", "Активен", "Порядок сортировки"
        ]
        
        rows_workers = [headers_workers]
        for w in workers:
            rows_workers.append([
                str(w["object_id"] or "Основной"),
                f"{w['last_name']} {w['first_name']}",
                clean_position(w["position"]),
                str(w["telegram_id"]),
                str(w["group_id"]),
                str(w["schedule"]),
                "Да" if w["needs_daily_fact"] else "Нет",
                "Да" if w["is_active"] else "Нет",
                str(w["sort_order"])
            ])
            
        try:
            ws_workers = sheet.worksheet("Сотрудники")
        except gspread.exceptions.WorksheetNotFound:
            ws_workers = sheet.add_worksheet(title="Сотрудники", rows="1000", cols="20")
            
        safe_update(ws_workers, rows_workers)
        
        # 2. Sync Reports to "Отчеты" and 3. Sync Analytics to "Аналитика"
        # All DB reads happen first and the connection is closed BEFORE any Sheets API
        # calls — holding a SQLite connection open across slow network I/O is what let a
        # concurrent admin write (e.g. deleting a worker) block on the DB and freeze the bot.
        conn = get_db()
        reports = conn.execute("""
            SELECT r.*, w.last_name, w.first_name, w.position, w.object_id
            FROM reports r
            LEFT JOIN workers w ON r.telegram_id = w.telegram_id
            ORDER BY r.id DESC
        """).fetchall()

        headers_reports = [
            "ID отчета", "Telegram ID", "Отдел", "ФИО сотрудника", "Должность",
            "Дата отчета", "Тип отчета", "Время отчёта", "Время получения",
            "Оценка (ОК)", "Опоздание", "Замечания", "Действия", "Обработанный отчёт"
        ]

        rows_reports = [headers_reports]
        for r in reports:
            # Check if unregistered user has pending info
            fio = ""
            if r["last_name"] and r["first_name"]:
                fio = f"{r['last_name']} {r['first_name']}"
            else:
                pending = conn.execute("SELECT first_name, last_name FROM pending_unregistered_users WHERE telegram_id = ?", (r["telegram_id"],)).fetchone()
                if pending:
                    fio = f"{pending['last_name']} {pending['first_name']} (Временный)"
                else:
                    fio = "Неизвестный сотрудник"

            rows_reports.append([
                str(r["id"]),
                str(r["telegram_id"]),
                str(r["object_id"] or "Основной"),
                fio,
                clean_position(r["position"]),
                str(r["report_date"]),
                str(r["report_type"]),
                str(r["slot_time"] or ""),
                str(r["received_at"]),
                "Да" if r["is_ok"] else "Нет",
                "Да" if r["is_late"] else "Нет",
                str(r["format_comment"] or ""),
                str(r["required_action"] or ""),
                str(r["raw_text"] or "")
            ])

        now = now_local()
        end_date_str = now.strftime("%Y-%m-%d")

        # Earliest report date fallback
        earliest_row = conn.execute("SELECT MIN(report_date) as min_date FROM reports").fetchone()
        earliest_date_str = (earliest_row["min_date"] if earliest_row and earliest_row["min_date"] else end_date_str) or end_date_str

        # Start date for last 30 days
        start_30_days = (now - dt_module.timedelta(days=30)).strftime("%Y-%m-%d")
        if start_30_days < earliest_date_str:
            start_30_days = earliest_date_str

        headers_analytics = [
            "Отдел", "ФИО сотрудника", "Должность",
            "Всего отчетов (за 30 дн)", "Сдано (за 30 дн)", "Опозданий (за 30 дн)", "Замечаний (за 30 дн)", "Пропущено (за 30 дн)", "% Сдачи (за 30 дн)",
            "Всего отчетов (всё время)", "Сдано (всё время)", "Опозданий (всё время)", "Замечаний (всё время)", "Пропущено (всё время)", "% Сдачи (всё время)"
        ]

        rows_analytics = [headers_analytics]
        for w in workers:
            if not w["is_active"]:
                continue

            try:
                stats_30 = calculate_worker_stats(w, start_30_days, end_date_str, conn)
                stats_all = calculate_worker_stats(w, earliest_date_str, end_date_str, conn)

                rows_analytics.append([
                    str(w["object_id"] or "Основной"),
                    f"{w['last_name']} {w['first_name']}",
                    clean_position(w["position"]),
                    str(stats_30["expected"]),
                    str(stats_30["submitted"]),
                    str(stats_30["lates"]),
                    str(stats_30["remarks"]),
                    str(stats_30["missed"]),
                    f"{stats_30['percent']:.1f}%",
                    str(stats_all["expected"]),
                    str(stats_all["submitted"]),
                    str(stats_all["lates"]),
                    str(stats_all["remarks"]),
                    str(stats_all["missed"]),
                    f"{stats_all['percent']:.1f}%"
                ])
            except Exception as exc:
                logging.getLogger(__name__).warning(
                    f"[Аналитика] Пропущен сотрудник {w['telegram_id']} ({w['last_name']} {w['first_name']}) из-за ошибки: {exc}"
                )

        conn.close()

        try:
            ws_reports = sheet.worksheet("Отчеты")
        except gspread.exceptions.WorksheetNotFound:
            ws_reports = sheet.add_worksheet(title="Отчеты", rows="5000", cols="20")
        safe_update(ws_reports, rows_reports)

        try:
            ws_analytics = sheet.worksheet("Аналитика")
        except gspread.exceptions.WorksheetNotFound:
            ws_analytics = sheet.add_worksheet(title="Аналитика", rows="1000", cols="20")
        safe_update(ws_analytics, rows_analytics)

        # 4. Sync Summary checkbox grid to "Сводка"
        try:
            ws_summary = sheet.worksheet("Сводка")
        except gspread.exceptions.WorksheetNotFound:
            ws_summary = sheet.add_worksheet(title="Сводка", rows="2000", cols="60")

        GREY = {"red": 0.85, "green": 0.85, "blue": 0.85}
        YELLOW = {"red": 1.0, "green": 0.95, "blue": 0.6}
        PINK = {"red": 0.96, "green": 0.78, "blue": 0.78}
        DEPT_BG = {"red": 0.85, "green": 0.82, "blue": 0.93}
        LEGEND = [(PINK, "Не сдал"), (YELLOW, "Сдал, но есть замечание"), (GREY, "Выходной, отпуск")]

        FROZEN_ROWS = len(LEGEND) + 1  # legend rows + the "Сотрудник/Время/даты" header row
        FROZEN_COLS = 2

        # Read the currently published grid (values + notes) before wiping it, so manual edits
        # survive the rewrite: a checkbox someone ticked by hand is never flipped back to unchecked,
        # and a manually-added note is carried forward if we don't have a fresh one to replace it.
        old_map = {}
        try:
            meta = sheet.fetch_sheet_metadata({"ranges": ws_summary.title, "includeGridData": "true"})
            old_grid = []
            for s in meta.get("sheets", []):
                if s.get("properties", {}).get("sheetId") == ws_summary.id:
                    data = s.get("data", [])
                    if data:
                        old_grid = data[0].get("rowData", [])
                    break

            if len(old_grid) > FROZEN_ROWS - 1:
                header_vals = old_grid[FROZEN_ROWS - 1].get("values", [])
                old_date_cols = {
                    idx: cell.get("formattedValue")
                    for idx, cell in enumerate(header_vals)
                    if idx >= 2 and cell.get("formattedValue")
                }
                current_name = None
                for row in old_grid[FROZEN_ROWS:]:
                    vals = row.get("values", [])
                    if not vals:
                        continue
                    name_cell_text = vals[0].get("formattedValue") if len(vals) > 0 else None
                    slot_cell_text = vals[1].get("formattedValue") if len(vals) > 1 else None
                    if name_cell_text:
                        current_name = name_cell_text
                    if not slot_cell_text or current_name is None:
                        continue
                    for idx, date_label in old_date_cols.items():
                        if idx >= len(vals):
                            continue
                        cell = vals[idx]
                        cell_format = cell.get("effectiveFormat", {}) or {}
                        old_map[(current_name, slot_cell_text, date_label)] = {
                            "bool": cell.get("effectiveValue", {}).get("boolValue"),
                            "note": cell.get("note"),
                            "color_tag": _color_to_tag(cell_format.get("backgroundColor"))
                        }
        except Exception:
            old_map = {}

        # Clear merges/formatting/checkboxes/notes left over from a previous sync before rebuilding
        # the grid, and pin the freeze pane to a known state (mergeCells rejects ranges that straddle
        # the frozen/non-frozen column boundary, so every merge below must respect FROZEN_COLS).
        # Note: ws.clear() (called later via safe_update) only clears cell VALUES — it does not
        # touch merges, formatting, data validation (checkboxes) or notes, so each of those has to
        # be cleared explicitly here or it survives indefinitely, including for deleted workers'
        # old rows (this was the cause of stale notes/comments lingering after a worker was removed).
        sheet.batch_update({"requests": [
            {"unmergeCells": {"range": {"sheetId": ws_summary.id}}},
            {"repeatCell": {"range": {"sheetId": ws_summary.id}, "cell": {"userEnteredFormat": {}, "note": ""}, "fields": "userEnteredFormat,note"}},
            {"setDataValidation": {"range": {"sheetId": ws_summary.id}}},
            {"updateSheetProperties": {
                "properties": {"sheetId": ws_summary.id, "gridProperties": {"frozenRowCount": FROZEN_ROWS, "frozenColumnCount": FROZEN_COLS}},
                "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount"
            }}
        ]})

        conn_sum = get_db()
        today_date = now.date()
        month_start = today_date.replace(day=1)
        date_list = []
        d_iter = month_start
        while d_iter <= today_date:
            if d_iter.weekday() < 5:
                date_list.append(d_iter)
            d_iter += dt_module.timedelta(days=1)

        headers_summary = ["Сотрудник", "Время"] + [d.strftime("%d.%m") for d in date_list]

        merge_requests = []
        format_requests = []
        note_requests = []
        checkbox_ranges = []
        sheet_sync_updates = []  # (telegram_id, report_date, slot_label, checked, note, color_tag)

        # Legend rows, matching the reference sheet: a colored swatch in column A, label in column B
        rows_summary = []
        for color, label in LEGEND:
            legend_row_idx = len(rows_summary)
            rows_summary.append(["", label] + [""] * (len(headers_summary) - 2))
            format_requests.append({
                "repeatCell": {
                    "range": {"sheetId": ws_summary.id, "startRowIndex": legend_row_idx, "endRowIndex": legend_row_idx + 1,
                              "startColumnIndex": 0, "endColumnIndex": 1},
                    "cell": {"userEnteredFormat": {"backgroundColor": color}},
                    "fields": "userEnteredFormat(backgroundColor)"
                }
            })
        rows_summary.append(headers_summary)
        header_row_idx = len(rows_summary) - 1
        format_requests.append({
            "repeatCell": {
                "range": {"sheetId": ws_summary.id, "startRowIndex": header_row_idx, "endRowIndex": header_row_idx + 1,
                          "startColumnIndex": 0, "endColumnIndex": len(headers_summary)},
                "cell": {"userEnteredFormat": {"textFormat": {"bold": True, "foregroundColor": {"red": 0, "green": 0, "blue": 0}}}},
                "fields": "userEnteredFormat.textFormat(bold,foregroundColor)"
            }
        })

        def _process_worker(w):
            slots = SCHEDULES.get(w["schedule"], SCHEDULE_A)
            join_row = conn_sum.execute(
                "SELECT MIN(report_date) as d FROM reports WHERE telegram_id = ?", (w["telegram_id"],)
            ).fetchone()
            hire_date = today_date
            if join_row and join_row["d"]:
                try:
                    hire_date = datetime.strptime(join_row["d"], "%Y-%m-%d").date()
                except ValueError:
                    pass

            start_str = date_list[0].strftime("%Y-%m-%d")
            end_str = date_list[-1].strftime("%Y-%m-%d")
            status_rows = conn_sum.execute(
                "SELECT report_date, slot_time, received_at, is_ok, format_comment, required_action FROM reports "
                "WHERE telegram_id = ? AND report_type = 'status' AND report_date >= ? AND report_date <= ?",
                (w["telegram_id"], start_str, end_str)
            ).fetchall()
            status_map = {(r["report_date"], r["slot_time"]): r for r in status_rows}

            fact_rows = conn_sum.execute(
                "SELECT report_date, received_at, is_ok, format_comment, required_action FROM reports "
                "WHERE telegram_id = ? AND report_type = 'daily_fact' AND report_date >= ? AND report_date <= ?",
                (w["telegram_id"], start_str, end_str)
            ).fetchall()
            fact_map = {r["report_date"]: r for r in fact_rows}
            include_fact_row = bool(w["needs_daily_fact"])

            not_working_rows = conn_sum.execute(
                "SELECT report_date, format_comment FROM reports WHERE telegram_id = ? AND report_type = 'not_working' "
                "AND report_date >= ? AND report_date <= ?",
                (w["telegram_id"], start_str, end_str)
            ).fetchall()
            not_working_dates = {r["report_date"] for r in not_working_rows}
            not_working_reasons = {r["report_date"]: r["format_comment"] for r in not_working_rows if r["format_comment"]}

            missed_reasons = {
                (r["report_date"], r["slot_time"]): r["reason"] for r in conn_sum.execute(
                    "SELECT report_date, slot_time, reason FROM missed_status_reasons "
                    "WHERE telegram_id = ? AND report_date >= ? AND report_date <= ?",
                    (w["telegram_id"], start_str, end_str)
                ).fetchall()
            }

            sync_rows = conn_sum.execute(
                "SELECT report_date, slot_label, checked, note, color_tag FROM sheet_cell_sync "
                "WHERE telegram_id = ? AND report_date >= ? AND report_date <= ?",
                (w["telegram_id"], start_str, end_str)
            ).fetchall()
            sync_map = {(r["report_date"], r["slot_label"]): dict(r) for r in sync_rows}

            name_text = f"{w['last_name']} {w['first_name']} ({clean_position(w['position'])})"
            worker_start_row = len(rows_summary)
            first_tracked_col = None

            for slot in slots:
                row_cells = ["", slot]
                hour, minute = map(int, slot.split(":"))
                for col_idx, d in enumerate(date_list):
                    abs_col = 2 + col_idx
                    d_str = d.strftime("%Y-%m-%d")
                    date_label = headers_summary[abs_col]
                    old_entry = old_map.get((name_text, slot, date_label))
                    fresh_note = None

                    if d < hire_date:
                        row_cells.append("-")
                        continue

                    if first_tracked_col is None or abs_col < first_tracked_col:
                        first_tracked_col = abs_col

                    if d_str in not_working_dates:
                        cell_value = False
                    elif d == today_date and (now.hour * 60 + now.minute) <= hour * 60 + minute + STATUS_LATE_TOLERANCE_MIN:
                        cell_value = ""
                    else:
                        rep = status_map.get((d_str, slot))
                        if rep is not None:
                            fresh_checked = True
                            note_parts = []
                            fresh_color_tag = "none"
                            received_at = rep["received_at"] or ""
                            try:
                                rh, rm = int(received_at[0:2]), int(received_at[3:5])
                                diff_mins = (rh * 60 + rm) - (hour * 60 + minute)
                                # BUG FIX: an early submission (diff_mins negative) must never
                                # be labelled "Прислал поздно" - опоздание only means sent
                                # LATER than the acceptance window (+60 мин от слота).
                                if diff_mins > STATUS_LATE_TOLERANCE_MIN:
                                    note_parts.append(f"Прислал поздно, в {received_at[:5]}")
                                    fresh_color_tag = "yellow"
                            except (ValueError, IndexError):
                                pass
                            if not rep["is_ok"]:
                                issues = extract_issue_lines(rep["format_comment"])
                                note_parts.extend(issues if issues else ["Есть замечание"])
                            fresh_note = "\n".join(note_parts) if note_parts else None
                        else:
                            fresh_checked = False
                            fresh_note = missed_reasons.get((d_str, slot))
                            fresh_color_tag = "pink"

                        # MAIN FIX: don't blindly apply the freshly computed value - only do
                        # so if this cell is unchanged since the bot's own last sync. If
                        # отдел контроля already ticked/unticked the box, edited the note, or
                        # changed the highlight color directly in Sheets since then, that
                        # manual edit is kept exactly as-is here instead of being overwritten.
                        cell_value, fresh_note, color_tag = _resolve_manual_preserving_cell(
                            sync_map, old_entry, (d_str, slot), fresh_checked, fresh_note, fresh_color_tag
                        )
                        color_dict = _color_tag_to_dict(color_tag)
                        if color_dict:
                            format_requests.append({
                                "repeatCell": {
                                    "range": {"sheetId": ws_summary.id, "startRowIndex": len(rows_summary),
                                              "endRowIndex": len(rows_summary) + 1,
                                              "startColumnIndex": abs_col, "endColumnIndex": abs_col + 1},
                                    "cell": {"userEnteredFormat": {"backgroundColor": color_dict}},
                                    "fields": "userEnteredFormat(backgroundColor)"
                                }
                            })
                        sheet_sync_updates.append(
                            (w["telegram_id"], d_str, slot, int(bool(cell_value)), fresh_note, color_tag)
                        )

                    row_cells.append(cell_value)

                    if fresh_note:
                        note_requests.append({
                            "updateCells": {
                                "range": {"sheetId": ws_summary.id, "startRowIndex": len(rows_summary),
                                          "endRowIndex": len(rows_summary) + 1,
                                          "startColumnIndex": abs_col, "endColumnIndex": abs_col + 1},
                                "rows": [{"values": [{"note": fresh_note}]}],
                                "fields": "note"
                            }
                        })

                rows_summary.append(row_cells)

            # "Факт" gets its own row per worker, independent of the per-slot status rows
            # above — own checkboxes, own notes, sourced from report_type='daily_fact' instead
            # of 'status'. Only shown for workers who are actually expected to submit one.
            if include_fact_row:
                row_cells = ["", "Факт"]
                last_slot = slots[-1]
                last_hour, last_minute = map(int, last_slot.split(":"))
                for col_idx, d in enumerate(date_list):
                    abs_col = 2 + col_idx
                    d_str = d.strftime("%Y-%m-%d")
                    date_label = headers_summary[abs_col]
                    old_entry = old_map.get((name_text, "Факт", date_label))
                    fresh_note = None

                    if d < hire_date:
                        row_cells.append("-")
                        continue

                    if first_tracked_col is None or abs_col < first_tracked_col:
                        first_tracked_col = abs_col

                    if d_str in not_working_dates:
                        cell_value = False
                    elif d == today_date and (now.hour * 60 + now.minute) <= last_hour * 60 + last_minute + 60:
                        cell_value = ""
                    else:
                        rep = fact_map.get(d_str)
                        if rep is not None:
                            fresh_checked = True
                            note_parts = []
                            if not rep["is_ok"]:
                                issues = extract_issue_lines(rep["format_comment"])
                                note_parts.extend(issues if issues else ["Есть замечание"])
                            fresh_note = "\n".join(note_parts) if note_parts else None
                            fresh_color_tag = "none"
                        else:
                            fresh_checked = False
                            fresh_note = None
                            fresh_color_tag = "pink"

                        # Same manual-edit-preserving check as the status rows above.
                        cell_value, fresh_note, color_tag = _resolve_manual_preserving_cell(
                            sync_map, old_entry, (d_str, "Факт"), fresh_checked, fresh_note, fresh_color_tag
                        )
                        color_dict = _color_tag_to_dict(color_tag)
                        if color_dict:
                            format_requests.append({
                                "repeatCell": {
                                    "range": {"sheetId": ws_summary.id, "startRowIndex": len(rows_summary),
                                              "endRowIndex": len(rows_summary) + 1,
                                              "startColumnIndex": abs_col, "endColumnIndex": abs_col + 1},
                                    "cell": {"userEnteredFormat": {"backgroundColor": color_dict}},
                                    "fields": "userEnteredFormat(backgroundColor)"
                                }
                            })
                        sheet_sync_updates.append(
                            (w["telegram_id"], d_str, "Факт", int(bool(cell_value)), fresh_note, color_tag)
                        )

                    row_cells.append(cell_value)

                    if fresh_note:
                        note_requests.append({
                            "updateCells": {
                                "range": {"sheetId": ws_summary.id, "startRowIndex": len(rows_summary),
                                          "endRowIndex": len(rows_summary) + 1,
                                          "startColumnIndex": abs_col, "endColumnIndex": abs_col + 1},
                                "rows": [{"values": [{"note": fresh_note}]}],
                                "fields": "note"
                            }
                        })
                rows_summary.append(row_cells)

            row_count = len(slots) + (1 if include_fact_row else 0)

            rows_summary[worker_start_row][0] = name_text
            merge_requests.append({
                "mergeCells": {
                    "range": {"sheetId": ws_summary.id, "startRowIndex": worker_start_row,
                              "endRowIndex": worker_start_row + row_count,
                              "startColumnIndex": 0, "endColumnIndex": 1},
                    "mergeType": "MERGE_ALL"
                }
            })

            if first_tracked_col is not None:
                checkbox_ranges.append(
                    (worker_start_row, worker_start_row + row_count, first_tracked_col, len(headers_summary))
                )

            # Merge consecutive "not working" (vacation/sick leave) days into a single grey
            # block spanning all of the worker's slot rows, instead of repeating the same
            # grey cell on every row for every day off — one note explains the whole span.
            not_working_col_idx = sorted(
                idx for idx, d in enumerate(date_list) if d.strftime("%Y-%m-%d") in not_working_dates
            )
            ranges = []
            for idx in not_working_col_idx:
                if ranges and idx == ranges[-1][1] + 1:
                    ranges[-1] = (ranges[-1][0], idx)
                else:
                    ranges.append((idx, idx))

            for r_start, r_end in ranges:
                c0, c1 = 2 + r_start, 2 + r_end + 1
                reasons = []
                for idx in range(r_start, r_end + 1):
                    reason = not_working_reasons.get(date_list[idx].strftime("%Y-%m-%d"))
                    if reason and reason not in reasons:
                        reasons.append(reason)
                merge_requests.append({
                    "mergeCells": {
                        "range": {"sheetId": ws_summary.id, "startRowIndex": worker_start_row,
                                  "endRowIndex": worker_start_row + row_count,
                                  "startColumnIndex": c0, "endColumnIndex": c1},
                        "mergeType": "MERGE_ALL"
                    }
                })
                format_requests.append({
                    "repeatCell": {
                        "range": {"sheetId": ws_summary.id, "startRowIndex": worker_start_row,
                                  "endRowIndex": worker_start_row + row_count,
                                  "startColumnIndex": c0, "endColumnIndex": c1},
                        "cell": {"userEnteredFormat": {"backgroundColor": GREY}},
                        "fields": "userEnteredFormat(backgroundColor)"
                    }
                })
                note_requests.append({
                    "updateCells": {
                        "range": {"sheetId": ws_summary.id, "startRowIndex": worker_start_row,
                                  "endRowIndex": worker_start_row + 1,
                                  "startColumnIndex": c0, "endColumnIndex": c0 + 1},
                        "rows": [{"values": [{"note": "; ".join(reasons) if reasons else "Выходной/отгул"}]}],
                        "fields": "note"
                    }
                })

        for dept, dept_workers in itertools.groupby(
            (w for w in workers if w["is_active"]),
            key=lambda w: str(w["object_id"] or "Основной")
        ):
            dept_row_idx = len(rows_summary)
            rows_summary.append([dept] + [""] * (len(headers_summary) - 1))
            # Two separate merges (frozen columns 0-1, then the rest) — Sheets rejects a merge
            # that spans both frozen and non-frozen columns.
            merge_requests.append({
                "mergeCells": {
                    "range": {"sheetId": ws_summary.id, "startRowIndex": dept_row_idx, "endRowIndex": dept_row_idx + 1,
                              "startColumnIndex": 0, "endColumnIndex": FROZEN_COLS},
                    "mergeType": "MERGE_ALL"
                }
            })
            merge_requests.append({
                "mergeCells": {
                    "range": {"sheetId": ws_summary.id, "startRowIndex": dept_row_idx, "endRowIndex": dept_row_idx + 1,
                              "startColumnIndex": FROZEN_COLS, "endColumnIndex": len(headers_summary)},
                    "mergeType": "MERGE_ALL"
                }
            })
            format_requests.append({
                "repeatCell": {
                    "range": {"sheetId": ws_summary.id, "startRowIndex": dept_row_idx, "endRowIndex": dept_row_idx + 1,
                              "startColumnIndex": 0, "endColumnIndex": len(headers_summary)},
                    "cell": {"userEnteredFormat": {"backgroundColor": DEPT_BG, "textFormat": {"bold": True, "foregroundColor": {"red": 0, "green": 0, "blue": 0}}}},
                    "fields": "userEnteredFormat(backgroundColor,textFormat)"
                }
            })

            for w in dept_workers:
                try:
                    _process_worker(w)
                except Exception as exc:
                    logging.getLogger(__name__).warning(
                        f"[Сводка] Пропущен сотрудник {w['telegram_id']} ({w['last_name']} {w['first_name']}) из-за ошибки: {exc}"
                    )

        # Persist what was actually rendered for every cell as the new "last known" baseline,
        # so the next sync can tell a genuine data update apart from a manual sheet edit.
        if sheet_sync_updates:
            conn_sum.executemany(
                "INSERT INTO sheet_cell_sync (telegram_id, report_date, slot_label, checked, note, color_tag) "
                "VALUES (?, ?, ?, ?, ?, ?) "
                "ON CONFLICT(telegram_id, report_date, slot_label) DO UPDATE SET "
                "checked=excluded.checked, note=excluded.note, color_tag=excluded.color_tag",
                sheet_sync_updates
            )
        conn_sum.commit()
        conn_sum.close()

        grid_last_row = len(rows_summary)
        BLACK = {"red": 0, "green": 0, "blue": 0}
        border_requests = []
        if grid_last_row > header_row_idx:
            # Thin line between every day column
            border_requests.append({
                "updateBorders": {
                    "range": {"sheetId": ws_summary.id, "startRowIndex": header_row_idx, "endRowIndex": grid_last_row,
                              "startColumnIndex": 2, "endColumnIndex": len(headers_summary)},
                    "innerVertical": {"style": "SOLID", "width": 1, "color": BLACK}
                }
            })
            # Thick line at the end of each work week (after every Friday column)
            for col_idx, d in enumerate(date_list):
                abs_col = 2 + col_idx
                if d.weekday() == 4 and abs_col + 1 < len(headers_summary):
                    border_requests.append({
                        "repeatCell": {
                            "range": {"sheetId": ws_summary.id, "startRowIndex": header_row_idx, "endRowIndex": grid_last_row,
                                      "startColumnIndex": abs_col, "endColumnIndex": abs_col + 1},
                            "cell": {"userEnteredFormat": {"borders": {"right": {"style": "SOLID_THICK", "width": 2, "color": BLACK}}}},
                            "fields": "userEnteredFormat.borders.right"
                        }
                    })

        safe_update(ws_summary, rows_summary)

        checkbox_requests = [
            {
                "setDataValidation": {
                    "range": {"sheetId": ws_summary.id, "startRowIndex": r0, "endRowIndex": r1,
                              "startColumnIndex": c0, "endColumnIndex": c1},
                    "rule": {"condition": {"type": "BOOLEAN"}, "strict": True}
                }
            }
            for (r0, r1, c0, c1) in checkbox_ranges
        ]

        all_requests = merge_requests + format_requests + border_requests + checkbox_requests + note_requests
        if all_requests:
            for i in range(0, len(all_requests), 400):
                sheet.batch_update({"requests": all_requests[i:i + 400]})

        return True, None

    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"[GSheets Sync Error] {e}")
        return False, str(e)
    finally:
        # Safety net for the leak described above - a no-op if the normal flow already
        # closed these (closing an already-closed sqlite3 connection is harmless).
        for c in (conn, conn_sum):
            if c is not None:
                try:
                    c.close()
                except Exception:
                    pass

_gsheets_sync_lock = None
_gsheets_sync_state_lock = None
_gsheets_sync_pending = False

def async_sync_gsheets_background():
    # BUG FIX: this used to spawn a brand-new thread on every call with no coordination.
    # save_report()/delete_worker()/etc. all call this after every single report, so at
    # end-of-day when several workers submit within seconds of each other, this could fire
    # 5-10 overlapping threads, each doing its own full read of `reports` plus dozens of
    # Google Sheets API calls (clear + rebuild) against the SAME spreadsheet concurrently -
    # wasted API quota, risk of 429 rate-limit errors, and interleaved writes to the same
    # sheet from independent runs. Now at most one sync runs at a time; a request that
    # arrives while one is already running just sets a "run once more" flag instead of
    # starting a second thread, and the running sync re-executes once more before exiting
    # so it still picks up whatever triggered the extra request.
    import threading
    global _gsheets_sync_lock, _gsheets_sync_state_lock, _gsheets_sync_pending
    if _gsheets_sync_lock is None:
        _gsheets_sync_lock = threading.Lock()
        _gsheets_sync_state_lock = threading.Lock()

    with _gsheets_sync_state_lock:
        if _gsheets_sync_lock.locked():
            _gsheets_sync_pending = True
            return

    def _run():
        global _gsheets_sync_pending
        with _gsheets_sync_lock:
            while True:
                sync_gsheets_task()
                with _gsheets_sync_state_lock:
                    if _gsheets_sync_pending:
                        _gsheets_sync_pending = False
                        continue
                    break

    threading.Thread(target=_run, daemon=True).start()
